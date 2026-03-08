#!/usr/bin/env python3
"""
GPO Cross-Domain Consolidation Analyzer v2.3.2
December 13, 2025

Parses GPOZaurr HTML reports to analyze Group Policy across multiple Active Directory domains.
Supports FIVE analysis modes for different audiences and use cases.

NEW IN v2.3.2 - IMPACT MODE (CLI Only):
- GPO replacement impact analysis for pre-execution planning
- Per-GPO delta showing what changes when replacing operation GPO with Enterprise Standard equivalent
- Risk assessment (HIGH/MEDIUM/LOW) based on settings lost or changed
- Answers: "What happens when I flip the switch?"

IMPACT MODE OUTPUT (5 tabs):
- Impact Summary: Per-GPO risk assessment with explanation
- Settings Retained: Settings that stay the same (no action)
- Settings Lost: CRITICAL - settings that will be lost if not migrated
- Settings Changed: CRITICAL - settings with different values (review needed)
- Settings Added: New settings from Enterprise Standard (awareness)

MIGRATION MODE:
- Setting-level extraction engine (78 categories validated against 10 production domains)
- Cross-domain comparison: Classifies settings as MIGRATE/DON'T MIGRATE/REVIEW
- Compares operation domain against baseline.corp baseline
- 75,355+ settings extracted with 0 errors across validation

MIGRATION MODE OUTPUT (4 tabs):
- Migration Summary: Overall classification counts and recommendations
- GPO Summary: Per-GPO priority classification (P1-P4) with action items
- Settings Analysis: Each setting with classification and source GPO info
- Review Required: Settings needing manual review (conflicts)

EXECUTIVE MODE:
- Migration Dashboard: Bucket breakdown by operation with migration readiness metrics
- Risk Assessment: Operations ranked by complexity (Mixed/Unknown/Links-only counts)
- Operations Comparison: Cross-operation bucket distribution for consistency analysis
- Enhanced Project Status with bucket-based metrics

LINKS-ONLY DETECTION (v2.3.2):
- GPOs linked to operation OUs but without operation code in name
  are now captured in "Review Required" tab to prevent missing anything
- Dual matching: Primary (DisplayName) + Secondary (Links) ensures complete coverage
- MatchType column added to identify how each GPO was matched

OPERATION FILTERING (v2.2.6):
- DisplayName-based filtering matches manual Excel filter exactly
- Word boundary matching for short codes prevents false positives
- name_prefixes in LOCATION_MAPPING for actual GPO naming conventions

BUCKET-BASED ANALYSIS:
- Dynamic OU parsing: Extracts operation and bucket from DN path
- GPOs categorized by link location (Server/Workstation/User/Domain Root)
- Multi-link handling: GPOs linked to multiple bucket types flagged as "Mixed"

BUCKET TYPES:
- Server: GPOs linked under Server OUs (MIGRATE - HIGH priority)
- Workstation: GPOs linked under Computer/Workstation OUs (MIGRATE - HIGH priority)  
- User: GPOs linked under User OUs (REVIEW - optimization candidates)
- Domain Root: GPOs linked at domain level (DON'T MIGRATE)
- Domain Controller: GPOs linked under Domain Controllers OU (DON'T MIGRATE)
- Mixed: GPOs linked to multiple bucket types (MANUAL REVIEW)
- Unknown: Cannot determine bucket (MANUAL REVIEW)

================================================================================
THREE-MODE ARCHITECTURE
================================================================================

EXECUTIVE MODE (7 tabs)
  Purpose: Cross-domain consolidation planning and project tracking
  Audience: Leadership, project managers
  Use Case: Monthly status updates, high-level project oversight
  Output: Strategic summary with standardization opportunities

DOMAIN MODE (5 tabs - BUCKET-FOCUSED)
  Purpose: Migration readiness assessment organized by GPO bucket type
  Audience: Domain/operations teams during GPO engagements
  Use Case: Per-domain/operation migration planning with clear scope
  Output: Bucket-organized analysis (Server, Workstation, User, Review)
  
  Tab 1: Bucket Overview - Summary counts by bucket type
  Tab 2: Server GPOs - MIGRATE (HIGH priority)
  Tab 3: Workstation GPOs - MIGRATE (HIGH priority)
  Tab 4: User GPOs - REVIEW (optimization candidates)
  Tab 5: Review Required - Domain Root, DC, Mixed, Unknown, AND Links-only GPOs

FULL MODE (19 tabs)
  Purpose: Complete technical analysis with all details
  Audience: Technical analysts, deep-dive troubleshooting
  Use Case: Quarterly comprehensive reviews, detailed analysis
  Output: All data, all associations, all drill-downs

================================================================================
KEY FEATURES
================================================================================

- Three-mode architecture for different audiences
- ENT filtering with per-domain tracking
- Shared GPO detection (baseline.corp <-> baseline.corp)
- Operations breakout within shared forest (10 locations)
- Domain-specific migration readiness assessment
- Single-domain mode detection with appropriate callouts
- Migration-focused approach (defer consolidation to Year 3+)

================================================================================
USAGE
================================================================================

Executive Mode:
    python gpo_analyzer.py --mode executive \
        --html-folder "C:\\GPO_Reports" \
        --output "executive.xlsx"

Domain Mode (regular domain):
    python gpo_analyzer.py --mode domain \
        --domain "echo.corp.com" \
        --html-folder "C:\\GPO_Reports" \
        --output "echo.xlsx"

Domain Mode (single operation - auto-detect domain):
    python gpo_analyzer.py --mode domain \
        --operation "OPF" \
        --html-folder "C:\\GPO_Reports" \
        --output "foxtrot.xlsx"

Domain Mode (single operation - explicit domain):
    python gpo_analyzer.py --mode domain \
        --domain "baseline.corp" \
        --operation "OPF" \
        --html-folder "C:\\GPO_Reports" \
        --output "foxtrot.xlsx"

Full Mode:
    python gpo_analyzer.py --mode full \
        --html-folder "C:\\GPO_Reports" \
        --output "full_analysis.xlsx"

================================================================================
OUTPUT TABS BY MODE
================================================================================

EXECUTIVE MODE (7 tabs):
  1. Project Status - Narrative summary with operations broken out
  2. Standardization Candidates - Settings in 8+ domains
  3. Conflict Resolution - Settings in 3-7 domains
  4. Work Distribution - GPO counts by domain/operation
  5. Performance Issues - WMI filters, empty GPOs by operation
  6. Infrastructure Dependencies - UNC paths, IPs
  7. Consolidation Roadmap - High-level timeline (illustrative)

DOMAIN MODE (5 tabs):
  1. Domain Overview - Migration scope, ENT filtering summary
  2. GPO Inventory - Complete list of GPOs to migrate
  3. Shared Configuration - Reference only (not actionable during migration)
  4. Domain-Specific Settings - Unique configuration
  5. Migration Checklist - Readiness assessment

FULL MODE (19 tabs):
  1. Executive Summary (operations broken out)
  2. Shared Enterprise GPOs
  3. Operations by Location
  4. Enterprise Standard GPOs
  5. Non-Standard ENT Names
  6-8. Enterprise Candidates, Inconsistent Settings, Domain Outliers
  9. Performance Issues (operations broken out)
  10. Migration Gotchas
  11. Security & Permissions
  12. OU Link Analysis
  13. Migration Roadmap
  14. Consolidation Recommendations
  14a. Settings Breakdown by Entity
  15. Gap Analysis
  16. Settings Detail View (pivot-ready)
  17-19. Raw Data

================================================================================
ARCHITECTURE NOTES
================================================================================

Data Flow:
  1. Parse HTML reports from GPOZaurr (one per domain)
  2. Filter to active GPOs (Enabled=True AND Linked=True)
  3. Detect shared GPOs (baseline.corp <-> baseline.corp by GUID)
  4. Filter ENT GPOs by name, track per-domain
  5. Extract settings from domain-unique GPOs
  6. Build association maps (setting -> entities, setting -> GPOs)
  7. Generate mode-specific Excel report

ENT Filtering:
  - Detects GPOs with "ENT" in DisplayName
  - Filters from domain reports (inherited from enterprise standards)
  - Tracks count per domain for reporting
  - Works in both single-domain and multi-domain scenarios

Migration Strategy:
  - Domain mode focuses on migration readiness, not consolidation
  - Move domains AS-IS during migration (Years 0-3)
  - Defer consolidation to Year 3+ after all domains migrated
  - Tab 3 (Shared Configuration) is reference only during migration

Single vs Multi-Domain Mode:
  - Single domain: Typical during GPO engagements, limited overlap analysis
  - Multi-domain: Fuller analysis with meaningful shared configuration data
  - Script detects scenario and adjusts callouts accordingly

================================================================================
VERSION HISTORY
================================================================================

v2.1.0 (December 2, 2025) - Current
  - Fixed: Unicode arrow characters causing SyntaxError
  - Fixed: Tab 5 ENT count now accurate (per-domain tracking)
  - Added: Single-domain mode detection and callouts
  - Updated: All version strings consistent

v2.0.2 (December 2, 2025)
  - Refactored: Domain mode for migration focus (not consolidation)
  - Refactored: ENT filtering into dedicated function
  - Added: Per-domain Enterprise Standard tracking dictionary
  - Changed: Tab 2 from "Enterprise Compliance" to "GPO Inventory"
  - Changed: Tab 5 from "P1-P5 Action Plan" to "Migration Checklist"
  - Fixed: TypeError on integer SettingValue

v1.9.1 (November 26, 2025)
  - Enhanced drill-down with domain/operation associations
  - Settings Detail View for pivot analysis
  - Performance Issues broken out by operation

"""
import argparse
import json
import re
import logging
from pathlib import Path
from collections import defaultdict, Counter
from typing import Dict, List, Any, Set, Tuple, Optional
from dataclasses import dataclass, field
from datetime import datetime
import sys

try:
    from bs4 import BeautifulSoup
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    print(f"Error: Required library not found: {e}")
    print("\nPlease install required libraries:")
    print("pip install beautifulsoup4 pandas openpyxl lxml")
    sys.exit(1)


# =============================================================================
# MIGRATION MODE - SETTING EXTRACTION ENGINE
# =============================================================================
# Phase 1A: Extracts granular settings from GPOZaurr HTML reports
# 78 categories validated against 10 production domains (75,355 settings, 0 errors)
# Phase 1B: Cross-domain comparison engine for migrate/don't-migrate decisions

# All 78 categories from GPOZaurr "Group Policy Content" section
SETTING_CATEGORIES = [
    # Original 45 categories
    'AccountPolicies', 'Audit', 'AutoPlay', 'ControlPanel', 'ControlPanelPersonalization',
    'Desktop', 'DnsClient', 'DriveMapping', 'FileExplorer', 'FolderRedirection',
    'GoogleChrome', 'GroupPolicy', 'InternetExplorer', 'LAPS', 'LocalGroups',
    'Logon', 'MicrosoftOutlook2003', 'Policies', 'Printers', 'PrintersPolicies',
    'PublicKeyPoliciesCertificates', 'PublicKeyPoliciesAutoEnrollment',
    'PublicKeyPoliciesEFS', 'PublicKeyPoliciesRootCA', 'RegistrySetting',
    'RegistrySettings', 'RemoteDesktopServices', 'RestrictedGroups', 'Scripts',
    'SecurityOptions', 'SoftwareInstallation', 'SystemServicesNT', 'TaskScheduler',
    'UserRightsAssignment', 'WindowsFirewallConnectionSecurityRules',
    'WindowsFirewallConnectionSecurityAuthentication', 'WindowsFirewallProfiles',
    'WindowsFirewallRules', 'WindowsLogon', 'WindowsMediaPlayer', 'WindowsMessenger',
    'WindowsRemoteManagement', 'WindowsTimeService', 'WindowsUpdate', 'InternetExplorerZones',
    # Discovered in baseline.corp.html (30 additional)
    'Autologon', 'Biometrics', 'ControlPanelAddRemove', 'ControlPanelDisplay',
    'ControlPanelPrinters', 'ControlPanelPrograms', 'ControlPanelRegional',
    'CredentialsDelegation', 'EventLog', 'FolderRedirectionPolicy', 'FSLogix',
    'InternetCommunicationManagement', 'LocalUsers', 'MicrosoftEdge',
    'MicrosoftManagementConsole', 'MicrosoftOutlook2010', 'MicrosoftOutlook2013',
    'MicrosoftOutlook2016', 'NetMeeting', 'OneDrive', 'OnlineAssistance',
    'PublicKeyPoliciesEnrollmentPolicy', 'RemoteAssistance', 'RSSFeeds',
    'SystemServices', 'TaskSchedulerPolicies', 'WindowsDefender',
    'WindowsHelloForBusiness', 'WindowsInstaller', 'WindowsPowerShell',
    # Discovered in additional domain HTML (3 additional)
    'Bitlocker', 'EventForwarding', 'EventLogService',
]

# Category schemas - how to extract key/value from each category type
CATEGORY_SCHEMAS = {
    # TYPE 1: Flat GPO - settings are columns in GPO row
    'AccountPolicies': {
        'type': 'flat_gpo',
        'setting_columns': [
            'ClearTextPassword', 'LockoutBadCount', 'LockoutDuration',
            'MaximumPasswordAge', 'MinimumPasswordAge', 'MinimumPasswordLength',
            'PasswordComplexity', 'PasswordHistorySize', 'ResetLockoutCount',
            'MaxClockSkew', 'MaxRenewAge', 'MaxServiceAge', 'MaxTicketAge',
            'TicketValidateClient'
        ]
    },
    'Audit': {
        'type': 'flat_gpo',
        'setting_columns': [
            'AuditAccountLogon', 'AuditAccountManage', 'AuditDSAccess',
            'AuditLogonEvents', 'AuditObjectAccess', 'AuditPolicyChange',
            'AuditPrivilegeUse', 'AuditProcessTracking', 'AuditSystemEvents',
            'AuditAccountLockout', 'AuditApplicationGenerated',
            'AuditApplicationGroupManagement', 'AuditAuditPolicyChange',
            'AuditAuthenticationPolicyChange', 'AuditAuthorizationPolicyChange',
            'AuditCertificationServices', 'AuditComputerAccountManagement',
            'AuditCredentialValidation', 'AuditDetailedDirectoryServiceReplication',
            'AuditDetailedFileShare', 'AuditDirectoryServiceAccess',
            'AuditDirectoryServiceChanges', 'AuditDirectoryServiceReplication',
            'AuditDistributionGroupManagement', 'AuditDPAPIActivity',
            'AuditFileShare', 'AuditFileSystem', 'AuditFilteringPlatformConnection',
            'AuditFilteringPlatformPacketDrop', 'AuditFilteringPlatformPolicyChange',
            'AuditGroupMembership', 'AuditHandleManipulation', 'AuditIPsecDriver',
            'AuditIPsecExtendedMode', 'AuditIPsecMainMode', 'AuditIPsecQuickMode',
            'AuditKerberosAuthenticationService', 'AuditKerberosServiceTicketOperations',
            'AuditKernelObject', 'AuditLogoff', 'AuditLogon',
            'AuditMPSSVCRuleLevelPolicyChange', 'AuditNetworkPolicyServer',
            'AuditNonSensitivePrivilegeUse', 'AuditOtherAccountLogonEvents',
            'AuditOtherAccountManagementEvents', 'AuditOtherLogonLogoffEvents',
            'AuditOtherObjectAccessEvents', 'AuditOtherPolicyChangeEvents',
            'AuditOtherPrivilegeUseEvents', 'AuditOtherSystemEvents',
            'AuditPNPActivity', 'AuditProcessCreation', 'AuditProcessTermination',
            'AuditRegistry', 'AuditRemovableStorage', 'AuditRPCEvents', 'AuditSAM',
            'AuditSecurityGroupManagement', 'AuditSecurityStateChange',
            'AuditSecuritySystemExtension', 'AuditSensitivePrivilegeUse',
            'AuditSpecialLogon', 'AuditSystemIntegrity', 'AuditUserDeviceClaims',
            'AuditUserAccountManagement'
        ]
    },
    'Logon': {'type': 'flat_gpo', 'setting_columns': ['RunTheseProgramsAtUserLogon', 'RunTheseProgramsAtUserLogonItemsToRunAtLogon']},
    'Autologon': {'type': 'flat_gpo', 'setting_columns': ['AutoAdminLogon', 'DefaultDomainName', 'DefaultUserName', 'DefaultPassword']},
    'EventLog': {'type': 'flat_gpo', 'setting_columns': ['ApplicationAuditLogRetentionPeriod', 'ApplicationMaximumLogSize', 'ApplicationRestrictGuestAccess', 'ApplicationRetentionDays', 'SystemAuditLogRetentionPeriod', 'SystemMaximumLogSize', 'SystemRestrictGuestAccess', 'SystemRetentionDays', 'SecurityAuditLogRetentionPeriod', 'SecurityMaximumLogSize', 'SecurityRestrictGuestAccess', 'SecurityRetentionDays']},
    'CredentialsDelegation': {'type': 'flat_gpo', 'setting_columns': ['AllowDelegatingDefaultCredentials', 'AllowDelegatingDefaultCredentialsConcatenateOSDefaultsWithInputAbove', 'EncryptionOracleRemediation', 'EncryptionOracleRemediationProtectionLevel']},
    'FolderRedirectionPolicy': {'type': 'flat_gpo', 'setting_columns': ['DoNotAutomaticallyMakeSpecificRedirectedFoldersAvailableOffline']},
    'TaskSchedulerPolicies': {'type': 'flat_gpo', 'setting_columns': ['PreventTaskRunOrEnd', 'ProhibitNewTaskCreation', 'ProhibitBrowse', 'ProhibitDragAndDrop', 'ProhibitTaskDeletion', 'HidePropertyPages']},
    'WindowsPowerShell': {'type': 'flat_gpo', 'setting_columns': ['TurnOnScriptExecution', 'TurnOnScriptExecutionExecutionPolicy']},
    'WindowsInstaller': {'type': 'flat_gpo', 'setting_columns': ['AlwaysInstallWithElevatedPrivileges', 'PreventRemovableMediaSourceForAnyInstallation', 'ProhibitRemovalOfUpdates', 'ProhibitRollback']},
    'WindowsDefender': {'type': 'flat_gpo', 'setting_columns': ['ConfigureWindowsDefenderSmartscreen']},
    'WindowsHelloForBusiness': {'type': 'flat_gpo', 'setting_columns': ['UseWindowsHelloForBusiness']},
    'OneDrive': {'type': 'flat_gpo', 'setting_columns': ['SaveDocumentsToOnedriveByDefault']},
    
    # TYPE 2: Per-Setting with KeyName/KeyValue
    'SecurityOptions': {'type': 'per_setting', 'key_field': 'KeyName', 'key_display': 'KeyDisplayName', 'value_field': 'KeyValue', 'fallback_key': 'SystemAccessPolicyName'},
    'UserRightsAssignment': {'type': 'per_setting', 'key_field': 'UserRightsAssignment', 'key_display': 'UserRightsAssignmentDescription', 'value_field': 'Name', 'fallback_key': 'Name'},
    'RegistrySettings': {'type': 'per_setting', 'key_field': 'Key', 'key_display': 'Name', 'value_field': 'Value'},
    'RegistrySetting': {'type': 'per_setting', 'key_field': 'KeyPath', 'key_display': 'Name', 'value_field': 'Value'},
    
    # TYPE 3: Per-Item configurations
    'DriveMapping': {'type': 'per_item', 'key_fields': ['Letter', 'Path'], 'value_fields': ['Action', 'Label', 'Persistent']},
    'Scripts': {'type': 'per_item', 'key_fields': ['Type', 'Command'], 'value_fields': ['Parameters', 'Order']},
    'Printers': {'type': 'per_item', 'key_fields': ['Name', 'Path'], 'value_fields': ['Action', 'Default']},
    'LocalGroups': {'type': 'per_item', 'key_fields': ['GroupName'], 'value_fields': ['Members', 'Action']},
    'RestrictedGroups': {'type': 'per_item', 'key_fields': ['GroupName'], 'value_fields': ['Members', 'MembersOf']},
    'FolderRedirection': {'type': 'per_item', 'key_fields': ['FolderType'], 'value_fields': ['DestinationPath', 'GrantExclusiveRights']},
    'TaskScheduler': {'type': 'per_item', 'key_fields': ['Name'], 'value_fields': ['Action', 'Enabled', 'Command']},
    'WindowsFirewallRules': {'type': 'per_item', 'key_fields': ['Name', 'Direction'], 'value_fields': ['Action', 'Enabled', 'Protocol', 'LocalPort', 'RemoteAddress']},
    'SystemServicesNT': {'type': 'per_item', 'key_fields': ['ServiceName'], 'value_fields': ['ServiceStartupType', 'ServiceAction']},
    'LocalUsers': {'type': 'per_item', 'key_fields': ['UserName'], 'value_fields': ['Action', 'NewName', 'FullName', 'Description', 'AccountIsDisabled']},
    'SystemServices': {'type': 'per_item', 'key_fields': ['ServiceName'], 'value_fields': ['ServiceStartUpMode', 'SecurityAuditingPresent', 'SecurityPermissionsPresent']},
    
    # TYPE 4: Generic - extract all non-standard fields
    '_default': {'type': 'generic', 'extract_all': True}
}

# Add generic handling for all categories not explicitly defined
for cat in SETTING_CATEGORIES:
    if cat not in CATEGORY_SCHEMAS:
        CATEGORY_SCHEMAS[cat] = {'type': 'generic', 'extract_all': True}


@dataclass
class ExtractedSetting:
    """Represents a single extracted setting from a GPO."""
    category: str
    gpo_name: str
    gpo_guid: str
    gpo_type: str  # Computer or User
    domain_name: str
    setting_key: str  # Normalized key for comparison
    setting_name: str  # Human-readable name
    setting_value: Any  # The actual value
    raw_data: Dict = field(default_factory=dict)
    links: str = ""
    linked: bool = True
    links_count: int = 0


@dataclass
class MigrationClassification:
    """Classification result for a setting comparison."""
    setting_key: str
    setting_name: str
    category: str
    classification: str  # MIGRATE, DONT_MIGRATE, REVIEW
    reason: str
    operation_value: Any
    enterprise_value: Any = None
    operation_gpo: str = ""
    enterprise_gpo: str = ""  # Only ENT GPOs - for Settings_Analysis
    shared_overlap_gpo: str = ""  # Operation GPOs in baseline.corp - for Column F


class SettingExtractionEngine:
    """
    Extracts granular settings from GPOZaurr HTML reports.
    Supports 78 categories validated against 10 production domains.
    
    Usage:
        engine = SettingExtractionEngine()
        settings = engine.extract_from_html_content(html_content, domain_name)
    """
    
    def __init__(self, verbose: bool = False):
        self.verbose = verbose
        self._html_content = None
    
    def log(self, message: str):
        """Print message if verbose mode is enabled."""
        if self.verbose:
            logging.debug(f"[ExtractionEngine] {message}")
    
    def extract_from_html_content(self, html_content: str, domain_name: str) -> List[ExtractedSetting]:
        """Extract all settings from HTML content."""
        self._html_content = html_content
        all_settings = []
        
        for category in SETTING_CATEGORIES:
            try:
                settings, error = self._extract_category(category, domain_name)
                if settings:
                    all_settings.extend(settings)
                    self.log(f"{category}: {len(settings)} settings")
            except Exception as e:
                self.log(f"{category}: Exception - {e}")
        
        self.log(f"Total settings extracted: {len(all_settings)}")
        return all_settings
    
    def _extract_json_array(self, category: str) -> Tuple[Optional[List[Dict]], Optional[str]]:
        """Extract the JSON data array for a category tab."""
        # Find the tab ID for this category
        tab_pattern = rf'<div id="(Tab-[a-z0-9]+)"[^>]*><span>{re.escape(category)}</span>'
        tab_match = re.search(tab_pattern, self._html_content)
        
        if not tab_match:
            return None, f"Tab not found"
        
        tab_id = tab_match.group(1)
        content_id = f"{tab_id}-Content"
        
        # Find the content section
        content_start = self._html_content.find(f'id="{content_id}"')
        if content_start == -1:
            return None, f"Content section not found"
        
        # Search for "data": [ within the next 10000 chars
        search_region = self._html_content[content_start:content_start + 10000]
        data_match = re.search(r'"data":\s*\[', search_region)
        
        if not data_match:
            return None, f"No data array found"
        
        # Find matching closing bracket
        start_pos = content_start + data_match.end() - 1
        bracket_count = 0
        end_pos = start_pos
        remaining = self._html_content[start_pos:]
        
        for i, char in enumerate(remaining):
            if char == '[':
                bracket_count += 1
            elif char == ']':
                bracket_count -= 1
                if bracket_count == 0:
                    end_pos = start_pos + i + 1
                    break
            if remaining[i:i+9] == '</script>':
                return None, f"Hit script boundary"
        
        if bracket_count != 0:
            return None, f"Unbalanced brackets"
        
        json_str = self._html_content[start_pos:end_pos]
        
        try:
            data = json.loads(json_str)
            return data, None
        except json.JSONDecodeError as e:
            return None, f"JSON parse error: {e}"
    
    def _extract_category(self, category: str, domain_name: str) -> Tuple[List[ExtractedSetting], Optional[str]]:
        """Extract settings from a specific category."""
        raw_data, error = self._extract_json_array(category)
        
        if error:
            return [], error
        
        if not raw_data:
            return [], "Empty data"
        
        schema = CATEGORY_SCHEMAS.get(category, CATEGORY_SCHEMAS['_default'])
        
        if schema['type'] == 'flat_gpo':
            return self._extract_flat_gpo(category, raw_data, schema, domain_name)
        elif schema['type'] == 'per_setting':
            return self._extract_per_setting(category, raw_data, schema, domain_name)
        elif schema['type'] == 'per_item':
            return self._extract_per_item(category, raw_data, schema, domain_name)
        else:
            return self._extract_generic(category, raw_data, domain_name)
    
    def _extract_flat_gpo(self, category: str, data: List[Dict], schema: Dict, domain_name: str) -> Tuple[List[ExtractedSetting], Optional[str]]:
        """Extract from flat GPO format where settings are columns."""
        settings = []
        setting_columns = schema.get('setting_columns', [])
        
        for row in data:
            gpo_name = row.get('DisplayName', 'Unknown')
            gpo_guid = row.get('GUID', '')
            gpo_type = row.get('GpoType', 'Unknown')
            links = row.get('Links', '')
            linked = row.get('Linked', 'True') == 'True'
            links_count = int(row.get('LinksCount', 0))
            
            for col in setting_columns:
                value = row.get(col)
                if value is not None and value != '' and value != 'Not Set':
                    settings.append(ExtractedSetting(
                        category=category,
                        gpo_name=gpo_name,
                        gpo_guid=gpo_guid,
                        gpo_type=gpo_type,
                        domain_name=domain_name,
                        setting_key=f"{category}:{col}",
                        setting_name=col,
                        setting_value=value,
                        raw_data=row,
                        links=links,
                        linked=linked,
                        links_count=links_count
                    ))
        
        return settings, None
    
    def _extract_per_setting(self, category: str, data: List[Dict], schema: Dict, domain_name: str) -> Tuple[List[ExtractedSetting], Optional[str]]:
        """Extract from per-setting format where each row is one setting."""
        settings = []
        key_field = schema.get('key_field', 'KeyName')
        key_display = schema.get('key_display', key_field)
        value_field = schema.get('value_field', 'KeyValue')
        fallback_key = schema.get('fallback_key')
        
        for row in data:
            gpo_name = row.get('DisplayName', 'Unknown')
            gpo_guid = row.get('GUID', '')
            gpo_type = row.get('GpoType', 'Unknown')
            links = row.get('Links', '')
            linked = row.get('Linked', 'True') == 'True'
            links_count = int(row.get('LinksCount', 0))
            
            key = row.get(key_field, '')
            if not key and fallback_key:
                key = row.get(fallback_key, '')
            
            if not key:
                continue
            
            display = row.get(key_display, key) or key
            value = row.get(value_field, '')
            
            settings.append(ExtractedSetting(
                category=category,
                gpo_name=gpo_name,
                gpo_guid=gpo_guid,
                gpo_type=gpo_type,
                domain_name=domain_name,
                setting_key=f"{category}:{key}",
                setting_name=display,
                setting_value=value,
                raw_data=row,
                links=links,
                linked=linked,
                links_count=links_count
            ))
        
        return settings, None
    
    def _extract_per_item(self, category: str, data: List[Dict], schema: Dict, domain_name: str) -> Tuple[List[ExtractedSetting], Optional[str]]:
        """Extract from per-item format (drives, scripts, printers, etc.)."""
        settings = []
        key_fields = schema.get('key_fields', [])
        value_fields = schema.get('value_fields', [])
        
        for row in data:
            gpo_name = row.get('DisplayName', 'Unknown')
            gpo_guid = row.get('GUID', '')
            gpo_type = row.get('GpoType', 'Unknown')
            links = row.get('Links', '')
            linked = row.get('Linked', 'True') == 'True'
            links_count = int(row.get('LinksCount', 0))
            
            # Build key from key fields
            key_parts = []
            for kf in key_fields:
                val = row.get(kf, '')
                if val:
                    key_parts.append(f"{kf}={val}")
            
            if not key_parts:
                continue
            
            key = "|".join(key_parts)
            
            # Build value from value fields
            value_parts = {}
            for vf in value_fields:
                val = row.get(vf)
                if val is not None and val != '':
                    value_parts[vf] = val
            
            settings.append(ExtractedSetting(
                category=category,
                gpo_name=gpo_name,
                gpo_guid=gpo_guid,
                gpo_type=gpo_type,
                domain_name=domain_name,
                setting_key=f"{category}:{key}",
                setting_name=key,
                setting_value=json.dumps(value_parts) if value_parts else "",
                raw_data=row,
                links=links,
                linked=linked,
                links_count=links_count
            ))
        
        return settings, None
    
    def _extract_generic(self, category: str, data: List[Dict], domain_name: str) -> Tuple[List[ExtractedSetting], Optional[str]]:
        """Generic extraction for undefined categories."""
        settings = []
        skip_fields = {'DisplayName', 'DomainName', 'GUID', 'GpoType', 'Filters', 'Linked', 'LinksCount', 'Links', 'CreatedTime', 'ModifiedTime', 'ReadTime', 'SecurityDescriptor', 'FilterDataAvailable'}
        
        for row in data:
            gpo_name = row.get('DisplayName', 'Unknown')
            gpo_guid = row.get('GUID', '')
            gpo_type = row.get('GpoType', 'Unknown')
            links = row.get('Links', '')
            linked = row.get('Linked', 'True') == 'True'
            links_count = int(row.get('LinksCount', 0))
            
            for field, value in row.items():
                if field in skip_fields:
                    continue
                if value is None or value == '' or value == 'Not Set':
                    continue
                
                settings.append(ExtractedSetting(
                    category=category,
                    gpo_name=gpo_name,
                    gpo_guid=gpo_guid,
                    gpo_type=gpo_type,
                    domain_name=domain_name,
                    setting_key=f"{category}:{field}",
                    setting_name=field,
                    setting_value=value,
                    raw_data=row,
                    links=links,
                    linked=linked,
                    links_count=links_count
                ))
        
        return settings, None


def is_enterprise_standard_gpo(gpo_name: str) -> bool:
    """
    Check if GPO name indicates an Enterprise Standard policy.

    Enterprise Standard GPOs follow patterns like:
    - "ENT - Policy Name"
    - "ENT-PolicyName"
    - "ENT_PolicyName"
    - Names containing "ENT" as a word or prefix

    This is the SINGLE SOURCE OF TRUTH for Enterprise Standard detection.
    Used by MigrationComparisonEngine and report generation.
    """
    if not gpo_name:
        return False
    # Match ENT at word boundary OR followed by underscore/hyphen/space
    return bool(re.search(r'(?:^|[\s\-_])ENT(?:[\s\-_]|$)', str(gpo_name), re.IGNORECASE))


class MigrationComparisonEngine:
    """
    Compares settings between operation domain and baseline.corp baseline.
    Classifies each setting as MIGRATE, DONT_MIGRATE, or REVIEW.
    Also identifies GAINED settings (Enterprise Standard settings operation will receive after migration).
    """
    
    def __init__(self):
        self.operation_settings: Dict[str, ExtractedSetting] = {}
        self.enterprise_settings: Dict[str, ExtractedSetting] = {}
        self.classifications: List[MigrationClassification] = []
    
    def load_operation_settings(self, settings: List[ExtractedSetting]):
        """Load settings from operation domain."""
        for s in settings:
            # Use setting_key as unique identifier
            # If multiple GPOs have same setting, keep first (or could merge)
            if s.setting_key not in self.operation_settings:
                self.operation_settings[s.setting_key] = s
    
    def load_enterprise_settings(self, settings: List[ExtractedSetting]):
        """
        Load settings from baseline.corp baseline.
        
        Loads ALL settings (ENT and operation GPOs). The filtering for 
        classification purposes happens in compare_and_classify().
        """
        for s in settings:
            if s.setting_key not in self.enterprise_settings:
                self.enterprise_settings[s.setting_key] = s
    
    def compare_and_classify(self) -> List[MigrationClassification]:
        """
        Compare operation settings against enterprise baseline.
        
        Classification Logic (only ENT GPOs affect classification):
        - DONT_MIGRATE: Setting exists in Enterprise Standard GPO with matching value
        - MIGRATE: Setting unique to operation domain OR only exists in operation GPOs
        - REVIEW: Setting exists in Enterprise Standard GPO but values differ (conflict)
        
        Operation GPOs in baseline.corp are tracked but don't affect classification.
        """
        self.classifications = []
        
        # Check each operation setting against enterprise
        for key, op_setting in self.operation_settings.items():
            if key in self.enterprise_settings:
                ent_setting = self.enterprise_settings[key]
                
                # Check if the matching enterprise GPO is an Enterprise Standard
                if is_enterprise_standard_gpo(ent_setting.gpo_name):
                    # Enterprise Standard GPO - affects classification
                    if str(op_setting.setting_value) == str(ent_setting.setting_value):
                        # Matching - don't migrate
                        self.classifications.append(MigrationClassification(
                            setting_key=key,
                            setting_name=op_setting.setting_name,
                            category=op_setting.category,
                            classification='DONT_MIGRATE',
                            reason='Setting exists in Enterprise Standard baseline with matching value',
                            operation_value=op_setting.setting_value,
                            enterprise_value=ent_setting.setting_value,
                            operation_gpo=op_setting.gpo_name,
                            enterprise_gpo=ent_setting.gpo_name
                        ))
                    else:
                        # Conflict with Enterprise Standard - review
                        self.classifications.append(MigrationClassification(
                            setting_key=key,
                            setting_name=op_setting.setting_name,
                            category=op_setting.category,
                            classification='REVIEW',
                            reason='Setting conflicts with Enterprise Standard baseline (different values)',
                            operation_value=op_setting.setting_value,
                            enterprise_value=ent_setting.setting_value,
                            operation_gpo=op_setting.gpo_name,
                            enterprise_gpo=ent_setting.gpo_name
                        ))
                else:
                    # Operation GPO in baseline.corp - does NOT affect classification
                    # Mark as MIGRATE, store in shared_overlap_gpo for Column F (not enterprise_gpo)
                    # Note: Do NOT populate enterprise_value - would be confusing without enterprise_gpo
                    self.classifications.append(MigrationClassification(
                        setting_key=key,
                        setting_name=op_setting.setting_name,
                        category=op_setting.category,
                        classification='MIGRATE',
                        reason='Setting unique to operation domain',
                        operation_value=op_setting.setting_value,
                        operation_gpo=op_setting.gpo_name,
                        shared_overlap_gpo=ent_setting.gpo_name  # For Column F only
                    ))
            else:
                # Unique to operation - migrate
                self.classifications.append(MigrationClassification(
                    setting_key=key,
                    setting_name=op_setting.setting_name,
                    category=op_setting.category,
                    classification='MIGRATE',
                    reason='Setting unique to operation domain',
                    operation_value=op_setting.setting_value,
                    operation_gpo=op_setting.gpo_name
                ))
        
        return self.classifications
    
    def get_summary(self) -> Dict[str, int]:
        """Get classification summary counts."""
        summary = {'MIGRATE': 0, 'DONT_MIGRATE': 0, 'REVIEW': 0}
        for c in self.classifications:
            summary[c.classification] = summary.get(c.classification, 0) + 1
        return summary
    
    def to_dataframe(self) -> pd.DataFrame:
        """Convert classifications to DataFrame for Excel export."""
        records = []
        for c in self.classifications:
            records.append({
                'Classification': c.classification,
                'Category': c.category,
                'Setting_Key': c.setting_key,
                'Setting_Name': c.setting_name,
                'Operation_Value': str(c.operation_value),
                'Enterprise_Value': str(c.enterprise_value) if c.enterprise_value else '',
                'Operation_GPO': c.operation_gpo,
                'Enterprise_GPO': c.enterprise_gpo,  # Only ENT GPOs
                'Shared_Forest_Overlap_GPO': c.shared_overlap_gpo,  # Operation GPOs in baseline.corp
                'Reason': c.reason
            })
        return pd.DataFrame(records)


# ============================================================================
# IMPACT MODE - GPO Replacement Impact Analysis
# ============================================================================

@dataclass
class GPOImpactAnalysis:
    """
    Impact analysis for replacing an operation GPO with its Enterprise Standard equivalent.
    
    Answers: "What happens when I decommission this operation GPO and rely on Enterprise Standard?"
    """
    operation_gpo: str
    ent_gpo: str  # "No Enterprise Standard equivalent" if no overlap
    retained_count: int = 0   # Same key, same value (no change)
    lost_count: int = 0       # In operation, not in Enterprise Standard (CRITICAL - must migrate)
    changed_count: int = 0    # Same key, different value (CRITICAL - review needed)
    added_count: int = 0      # In Enterprise Standard, not in operation (awareness)
    risk_level: str = ""      # HIGH, MEDIUM, LOW
    risk_notes: str = ""      # Human-readable explanation
    
    # Detail lists for drill-down tabs
    retained: List[ExtractedSetting] = field(default_factory=list)
    lost: List[ExtractedSetting] = field(default_factory=list)
    changed: List[Tuple[ExtractedSetting, ExtractedSetting]] = field(default_factory=list)  # (operation, eit)
    added: List[ExtractedSetting] = field(default_factory=list)


class ImpactAnalysisEngine:
    """
    Analyzes impact of replacing operation GPOs with Enterprise Standard equivalents.
    
    Per-GPO pair analysis showing what changes when you "flip the switch":
    - RETAINED: Settings that stay the same
    - LOST: Settings that will be lost (must migrate elsewhere)
    - CHANGED: Settings with different values (need review)
    - ADDED: New settings from Enterprise Standard (awareness)
    
    Usage:
        engine = ImpactAnalysisEngine()
        engine.load_operation_settings(op_settings)
        engine.load_ent_settings(ent_settings)
        engine.detect_overlaps()
        analyses = engine.analyze_all()
    """
    
    def __init__(self):
        # Settings grouped by GPO name
        self.operation_gpos: Dict[str, List[ExtractedSetting]] = {}
        self.ent_gpos: Dict[str, List[ExtractedSetting]] = {}
        # Detected overlaps: operation_gpo -> ent_gpo with most overlap
        self.overlaps: Dict[str, str] = {}
        # Analysis results
        self.analyses: List[GPOImpactAnalysis] = []
    
    def load_operation_settings(self, settings: List[ExtractedSetting]):
        """Group operation settings by GPO name."""
        self.operation_gpos = {}
        for s in settings:
            if s.gpo_name not in self.operation_gpos:
                self.operation_gpos[s.gpo_name] = []
            self.operation_gpos[s.gpo_name].append(s)
    
    def load_ent_settings(self, settings: List[ExtractedSetting]):
        """Load and group Enterprise Standard settings by GPO name."""
        self.ent_gpos = {}
        for s in settings:
            if is_enterprise_standard_gpo(s.gpo_name):
                if s.gpo_name not in self.ent_gpos:
                    self.ent_gpos[s.gpo_name] = []
                self.ent_gpos[s.gpo_name].append(s)
    
    def detect_overlaps(self):
        """
        Detect which operation GPOs overlap with which ENT GPOs.
        Overlap = shared setting_keys.
        Picks the Enterprise Standard GPO with the MOST overlap (strongest relationship).
        """
        self.overlaps = {}
        
        for op_gpo, op_settings in self.operation_gpos.items():
            op_keys = {s.setting_key for s in op_settings}
            
            best_ent_gpo = None
            best_overlap_count = 0
            
            for ent_gpo, ent_settings in self.ent_gpos.items():
                ent_keys = {s.setting_key for s in ent_settings}
                overlap_count = len(op_keys & ent_keys)
                
                if overlap_count > best_overlap_count:
                    best_overlap_count = overlap_count
                    best_ent_gpo = ent_gpo
            
            self.overlaps[op_gpo] = best_ent_gpo  # None if no overlap
    
    def _calculate_risk(self, lost: int, changed: int) -> Tuple[str, str]:
        """
        Calculate risk level and generate human-readable notes.
        
        Risk Logic:
        - HIGH: Lost > 5 OR Changed > 3 — Significant settings will be lost or changed
        - MEDIUM: Lost > 0 OR Changed > 0 — Some settings require action
        - LOW: Lost = 0 AND Changed = 0 — Safe replacement
        """
        notes = []
        
        if lost > 0:
            notes.append(f"{lost} setting{'s' if lost != 1 else ''} will be lost if not migrated")
        if changed > 0:
            notes.append(f"{changed} setting{'s' if changed != 1 else ''} have different values (review needed)")
        
        if lost > 5 or changed > 3:
            level = "HIGH"
        elif lost > 0 or changed > 0:
            level = "MEDIUM"
        else:
            level = "LOW"
            notes.append("Safe replacement - no settings lost or changed")
        
        return level, "; ".join(notes)
    
    def analyze_gpo_pair(self, op_gpo: str, ent_gpo: Optional[str]) -> GPOImpactAnalysis:
        """
        Analyze impact for a single operation GPO -> Enterprise Standard GPO replacement.
        """
        op_settings = self.operation_gpos.get(op_gpo, [])
        
        # No Enterprise Standard equivalent - all settings are "lost" if GPO is decommissioned
        if not ent_gpo:
            risk_level, risk_notes = self._calculate_risk(len(op_settings), 0)
            return GPOImpactAnalysis(
                operation_gpo=op_gpo,
                ent_gpo="No Enterprise Standard equivalent",
                lost_count=len(op_settings),
                lost=op_settings,
                risk_level=risk_level,
                risk_notes=f"No Enterprise Standard equivalent - all {len(op_settings)} settings would be lost"
            )
        
        ent_settings = self.ent_gpos.get(ent_gpo, [])
        
        # Build lookup dicts
        op_by_key = {s.setting_key: s for s in op_settings}
        ent_by_key = {s.setting_key: s for s in ent_settings}
        
        op_keys = set(op_by_key.keys())
        ent_keys = set(ent_by_key.keys())
        
        retained = []
        lost = []
        changed = []
        added = []
        
        # Settings in operation
        for key in op_keys:
            op_setting = op_by_key[key]
            if key in ent_keys:
                ent_setting = ent_by_key[key]
                if str(op_setting.setting_value) == str(ent_setting.setting_value):
                    retained.append(op_setting)
                else:
                    changed.append((op_setting, ent_setting))
            else:
                lost.append(op_setting)
        
        # Settings only in Enterprise Standard
        for key in ent_keys - op_keys:
            added.append(ent_by_key[key])
        
        risk_level, risk_notes = self._calculate_risk(len(lost), len(changed))
        
        return GPOImpactAnalysis(
            operation_gpo=op_gpo,
            ent_gpo=ent_gpo,
            retained_count=len(retained),
            lost_count=len(lost),
            changed_count=len(changed),
            added_count=len(added),
            risk_level=risk_level,
            risk_notes=risk_notes,
            retained=retained,
            lost=lost,
            changed=changed,
            added=added
        )
    
    def analyze_all(self) -> List[GPOImpactAnalysis]:
        """
        Analyze all operation GPOs against their Enterprise Standard overlaps.
        Must call detect_overlaps() first.
        """
        self.analyses = []
        
        for op_gpo in sorted(self.operation_gpos.keys()):
            ent_gpo = self.overlaps.get(op_gpo)
            analysis = self.analyze_gpo_pair(op_gpo, ent_gpo)
            self.analyses.append(analysis)
        
        return self.analyses
    
    def get_summary_dataframe(self) -> 'pd.DataFrame':
        """Generate Impact Summary tab dataframe."""
        import pandas as pd
        
        records = []
        for a in self.analyses:
            records.append({
                'Operation_GPO': a.operation_gpo,
                'ENT_Replacement': a.ent_gpo,
                'Retained': a.retained_count,
                'Lost': a.lost_count,
                'Changed': a.changed_count,
                'Added': a.added_count,
                'Risk_Level': a.risk_level,
                'Risk_Notes': a.risk_notes
            })
        
        df = pd.DataFrame(records)
        # Sort by risk level (HIGH first), then by lost count
        risk_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2}
        df['_risk_sort'] = df['Risk_Level'].map(risk_order)
        df = df.sort_values(['_risk_sort', 'Lost'], ascending=[True, False])
        df = df.drop('_risk_sort', axis=1)
        
        return df
    
    def get_retained_dataframe(self) -> 'pd.DataFrame':
        """Generate Settings Retained tab dataframe."""
        import pandas as pd
        
        records = []
        for a in self.analyses:
            for s in a.retained:
                records.append({
                    'Operation_GPO': a.operation_gpo,
                    'ENT_GPO': a.ent_gpo,
                    'Category': s.category,
                    'Setting_Key': s.setting_key,
                    'Setting_Name': s.setting_name,
                    'Value': str(s.setting_value)
                })
        
        return pd.DataFrame(records)
    
    def get_lost_dataframe(self) -> 'pd.DataFrame':
        """Generate Settings Lost tab dataframe (CRITICAL)."""
        import pandas as pd
        
        records = []
        for a in self.analyses:
            for s in a.lost:
                records.append({
                    'Operation_GPO': a.operation_gpo,
                    'ENT_GPO': a.ent_gpo,
                    'Category': s.category,
                    'Setting_Key': s.setting_key,
                    'Setting_Name': s.setting_name,
                    'Current_Value': str(s.setting_value),
                    'Action': 'MIGRATE - Setting will be lost if not moved elsewhere'
                })
        
        return pd.DataFrame(records)
    
    def get_changed_dataframe(self) -> 'pd.DataFrame':
        """Generate Settings Changed tab dataframe (CRITICAL)."""
        import pandas as pd
        
        records = []
        for a in self.analyses:
            for op_setting, ent_setting in a.changed:
                records.append({
                    'Operation_GPO': a.operation_gpo,
                    'ENT_GPO': a.ent_gpo,
                    'Category': op_setting.category,
                    'Setting_Key': op_setting.setting_key,
                    'Setting_Name': op_setting.setting_name,
                    'Current_Value': str(op_setting.setting_value),
                    'New_Value': str(ent_setting.setting_value),
                    'Action': 'REVIEW - Value will change after replacement'
                })
        
        return pd.DataFrame(records)
    
    def get_added_dataframe(self) -> 'pd.DataFrame':
        """Generate Settings Added tab dataframe (AWARENESS)."""
        import pandas as pd
        
        records = []
        for a in self.analyses:
            for s in a.added:
                records.append({
                    'Operation_GPO': a.operation_gpo,
                    'ENT_GPO': a.ent_gpo,
                    'Category': s.category,
                    'Setting_Key': s.setting_key,
                    'Setting_Name': s.setting_name,
                    'New_Value': str(s.setting_value)
                })
        
        return pd.DataFrame(records)


# Location code mapping for operations analysis
# CRITICAL: Operations within baseline.corp/shared forest ARE IN THE baseline.corp HTML file
# There is no separate baseline.corp HTML file - it's the same as baseline.corp
LOCATION_MAPPING = {
    # Standalone domains (HTML file = domain)
    # shared_prefixes: if set, also pull matching GPOs from baseline.corp into this operation's report
    'OPA': {'full_name': 'Operation Alpha', 'source_domain': 'corp.alpha.com', 'target_domain': 'corp.alpha.com', 'shared_prefixes': ['OPA', 'Alpha']},
    'OPB': {'full_name': 'Operation Bravo', 'source_domain': 'corp.bravo.com', 'target_domain': 'corp.bravo.com', 'shared_prefixes': ['OPB', 'Bravo']},
    'OPC': {'full_name': 'Operation Charlie', 'source_domain': 'charlie.local', 'target_domain': 'charlie.local'},
    'OPD': {'full_name': 'Operation Delta', 'source_domain': 'delta.local', 'target_domain': 'delta.local'},
    'OPE': {'full_name': 'Operation Echo', 'source_domain': 'echo.corp.com', 'target_domain': 'echo.corp.com'},

    # Shared forest operations (in baseline.corp HTML, tagged with location codes)
    # The HTML file is baseline.corp.html, DomainName field = "baseline.corp"
    # name_prefixes: prefixes used in GPO DisplayNames for this operation
    'OPF': {'full_name': 'Operation Foxtrot', 'source_domain': 'baseline.corp', 'target_domain': 'baseline.corp', 'name_prefixes': ['OPF', 'Foxtrot']},
    'OPG': {'full_name': 'Operation Golf', 'source_domain': 'baseline.corp', 'target_domain': 'baseline.corp', 'name_prefixes': ['OPG', 'Golf']},
    'OPH': {'full_name': 'Operation Hotel', 'source_domain': 'baseline.corp', 'target_domain': 'baseline.corp', 'name_prefixes': ['OPH', 'Hotel']},
    'OPI': {'full_name': 'Operation India', 'source_domain': 'baseline.corp', 'target_domain': 'baseline.corp', 'name_prefixes': ['OPI', 'India']},
    'OPJ': {'full_name': 'Operation Juliet', 'source_domain': 'baseline.corp', 'target_domain': 'baseline.corp', 'name_prefixes': ['OPJ', 'Juliet']},
}

# Shared Forest operations list for validation
SHARED_FOREST_OPERATIONS = ['OPF', 'OPG', 'OPH', 'OPI', 'OPJ']

# ================================================================================
# BUCKET DETECTION - Dynamic OU-based categorization
# ================================================================================

# Keywords for bucket type detection (case-insensitive matching)
# Order matters: more specific matches first
BUCKET_KEYWORDS = {
    'Domain Controller': ['Domain Controllers', 'DomainControllers'],
    'Server': ['Server', 'Servers'],
    'Workstation': ['Computer', 'Computers', 'Workstation', 'Workstations', 'Desktop', 'Desktops', 'PC', 'PCs', 'Laptop', 'Laptops'],
    'User': ['User', 'Users', 'People'],
}

def parse_dn_path(link_path: str) -> dict:
    """
    Parse a Distinguished Name path to extract OU hierarchy.
    DN paths read right-to-left (domain root first, specific OU last).
    
    Example:
        Input: "OU=Admin,OU=Workstations,OU=Computers,OU=Foxtrot,DC=baseline,DC=corp"
        Output: {
            'domain': 'baseline.corp',
            'ou_hierarchy': ['Foxtrot', 'Computers', 'Workstations', 'Admin'],  # Bottom-up order
            'raw_path': 'OU=Admin,OU=Workstations,OU=Computers,OU=Foxtrot,DC=baseline,DC=corp'
        }
    """
    if not link_path or pd.isna(link_path):
        return {'domain': '', 'ou_hierarchy': [], 'raw_path': ''}
    
    link_path = str(link_path).strip()
    
    # Extract domain components (DC=xxx)
    dc_parts = re.findall(r'DC=([^,]+)', link_path, re.IGNORECASE)
    domain = '.'.join(dc_parts) if dc_parts else ''
    
    # Extract OU components - order in DN is leaf-to-root, so reverse for root-to-leaf
    ou_parts = re.findall(r'OU=([^,]+)', link_path, re.IGNORECASE)
    # Reverse to get hierarchy from domain root down to linked OU
    ou_hierarchy = list(reversed(ou_parts))
    
    return {
        'domain': domain,
        'ou_hierarchy': ou_hierarchy,
        'raw_path': link_path
    }

def detect_bucket_from_path(ou_hierarchy: list) -> str:
    """
    Detect bucket type by searching OU hierarchy for keywords.
    
    IMPORTANT: Scans from LEAF to ROOT (most specific OU first).
    This ensures that OU=Servers,OU=Computers,OU=Alpha gets detected as
    'Server' not 'Workstation', because Servers is the most specific OU.

    Args:
        ou_hierarchy: List of OUs from root to leaf (e.g., ['Foxtrot', 'Computers', 'Workstations', 'Admin'])
    
    Returns:
        Bucket type: 'Server', 'Workstation', 'User', 'Domain Controller', 'Domain Root', or 'Unknown'
    """
    if not ou_hierarchy:
        return 'Domain Root'
    
    # Search from LEAF to ROOT (reverse order) - most specific OU takes priority
    for ou_name in reversed(ou_hierarchy):
        ou_lower = ou_name.lower()
        
        # Check each bucket type's keywords
        for bucket_type, keywords in BUCKET_KEYWORDS.items():
            for keyword in keywords:
                if keyword.lower() in ou_lower:
                    return bucket_type
    
    # No bucket keyword found
    return 'Unknown'

def detect_operation_from_path(ou_hierarchy: list, domain: str) -> str:
    """
    Extract operation name from OU hierarchy.
    
    For baseline.corp/shared forest: First OU after domain (before bucket keywords)
    For standalone domains: Use domain name as operation
    
    Args:
        ou_hierarchy: List of OUs from root to leaf
        domain: Domain name (e.g., 'baseline.corp', 'echo.corp.com')

    Returns:
        Operation name (e.g., 'Foxtrot', 'Golf', 'echo.corp.com')
    """
    # Standalone domains: domain IS the operation
    if domain.lower() not in ['baseline.corp', 'baseline.corp']:
        return domain
    
    # baseline.corp: First OU is typically the operation/location
    if ou_hierarchy:
        # First OU (closest to domain root) is usually the operation
        # E.g., ['Region1', 'Foxtrot', 'Computers', 'Admin'] -> 'Region1' is region, 'Foxtrot' is operation
        # But sometimes: ['Foxtrot', 'Computers', 'Admin'] -> 'Foxtrot' is operation
        
        # Heuristic: Skip common regional OUs, find first non-bucket OU
        regional_ous = ['Region1', 'Region2', 'Sites', 'ENT', 'Operations']
        
        for ou in ou_hierarchy:
            # Skip regional/structural OUs
            if ou in regional_ous:
                continue
            
            # Check if this OU is a bucket keyword (skip if so)
            is_bucket = False
            ou_lower = ou.lower()
            for keywords in BUCKET_KEYWORDS.values():
                for keyword in keywords:
                    if keyword.lower() in ou_lower:
                        is_bucket = True
                        break
                if is_bucket:
                    break
            
            if not is_bucket:
                return ou  # This is the operation name
        
        # Fallback: first OU if nothing else matches
        return ou_hierarchy[0] if ou_hierarchy else 'Unknown'
    
    return 'Unknown'

def analyze_gpo_links(links_field: str) -> dict:
    """
    Analyze all links for a GPO and determine bucket type(s).
    
    Handles multi-link GPOs by detecting if linked to multiple bucket types.
    
    Args:
        links_field: The Links field from GPOZaurr (may contain multiple link paths)
    
    Returns:
        {
            'bucket': 'Server' | 'Workstation' | 'User' | 'Domain Root' | 'Domain Controller' | 'Mixed' | 'Unknown',
            'operation': 'Foxtrot' | 'Golf' | etc.,
            'bucket_details': {'Server': 2, 'Workstation': 1} for Mixed,
            'link_paths': [list of parsed link info]
        }
    """
    if not links_field or pd.isna(links_field):
        return {
            'bucket': 'Unknown',
            'operation': 'Unknown',
            'bucket_details': {},
            'link_paths': []
        }
    
    links_str = str(links_field)
    
    # Parse individual link paths
    # Links field may contain multiple paths separated by various delimiters
    # Common patterns: JSON-like structure or semicolon-separated
    
    link_paths = []
    bucket_counts = defaultdict(int)
    operations = set()
    
    # Try to extract individual DN paths
    # Pattern: Match complete DN paths (starts with OU=/CN=, ends with DC= components)
    dn_pattern = r'(?:OU|CN)=[^,]+(?:,(?:OU|CN)=[^,]+)*,DC=[^,]+(?:,DC=[^,]+)+'
    dn_matches = re.findall(dn_pattern, links_str, re.IGNORECASE)
    
    if not dn_matches:
        # Fallback: treat whole string as single path
        dn_matches = [links_str] if 'DC=' in links_str else []
    
    for dn_path in dn_matches:
        parsed = parse_dn_path(dn_path)
        bucket = detect_bucket_from_path(parsed['ou_hierarchy'])
        operation = detect_operation_from_path(parsed['ou_hierarchy'], parsed['domain'])
        
        link_paths.append({
            'path': dn_path,
            'bucket': bucket,
            'operation': operation,
            'ou_hierarchy': parsed['ou_hierarchy']
        })
        
        bucket_counts[bucket] += 1
        operations.add(operation)
    
    # Determine final bucket type
    if not bucket_counts:
        final_bucket = 'Unknown'
    elif len(bucket_counts) == 1:
        final_bucket = list(bucket_counts.keys())[0]
    else:
        # Multiple bucket types = Mixed
        final_bucket = 'Mixed'
    
    # Determine operation (use most common, or first if tie)
    final_operation = list(operations)[0] if operations else 'Unknown'
    
    return {
        'bucket': final_bucket,
        'operation': final_operation,
        'bucket_details': dict(bucket_counts),
        'link_paths': link_paths
    }


def format_links_for_display(links_str: str, max_links: int = None) -> str:
    """
    Format GPO Links field for readable display in Excel.
    
    Converts comma-separated DN paths into semicolon+newline separated, simplified paths.
    
    Args:
        links_str: Raw Links field from GPOZaurr (comma-separated DN paths)
        max_links: Maximum number of links to show (None = show all)
    
    Returns:
        Formatted string with each link separated by semicolon and newline
    
    Example:
        Input: "OU=Admin,OU=Workstations,OU=Foxtrot,DC=enterprise,DC=ad, OU=Servers,OU=Golf,DC=enterprise,DC=ad"
        Output: "Foxtrot > Workstations > Admin;\nGolf > Servers"
    """
    if not links_str or pd.isna(links_str):
        return ''
    
    links_str = str(links_str).strip()
    if not links_str:
        return ''
    
    # Parse individual DN paths - each DN starts with OU=/CN= and ends with DC=xxx,DC=yyy
    # This regex matches complete Distinguished Name paths
    dn_pattern = r'(?:OU|CN)=[^,]+(?:,(?:OU|CN)=[^,]+)*,DC=[^,]+(?:,DC=[^,]+)+'
    dn_matches = re.findall(dn_pattern, links_str, re.IGNORECASE)
    
    if not dn_matches:
        # Fallback: treat whole string as single path
        dn_matches = [links_str] if 'DC=' in links_str else []
    
    # Process all links (or limited if max_links specified)
    links_to_process = dn_matches if max_links is None else dn_matches[:max_links]
    
    formatted_links = []
    for dn_path in links_to_process:
        parsed = parse_dn_path(dn_path)
        
        if parsed['ou_hierarchy']:
            # Format as: Operation > Level1 > Level2 > ... (root-to-leaf order)
            formatted = ' > '.join(parsed['ou_hierarchy'])
            formatted_links.append(formatted)
        else:
            # Domain root link
            formatted_links.append(f"[Domain Root: {parsed['domain']}]")
    
    # Add indicator if there are more links (only if max_links was specified)
    if max_links is not None and len(dn_matches) > max_links:
        formatted_links.append(f"... and {len(dn_matches) - max_links} more")
    
    # Join with semicolon + newline for clear separation in Excel
    return ';\n'.join(formatted_links)


class GPOAnalyzer:
    """Main class for analyzing GPO data across domains - v2.3.2
    
    Supports three modes:
    - executive: Cross-domain consolidation summary (7 tabs)
    - domain: Bucket-focused migration plan (5 tabs) - Server/Workstation/User/Review
    - full: Complete v1.9.1 report (19 tabs) - default fallback
    
    NEW in v2.3.2:
    - Bucket detection from OU link paths
    - Operation detection from OU structure (no name pattern matching)
    - Domain mode tabs restructured by bucket type
    """

    def __init__(self, html_folder: Path, mode: str = 'full', target_domain: str = None, operation: str = None):
        self.html_folder = html_folder
        self.mode = mode  # 'executive', 'domain', or 'full'
        self.target_domain = target_domain  # Required for domain mode (unless operation specified)
        self.operation = operation  # Optional: target specific operation (OPF, OPG, etc.)
        self.domains = {}
        self.active_gpos = pd.DataFrame()
        self.enterprise_standard_gpos = pd.DataFrame()
        self.ent_nonstandard_gpos = pd.DataFrame()
        self.domain_gpos = pd.DataFrame()  # Master list: all standalone domain GPOs (not shared, not enterprise/shared forest)
        self.report_gpos = pd.DataFrame()   # Filtered subset for domain/operation mode reporting
        self.links_only_gpos = pd.DataFrame()  # GPOs linked to operation OU but name doesn't match - for review
        self.shared_gpos = pd.DataFrame()
        self.operations_gpos = pd.DataFrame()
        self.all_settings = []
        self.operation_settings = {}  # Settings by operation
        self.enterprise_standards = []  # Settings in 8+ domains/operations
        # Enhanced tracking for drill-down
        self.setting_to_entities = {}  # Maps (SettingName, SettingValue) -> list of entities
        self.setting_to_gpos = {}      # Maps (SettingName, SettingValue) -> list of GPO names
        self.entity_setting_details = []  # Full detail: entity, setting, value, category, GPO
        
        # Track total GPOs for inactive calculation
        self.total_gpos_before_filter = 0

    @staticmethod
    def normalize_guid(guid):
        """Normalize GUID for comparison - lowercase, no dashes"""
        if pd.isna(guid):
            return ''
        guid_str = str(guid).lower().strip()
        # Remove dashes and braces
        guid_str = guid_str.replace('-', '').replace('{', '').replace('}', '')
        return guid_str

    @staticmethod
    def extract_location_code(gpo_name: str) -> List[str]:
        """Extract location codes from GPO name"""
        if pd.isna(gpo_name):
            return []
        
        locations = []
        name_upper = str(gpo_name).upper()
        
        # Check for location codes at start of name (e.g., "OPF - ", "OPG - ")
        for code in LOCATION_MAPPING.keys():
            # Pattern: "CODE - " at start or " - CODE - " anywhere
            if name_upper.startswith(f"{code} - ") or f" - {code} - " in name_upper:
                locations.append(code)
            # Pattern: "CODE-" at start
            elif name_upper.startswith(f"{code}-"):
                locations.append(code)
        
        return list(set(locations))  # Remove duplicates

    def parse_html_reports(self):
        """Parse all HTML reports in the folder"""
        html_files = list(self.html_folder.glob("*.html"))

        if not html_files:
            print(f"Error: No HTML files found in {self.html_folder}")
            sys.exit(1)

        print(f"\nFound {len(html_files)} HTML report(s)")

        for html_file in html_files:
            print(f"Processing: {html_file.name}")
            domain_name = self._extract_domain_name(html_file)
            self.domains[domain_name] = self._parse_single_report(html_file)

    def _extract_domain_name(self, html_file: Path) -> str:
        """Extract domain name from HTML content (preferred) or filename (fallback)"""
        # First, try to parse from HTML content (most accurate)
        with open(html_file, 'r', encoding='utf-8') as f:
            content = f.read()
            soup = BeautifulSoup(content, 'html.parser')

            # Try to find domain name in the data
            script_tags = soup.find_all('script')
            for script in script_tags:
                if script.string and 'DomainName' in script.string:
                    match = re.search(r'"DomainName":\s*"([^"]+)"', script.string)
                    if match:
                        domain_from_html = match.group(1)
                        logging.info(f"Extracted domain from HTML content: {domain_from_html}")
                        return domain_from_html

        # Fallback to filename if HTML parsing fails
        domain = html_file.stem
        if domain and domain != "invoke_gpozaurr_report":
            # Remove common suffixes that might be in filename
            domain = domain.replace('_report', '').replace('-report', '')
            logging.warning(f"Could not extract domain from HTML, using filename: {domain}")
            return domain

        logging.error(f"Could not determine domain name for {html_file}")
        return domain

    def _parse_single_report(self, html_file: Path) -> Dict[str, pd.DataFrame]:
        """Parse a single GPOZaurr HTML report"""
        with open(html_file, 'r', encoding='utf-8') as f:
            content = f.read()

        soup = BeautifulSoup(content, 'html.parser')

        report_data = {}

        # Parse GP Summary
        report_data['summary'] = self._extract_gp_summary(soup, content)

        # Parse GP Links
        report_data['links'] = self._extract_gp_links(soup, content)

        # Parse GP Content (all categories)
        report_data['content'] = self._extract_gp_content(soup, content)

        # Parse GP Permissions
        report_data['permissions'] = self._extract_gp_permissions(soup, content)

        return report_data

    def _extract_table_from_tab(self, content: str, tab_name: str,
                                required_columns: List[str]) -> pd.DataFrame:
        """Extract DataTable data from a specific tab by finding it in context"""

        # Find the tab ID for the given tab name
        tab_pattern = rf'<div id="(Tab-[\w]+)"[^>]*><span>{re.escape(tab_name)}</span></div>'
        tab_match = re.search(tab_pattern, content)

        if not tab_match:
            return pd.DataFrame()

        tab_id = tab_match.group(1)
        content_div_id = f"{tab_id}-Content"

        # Find where this tab's content starts
        content_start = content.find(f'id="{content_div_id}"')
        if content_start == -1:
            return pd.DataFrame()

        # Find the next tab's content (or end of content section)
        next_tab_match = re.search(r'<div id="Tab-[\w]+-Content"', content[content_start + 100:])
        if next_tab_match:
            content_end = content_start + 100 + next_tab_match.start()
        else:
            content_end = len(content)

        # Extract just this tab's content
        tab_content = content[content_start:content_end]

        # Find DataTable in this section
        table_pattern = r"'#(DT-[\w]+)'.*?\"data\":\s*\[(.*?)\],\s*\"columns\""
        match = re.search(table_pattern, tab_content, re.DOTALL)

        if not match:
            return pd.DataFrame()

        data_str = '[' + match.group(2) + ']'

        try:
            data = json.loads(data_str)
            if data and len(data) > 0:
                # Verify it has the required columns
                first_row = data[0]
                if all(col in first_row for col in required_columns):
                    return pd.DataFrame(data)
        except json.JSONDecodeError as e:
            print(f"    Warning: Could not parse JSON for {tab_name}: {e}")

        return pd.DataFrame()

    def _extract_gp_summary(self, soup: BeautifulSoup, content: str) -> pd.DataFrame:
        """Extract Group Policy Summary data"""
        print("  - Extracting GP Summary...")

        df = self._extract_table_from_tab(
            content,
            "Group Policy Summary",
            ['DisplayName', 'DomainName', 'GUID', 'Enabled', 'Linked']
        )

        if not df.empty:
            print(f"    Found {len(df)} GPOs")
        else:
            print(f"    WARNING: Could not extract GP Summary data")

        return df

    def _extract_gp_links(self, soup: BeautifulSoup, content: str) -> pd.DataFrame:
        """Extract Group Policy Links data"""
        print("  - Extracting GP Links...")

        df = self._extract_table_from_tab(
            content,
            "Group Policy Links",
            ['DisplayName', 'DomainName', 'GUID', 'Linked']
        )

        if not df.empty:
            print(f"    Found {len(df)} GPO link records")

        return df

    def _extract_gp_content(self, soup: BeautifulSoup, content: str) -> Dict[str, pd.DataFrame]:
        """Extract Group Policy Content data (all subcategories)"""
        print("  - Extracting GP Content (40+ categories)...")

        # Find the Group Policy Content tab
        tab_pattern = r'<div id="(Tab-[\w]+)"[^>]*><span>Group Policy Content</span></div>'
        tab_match = re.search(tab_pattern, content)

        if not tab_match:
            print("    WARNING: Could not find Group Policy Content tab")
            return {}

        tab_id = tab_match.group(1)
        content_div_id = f"{tab_id}-Content"

        # Find where this tab's content starts
        content_start = content.find(f'id="{content_div_id}"')
        if content_start == -1:
            print("    WARNING: Could not find Group Policy Content section")
            return {}

        search_from = content_start + 100
        end_pattern = r'</div></div></div></div><div id="Tab-[\w]+-Content"'
        end_match = re.search(end_pattern, content[search_from:])

        if end_match:
            content_end = search_from + end_match.start() + 28
        else:
            content_end = len(content)

        gp_content_section = content[content_start:content_end]

        # Find all nested tabs within GP Content
        nested_tab_pattern = r'<div id="(Tab-[\w]+)"[^>]*><span>([^<]+)</span></div>'
        nested_tabs = re.findall(nested_tab_pattern, gp_content_section)

        logging.info(f"Found {len(nested_tabs)} nested tabs in GP Content")

        categories = {}

        for tab_id, tab_name in nested_tabs:
            nested_content_id = f"{tab_id}-Content"

            nested_start = gp_content_section.find(f'id="{nested_content_id}"')
            if nested_start == -1:
                continue

            next_nested = re.search(r'<div id="Tab-[\w]+-Content"', gp_content_section[nested_start + 100:])
            if next_nested:
                nested_end = nested_start + 100 + next_nested.start()
            else:
                nested_end = len(gp_content_section)

            nested_section = gp_content_section[nested_start:nested_end]

            data_pattern = r'"data":\s*\[(.*?)\],\s*"columns"'
            match = re.search(data_pattern, nested_section, re.DOTALL)

            if match:
                data_str = '[' + match.group(1) + ']'
                try:
                    data = json.loads(data_str)
                    if data and len(data) > 0:
                        first_row = data[0]
                        if 'DisplayName' in first_row or 'GUID' in first_row or 'DomainName' in first_row:
                            df = pd.DataFrame(data)
                            categories[tab_name] = df
                            logging.debug(f"  Category '{tab_name}': {len(df)} rows")
                except json.JSONDecodeError as e:
                    logging.warning(f"  Could not parse JSON for category '{tab_name}': {e}")
                    continue

        print(f"    Found {len(categories)} content categories")
        return categories

    def _extract_gp_permissions(self, soup: BeautifulSoup, content: str) -> pd.DataFrame:
        """Extract Group Policy Permissions Analysis data"""
        print("  - Extracting GP Permissions...")

        df = self._extract_table_from_tab(
            content,
            "Group Policy Permissions Analysis",
            ['DisplayName', 'DomainName']
        )

        if not df.empty:
            print(f"    Found {len(df)} permission records")

        return df

    def filter_active_gpos(self):
        """Filter to active GPOs, detect shared GPOs, split ENT vs Domain vs Operations"""
        print("\n=== Filtering Active GPOs ===")
        
        # Step 1: Collect all active GPOs
        self._collect_active_gpos()
        
        # Step 2: Detect shared GPOs (baseline.corp â†” baseline.corp)
        self._detect_shared_gpos()
        
        # Step 3: Create domain_gpos (not shared, not enterprise/shared forest)
        self._create_domain_gpos()
        
        # Step 4: Filter ENT from ALL domains, track per-domain
        self._filter_and_track_ent()
        
        # Step 5: Categorize shared GPOs (operations vs Enterprise Standard)
        self._categorize_shared_gpos()
    
    def _collect_active_gpos(self):
        """Collect all active GPOs from all domains"""
        all_gpos = []

        for domain, data in self.domains.items():
            if 'summary' in data and not data['summary'].empty:
                df = data['summary'].copy()
                df['Domain'] = domain
                all_gpos.append(df)

        if not all_gpos:
            print("Warning: No GPO summary data found")
            return

        combined = pd.concat(all_gpos, ignore_index=True)
        
        # Track total for inactive calculation
        self.total_gpos_before_filter = len(combined)

        print(f"Total GPOs across all domains: {len(combined)}")

        # Check which columns we have
        has_enabled = 'Enabled' in combined.columns
        has_linked = 'Linked' in combined.columns

        if not has_enabled or not has_linked:
            print(f"Warning: Missing required columns. Enabled={has_enabled}, Linked={has_linked}")
            self.active_gpos = combined
        else:
            combined['Enabled_bool'] = combined['Enabled'].astype(str).str.lower().isin(['true', 'yes', '1'])
            combined['Linked_bool'] = combined['Linked'].astype(str).str.lower().isin(['true', 'yes', '1'])

            self.active_gpos = combined[
                (combined['Enabled_bool'] == True) &
                (combined['Linked_bool'] == True)
                ].copy()

            print(f"Active GPOs (Enabled AND Linked): {len(self.active_gpos)}")
        
        # Add bucket analysis to all active GPOs
        self._analyze_gpo_buckets()
    
    def _analyze_gpo_buckets(self):
        """Analyze bucket type, operation, and configuration target for all active GPOs"""
        print("\n=== Analyzing GPO Buckets (OU-based) ===")
        
        if self.active_gpos.empty:
            print("  No active GPOs to analyze")
            return
        
        # Check if Links column exists
        if 'Links' not in self.active_gpos.columns:
            print("  WARNING: No 'Links' column found - cannot determine buckets")
            self.active_gpos['Bucket'] = 'Unknown'
            self.active_gpos['DetectedOperation'] = 'Unknown'
            self.active_gpos['BucketDetails'] = ''
            self.active_gpos['AppliesTo'] = 'Unknown'
            self.active_gpos['ConfigMismatch'] = False
            return
        
        # Analyze each GPO's links and configuration
        buckets = []
        operations = []
        bucket_details = []
        applies_to_list = []
        config_mismatch_list = []
        
        for _, gpo in self.active_gpos.iterrows():
            links_field = gpo.get('Links', '')
            analysis = analyze_gpo_links(links_field)
            
            buckets.append(analysis['bucket'])
            operations.append(analysis['operation'])
            bucket_details.append(str(analysis['bucket_details']) if analysis['bucket_details'] else '')
            
            # Determine AppliesTo based on ComputerEnabled/UserEnabled
            computer_enabled = str(gpo.get('ComputerEnabled', 'True')).lower() == 'true'
            user_enabled = str(gpo.get('UserEnabled', 'True')).lower() == 'true'
            
            if computer_enabled and user_enabled:
                applies_to = 'Both'
            elif computer_enabled and not user_enabled:
                applies_to = 'Computer'
            elif user_enabled and not computer_enabled:
                applies_to = 'User'
            else:
                applies_to = 'None'  # Shouldn't happen if GPO is Enabled
            
            applies_to_list.append(applies_to)
            
            # Detect configuration mismatch
            # Mismatch = GPO targets don't align with OU bucket
            bucket = analysis['bucket']
            mismatch = False
            
            # Computer-targeted OUs (Server, Workstation, Domain Controller) with User-only GPO
            if bucket in ['Server', 'Workstation', 'Domain Controller'] and applies_to == 'User':
                mismatch = True
            # User-targeted OU with Computer-only GPO
            elif bucket == 'User' and applies_to == 'Computer':
                mismatch = True
            
            config_mismatch_list.append(mismatch)
        
        self.active_gpos['Bucket'] = buckets
        self.active_gpos['DetectedOperation'] = operations
        self.active_gpos['BucketDetails'] = bucket_details
        self.active_gpos['AppliesTo'] = applies_to_list
        self.active_gpos['ConfigMismatch'] = config_mismatch_list
        
        # Print summary
        bucket_counts = self.active_gpos['Bucket'].value_counts()
        print(f"  Bucket distribution:")
        for bucket, count in bucket_counts.items():
            print(f"    {bucket}: {count}")
        
        # Print AppliesTo summary
        applies_counts = self.active_gpos['AppliesTo'].value_counts()
        print(f"\n  Configuration target (AppliesTo):")
        for target, count in applies_counts.items():
            print(f"    {target}: {count}")
        
        # Print mismatch count
        mismatch_count = self.active_gpos['ConfigMismatch'].sum()
        if mismatch_count > 0:
            print(f"\n  âš  Configuration mismatches detected: {mismatch_count}")
            print(f"    (GPOs where target doesn't match OU type)")
        
        # Print discovered operations
        discovered_ops = self.active_gpos['DetectedOperation'].value_counts()
        print(f"\n  Discovered operations ({len(discovered_ops)} total):")
        for op, count in discovered_ops.head(15).items():
            print(f"    {op}: {count}")
        if len(discovered_ops) > 15:
            print(f"    ... and {len(discovered_ops) - 15} more")
    
    def _detect_shared_gpos(self):
        """Detect shared GPOs between baseline.corp and baseline.corp"""
        print("\n=== Detecting Shared GPOs (baseline.corp â†” baseline.corp) ===")
        
        self.active_gpos['GUID_normalized'] = self.active_gpos['GUID'].apply(self.normalize_guid)
        
        # Get GUIDs from baseline.corp and baseline.corp
        ent_ad = self.active_gpos[self.active_gpos['Domain'] == 'baseline.corp']
        shared_forest = self.active_gpos[self.active_gpos['Domain'] == 'baseline.corp']

        if not ent_ad.empty and not shared_forest.empty:
            ent_guids = set(ent_ad['GUID_normalized'].values)
            shared_guids_set = set(shared_forest['GUID_normalized'].values)
            self.shared_guids = ent_guids & shared_guids_set
            
            print(f"Shared GPOs between baseline.corp and baseline.corp: {len(self.shared_guids)}")
            
            # Mark shared GPOs (use baseline.corp as canonical source)
            self.shared_gpos = ent_ad[ent_ad['GUID_normalized'].isin(self.shared_guids)].copy()
        else:
            print("Note: baseline.corp or baseline.corp not found - no shared GPO detection")
            self.shared_guids = set()
            self.shared_gpos = pd.DataFrame()
    
    def _create_domain_gpos(self):
        """Create domain_gpos - all active GPOs except shared and enterprise/shared forest"""
        # Exclude: shared GPOs + baseline.corp domain + baseline.corp domain
        self.domain_gpos = self.active_gpos[
            ~self.active_gpos['GUID_normalized'].isin(self.shared_guids) &
            ~self.active_gpos['Domain'].isin(['baseline.corp', 'baseline.corp'])
        ].copy()
        
        print(f"\n=== Creating Domain GPOs ===")
        print(f"Domain GPOs (before ENT filter): {len(self.domain_gpos)}")
    
    def _filter_and_track_ent(self):
        """Filter ENT GPOs from domain_gpos and track counts per domain"""
        print("\n=== Filtering ENT GPOs by Name ===")
        
        self.ent_filtered_per_domain = {}
        
        if self.domain_gpos.empty:
            print("No domain GPOs to filter")
            return
        
        # Process each domain
        for domain in self.domain_gpos['Domain'].unique():
            # Count before filtering
            before = len(self.domain_gpos[self.domain_gpos['Domain'] == domain])
            
            # Create mask: GPOs from this domain that contain "ENT"
            domain_mask = self.domain_gpos['Domain'] == domain
            ent_mask = self.domain_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
            remove_mask = domain_mask & ent_mask
            
            # Filter out ENT GPOs
            self.domain_gpos = self.domain_gpos[~remove_mask].copy()
            
            # Count after filtering
            after = len(self.domain_gpos[self.domain_gpos['Domain'] == domain])
            filtered_count = before - after
            self.ent_filtered_per_domain[domain] = filtered_count
            
            if filtered_count > 0:
                print(f"  {domain}: Filtered out {filtered_count} ENT GPOs")
        
        print(f"Domain GPOs (after ENT filter): {len(self.domain_gpos)}")
    
    def _categorize_shared_gpos(self):
        """Split shared GPOs into Enterprise Standard vs Operations"""
        print("\n=== Categorizing Shared GPOs: ENT vs Operations ===")
        
        # If shared detection worked (two domains found), use shared_gpos
        # Otherwise, use all baseline.corp GPOs directly
        if self.shared_gpos.empty:
            print("No shared GPOs detected (two-domain intersection)")
            print("Looking for baseline.corp GPOs directly...")
            
            # Get all baseline.corp GPOs from active_gpos
            enterprise_gpos = self.active_gpos[
                self.active_gpos['Domain'] == 'baseline.corp'
            ].copy()
            
            if enterprise_gpos.empty:
                print("No baseline.corp GPOs found")
                self.enterprise_standard_gpos = pd.DataFrame()
                self.ent_nonstandard_gpos = pd.DataFrame()
                self.operations_gpos = pd.DataFrame()
                return
            
            print(f"Found {len(enterprise_gpos)} baseline.corp GPOs to categorize")
            source_gpos = enterprise_gpos
        else:
            print(f"Using {len(self.shared_gpos)} shared GPOs from two-domain detection")
            source_gpos = self.shared_gpos
        
        # Enterprise Standard
        ent_standard_mask = source_gpos['DisplayName'].str.startswith('ENT - ', na=False)
        self.enterprise_standard_gpos = source_gpos[ent_standard_mask].copy()
        
        # ENT Non-Standard
        ent_any_mask = source_gpos['DisplayName'].str.upper().str.startswith('ENT', na=False)
        ent_nonstandard_mask = ent_any_mask & ~ent_standard_mask
        self.ent_nonstandard_gpos = source_gpos[ent_nonstandard_mask].copy()
        
        # Operations (shared but not ENT)
        self.operations_gpos = source_gpos[~ent_any_mask].copy()
        
        # Extract location codes for operations GPOs
        self.operations_gpos['LocationCodes'] = self.operations_gpos['DisplayName'].apply(self.extract_location_code)
        self.operations_gpos['LocationCodesStr'] = self.operations_gpos['LocationCodes'].apply(lambda x: ', '.join(x) if x else 'Unknown')
        
        print(f"Enterprise Standard: {len(self.enterprise_standard_gpos)}")
        print(f"ENT Non-Standard: {len(self.ent_nonstandard_gpos)}")
        print(f"Operations GPOs: {len(self.operations_gpos)}")
        
        if len(self.ent_nonstandard_gpos) > 0:
            print(f"âš ï¸  WARNING: {len(self.ent_nonstandard_gpos)} GPOs have non-standard ENT naming!")


    def analyze_settings_patterns(self):
        """Analyze setting patterns - DOMAIN-UNIQUE GPOs ONLY"""
        print("\n=== Analyzing Setting Patterns (Domain-Unique GPOs Only) ===")

        # Analyze only domain_gpos (excludes all shared enterprise/shared forest GPOs)
        guid_col = None
        if not self.domain_gpos.empty:
            for possible_col in ['GUID', 'Id', 'GPOGuid', 'Guid']:
                if possible_col in self.domain_gpos.columns:
                    guid_col = possible_col
                    print(f"Using '{guid_col}' as GPO identifier")
                    break

        if not guid_col:
            print("Warning: No GUID/Id column found")

        # Collect settings from domain-unique GPOs only
        for domain, data in self.domains.items():
            if 'content' not in data:
                continue

            # Get domain-specific GPO IDs only
            if not self.domain_gpos.empty and guid_col:
                domain_ids = set(
                    self.domain_gpos[self.domain_gpos['Domain'] == domain][guid_col].values
                )
            else:
                domain_ids = None

            for category_name, df in data['content'].items():
                if df.empty:
                    continue

                # Find identifier column
                df_id_col = None
                if guid_col and guid_col in df.columns:
                    df_id_col = guid_col
                else:
                    for possible_col in ['GUID', 'Id', 'GPOGuid', 'Guid']:
                        if possible_col in df.columns:
                            df_id_col = possible_col
                            break

                # Filter to domain-unique GPOs only
                if df_id_col and domain_ids:
                    df_normalized = df.copy()
                    df_normalized['GUID_normalized'] = df_normalized[df_id_col].apply(self.normalize_guid)
                    domain_ids_normalized = set(self.normalize_guid(g) for g in domain_ids)

                    df = df_normalized[df_normalized['GUID_normalized'].isin(domain_ids_normalized)]
                    if not df.empty:
                        df = df.drop(columns=['GUID_normalized'])

                if df.empty:
                    continue

                # Extract settings
                for _, row in df.iterrows():
                    gpo_name = row.get('DisplayName', row.get('Name', 'Unknown'))

                    skip_cols = ['DisplayName', 'Name', 'DomainName', 'GUID', 'Id', 'GPOGuid', 'Guid',
                                 'GpoType', 'Filters', 'Linked', 'LinksCount', 'Links']

                    for col, val in row.items():
                        if col not in skip_cols:
                            if pd.notna(val) and str(val).strip() and str(val) != 'Not Set':
                                setting_key = (col, str(val))
                                
                                # Track individual settings (existing)
                                setting_record = {
                                    'Domain': domain,
                                    'GPOName': gpo_name,
                                    'Category': category_name,
                                    'SettingName': col,
                                    'SettingValue': val
                                }
                                self.all_settings.append(setting_record)
                                
                                # NEW v1.9.1: Track associations for drill-down
                                # Map setting -> entities
                                if setting_key not in self.setting_to_entities:
                                    self.setting_to_entities[setting_key] = set()
                                self.setting_to_entities[setting_key].add(domain)
                                
                                # Map setting -> GPO names
                                if setting_key not in self.setting_to_gpos:
                                    self.setting_to_gpos[setting_key] = set()
                                self.setting_to_gpos[setting_key].add(gpo_name)
                                
                                # Track detailed record for pivot analysis
                                self.entity_setting_details.append({
                                    'Entity': domain,
                                    'SettingName': col,
                                    'SettingValue': str(val),
                                    'Category': category_name,
                                    'GPOName': gpo_name,
                                    'SettingKey': f"{col}={val}"
                                })

        print(f"Total settings extracted (Domain-Unique GPOs only): {len(self.all_settings)}")

    def analyze_for_decisions(self):
        """Analyze settings to generate consolidation recommendations"""
        print("\n=== Analyzing for Consolidation Decisions ===")
        
        if not self.all_settings:
            print("No settings to analyze for decisions")
            return
        
        # Group settings by operation/domain
        print("  - Grouping settings by operation/domain...")
        settings_df = pd.DataFrame(self.all_settings)
        
        for entity in settings_df['Domain'].unique():
            entity_settings = settings_df[settings_df['Domain'] == entity]
            self.operation_settings[entity] = entity_settings
        
        # Also add settings from operations in shared GPOs
        if not self.operations_gpos.empty:
            print("  - Analyzing operations in shared enterprise GPOs...")
            # Extract location-based groupings
            for code, info in LOCATION_MAPPING.items():
                ops_with_location = self.operations_gpos[
                    self.operations_gpos['LocationCodes'].apply(lambda x: code in x)
                ]
                if len(ops_with_location) > 0:
                    operation_name = f"{code} (Shared Forest)"
                    # Note: We'd need settings data from these GPOs
                    # For now, just track the GPO count
                    if operation_name not in self.operation_settings:
                        self.operation_settings[operation_name] = pd.DataFrame()
        
        # Identify enterprise standard candidates (settings in 8+ entities)
        print("  - Identifying enterprise standard candidates...")
        setting_counts = settings_df.groupby(['SettingName', 'SettingValue']).agg({
            'Domain': lambda x: len(set(x))
        }).reset_index()
        setting_counts.columns = ['SettingName', 'SettingValue', 'EntityCount']
        
        # Enterprise threshold: 8+ domains/operations
        self.enterprise_standards = setting_counts[setting_counts['EntityCount'] >= 8].to_dict('records')
        
        print(f"  - Found {len(self.enterprise_standards)} enterprise standard candidates")
        print(f"  - Tracking settings for {len(self.operation_settings)} operations/domains")

    def generate_excel_report(self, output_path: Path):
        """Generate Excel report based on mode"""
        print(f"\n=== Generating {self.mode.upper()} Report: {output_path} ===")
        
        if self.mode == 'executive':
            self._generate_executive_report(output_path)
        elif self.mode == 'domain':
            # Domain mode requires either --domain or --operation
            if not self.target_domain and not self.operation:
                print("ERROR: Domain mode requires either --domain or --operation parameter")
                sys.exit(1)
            
            # If operation specified, auto-detect domain from LOCATION_MAPPING
            if self.operation:
                if self.operation not in LOCATION_MAPPING:
                    print(f"ERROR: Unknown operation code '{self.operation}'")
                    print(f"Valid operations: {', '.join(sorted(LOCATION_MAPPING.keys()))}")
                    sys.exit(1)
                
                # Get source domain for operation
                source_domain = LOCATION_MAPPING[self.operation]['source_domain']
                
                # If operation is shared forest operation, ensure we're analyzing correct domain
                if self.target_domain and self.target_domain != source_domain:
                    print(f"WARNING: Operation {self.operation} is in {source_domain}, not {self.target_domain}")
                    print(f"Using {source_domain} as source domain")
                
                self.target_domain = source_domain
                print(f"\nOperation Mode: {LOCATION_MAPPING[self.operation]['full_name']} ({self.operation})")
                print(f"Source Domain: {source_domain} (loads from baseline.corp or baseline.corp HTML file)")
                print(f"Target Domain: {LOCATION_MAPPING[self.operation]['target_domain']}")
                
                # Helpful note for shared forest operations
                if self.operation in SHARED_FOREST_OPERATIONS:
                    print(f"NOTE: {self.operation} GPOs are in the baseline.corp HTML file (DomainName field)")
                    print(f"      The file may be named 'baseline.corp.html' but contains DomainName='baseline.corp'")
            
            self._generate_domain_report(output_path)
        else:
            # Full v1.9.1 report (default fallback)
            self._generate_full_report(output_path)
    
    def _generate_full_report(self, output_path: Path):
        """Generate comprehensive v1.9.1 Excel report with all 19 tabs"""
        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        # Tab 1: Executive Summary (with deduplication)
        self._create_executive_summary(writer)

        # Tab 2: Shared Enterprise GPOs (552 counted once)
        self._create_shared_gpos_inventory(writer)

        # Tab 3: Shared Forest Operations by Location
        self._create_operations_by_location(writer)

        # Tab 4: Enterprise Standard GPOs
        self._create_ent_inventory(writer)

        # Tab 5: Non-Standard ENT Names
        self._create_nonstandard_ent(writer)

        # Tab 6-8: Setting Analysis (Domain-Unique only)
        self._create_setting_analysis(writer)

        # Tab 9: Performance Issues
        self._create_performance_analysis(writer)

        # Tab 10: Migration Gotchas
        self._create_migration_gotchas(writer)

        # Tab 11: Security & Permissions
        self._create_security_analysis(writer)

        # Tab 12: OU Link Analysis
        self._create_link_analysis(writer)

        # Tab 13: Migration Roadmap
        self._create_migration_roadmap(writer)
        
        # Tab 14: Consolidation Recommendations
        self._create_consolidation_recommendations(writer)
        
        # Tab 14a: Settings Breakdown by Entity (NEW in v1.9.1)
        self._create_settings_breakdown(writer)
        
        # Tab 15: Gap Analysis
        self._create_gap_analysis(writer)
        
        # Tab 16: Settings Detail View (NEW in v1.9.1)
        self._create_settings_detail_view(writer)

        # Tab 17-18: Raw Data
        self._create_raw_data(writer)

        writer.close()
        print(f"\nâœ“ Report generated successfully: {output_path}")
    
    def _generate_executive_report(self, output_path: Path):
        """Generate executive cross-domain summary (7 tabs) - ENHANCED v2.3.2"""
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        print("  - Creating Tab 1: Migration Dashboard...")
        self._create_exec_tab1_migration_dashboard(writer)
        
        print("  - Creating Tab 2: Operations Summary...")
        self._create_exec_tab2_operations_summary(writer)
        
        print("  - Creating Tab 3: Risk Assessment...")
        self._create_exec_tab3_risk_assessment(writer)
        
        print("  - Creating Tab 4: Bucket Comparison...")
        self._create_exec_tab4_bucket_comparison(writer)
        
        print("  - Creating Tab 5: Performance Issues...")
        self._create_exec_tab5_performance(writer)
        
        print("  - Creating Tab 6: Infrastructure Dependencies...")
        self._create_exec_tab6_infrastructure(writer)
        
        print("  - Creating Tab 7: Consolidation Roadmap...")
        self._create_exec_tab7_roadmap(writer)
        
        writer.close()
        print(f"\nâœ“ Executive report generated: {output_path}")
    
    def _filter_to_operation(self):
        """Filter GPOs to only include those for specified operation using OU-based detection"""
        operation_name = self.operation
        
        print(f"\n=== Filtering to Operation: {operation_name} (OU-based) ===")
        logging.info(f"Filtering to operation: {operation_name}")
        
        # CRITICAL: Check source_domain to determine where GPOs live
        # Shared Forest operations (OPF, OPG, OPH, etc.) -> GPOs in baseline.corp
        # Standalone operations (OPA, OPB, OPC, etc.) -> GPOs in their own domain
        source_domain = None
        if operation_name in LOCATION_MAPPING:
            source_domain = LOCATION_MAPPING[operation_name].get('source_domain')
            logging.info(f"Operation {operation_name} source_domain: {source_domain}")
        
        # STANDALONE DOMAIN HANDLING
        # If source_domain is NOT baseline.corp, get GPOs directly from that domain
        if source_domain and source_domain != 'baseline.corp':
            print(f"  Standalone operation: Getting GPOs from {source_domain}")
            logging.info(f"Standalone domain mode for {operation_name}")
            
            # Get GPOs from the standalone domain
            domain_mask = self.active_gpos['Domain'].str.lower() == source_domain.lower()
            self.report_gpos = self.active_gpos[domain_mask].copy()
            
            if self.report_gpos.empty:
                print(f"  WARNING: No GPOs found in domain {source_domain}")
                available_domains = self.active_gpos['Domain'].unique().tolist()
                print(f"  Available domains: {available_domains}")
                return
            
            # Filter out ENT GPOs
            ent_mask = self.report_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
            if ent_mask.any():
                ent_count = ent_mask.sum()
                print(f"  Filtering out {ent_count} ENT GPOs")
                self.report_gpos = self.report_gpos[~ent_mask].copy()
            
            # Mark all as Domain GPO (standalone domains don't have Links-only distinction)
            self.report_gpos['MatchType'] = 'Domain GPO'
            self.links_only_gpos = pd.DataFrame()  # No links-only for standalone
            
            # Show results
            print(f"  Total GPOs for {operation_name}: {len(self.report_gpos)}")
            if not self.report_gpos.empty and 'Bucket' in self.report_gpos.columns:
                bucket_counts = self.report_gpos['Bucket'].value_counts()
                print(f"\n  Bucket breakdown for {operation_name}:")
                for bucket, count in bucket_counts.items():
                    print(f"    {bucket}: {count}")
            
            # Filter settings to this operation
            if self.all_settings:
                settings_before = len(self.all_settings)
                operation_gpo_names = set(self.report_gpos['DisplayName'].values)
                self.all_settings = [
                    s for s in self.all_settings 
                    if s.get('GPOName') in operation_gpo_names
                ]
                settings_after = len(self.all_settings)
                print(f"  Filtered settings: {settings_before} -> {settings_after}")
            
            return  # Done - standalone handling complete
        
        # SHARED FOREST OPERATION HANDLING (original logic)
        # For operations in baseline.corp/baseline.corp, use the DetectedOperation column
        # which was populated by _analyze_gpo_buckets()
        
        if 'DetectedOperation' not in self.active_gpos.columns:
            print("  ERROR: DetectedOperation column not found - bucket analysis may have failed")
            logging.error("DetectedOperation column not found in active_gpos")
            return
        
        # Get baseline.corp OR baseline.corp GPOs (they're the same forest, different domain names in reports)
        enterprise_gpos = self.active_gpos[
            self.active_gpos['Domain'].isin(['baseline.corp', 'baseline.corp'])
        ].copy()
        
        if enterprise_gpos.empty:
            print("  ERROR: No baseline.corp or baseline.corp GPOs found")
            logging.error("No baseline.corp/baseline.corp GPOs found for operation filtering")
            # Show what domains ARE available
            available_domains = self.active_gpos['Domain'].unique().tolist()
            print(f"  Available domains: {available_domains}")
            logging.info(f"Available domains in active_gpos: {available_domains}")
            return
        
        before_count = len(enterprise_gpos)
        domains_found = enterprise_gpos['Domain'].unique().tolist()
        print(f"  Source domains: {domains_found}")
        logging.info(f"Operation filter using domains: {domains_found}")
        
        # Build list of search terms from the operation input
        # Include: operation code, full name, and all name prefixes used in GPO names
        search_terms = [operation_name.lower()]
        
        # If it's a LOCATION_MAPPING code, add the full name and name prefixes
        if operation_name in LOCATION_MAPPING:
            op_config = LOCATION_MAPPING[operation_name]
            full_name = op_config['full_name']
            search_terms.append(full_name.lower())  # "Operation Foxtrot"
            search_terms.append(full_name.lower().replace(' ', ''))  # "operationfoxtrot"
            
            # Add name_prefixes if defined (e.g., OPH for Hotel, OPI for India)
            if 'name_prefixes' in op_config:
                for prefix in op_config['name_prefixes']:
                    if prefix.lower() not in search_terms:
                        search_terms.append(prefix.lower())
        
        # Remove duplicates while preserving order
        seen = set()
        search_terms = [x for x in search_terms if not (x in seen or seen.add(x))]
        
        logging.info(f"Operation search terms: {search_terms}")
        
        # Match by DisplayName - this matches manual filtering exactly
        # GPOs named "OPF-*" or "Foxtrot*" are clearly operation-specific
        # This is exactly what a human would do when manually filtering in Excel
        name_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
        
        for term in search_terms:
            if len(term) < 2:
                continue  # Skip single-char terms
            
            if len(term) <= 3:
                # Short codes (2-3 chars): require word boundary to avoid false positives
                # e.g., "OPJ" should match "OPJ-Printers" but not "Guest", "Forest", "Desktop"
                # Pattern: start of string OR preceded by space/dash/underscore
                #          followed by end of string OR space/dash/underscore
                pattern = r'(^|[\s\-_])' + re.escape(term) + r'($|[\s\-_])'
                term_in_name = enterprise_gpos['DisplayName'].str.contains(
                    pattern, case=False, na=False, regex=True
                )
            else:
                # Longer terms (4+ chars): simple contains match is safe
                term_in_name = enterprise_gpos['DisplayName'].str.contains(
                    term, case=False, na=False, regex=False
                )
            
            name_mask = name_mask | term_in_name
        
        name_count = name_mask.sum()
        logging.info(f"DisplayName matching: {name_count} GPOs")
        print(f"  DisplayName matching: {name_count} GPOs")
        
        # SECONDARY: Match by Links column (catches GPOs linked to operation OUs)
        # These are GPOs that APPLY to the operation but don't have the operation code in their name
        # e.g., "No Logon Banner" linked to OU=Foxtrot should be reviewed for Foxtrot
        links_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
        
        if 'Links' in enterprise_gpos.columns:
            for term in search_terms:
                # Only check terms 4+ chars to avoid false positives in Links text
                if len(term) >= 4:
                    term_in_links = enterprise_gpos['Links'].str.contains(
                        term, case=False, na=False, regex=False
                    )
                    links_mask = links_mask | term_in_links
        
        # GPOs that are linked to operation OUs but NOT matched by name
        # These need manual review - they apply to this operation but aren't named for it
        links_only_mask = links_mask & ~name_mask
        links_only_count = links_only_mask.sum()
        
        if links_only_count > 0:
            logging.info(f"Links-only matching (for review): {links_only_count} GPOs")
            print(f"  Linked to operation OUs (review required): {links_only_count} GPOs")
        
        # Store the links-only GPOs separately for the Review Required tab
        self.links_only_gpos = enterprise_gpos[links_only_mask].copy()
        
        # Mark them with a flag for the report
        if not self.links_only_gpos.empty:
            self.links_only_gpos['ReviewReason'] = 'Linked to ' + operation_name + ' OU but name does not contain operation code'
        
        # Combine both for report_gpos (name matches + links-only)
        combined_mask = name_mask | links_only_mask
        self.report_gpos = enterprise_gpos[combined_mask].copy()
        
        # Add a column to identify which GPOs need review
        # Use the original index to properly assign MatchType
        self.report_gpos['MatchType'] = 'Name Match'
        # GPOs that matched by links but not by name need review
        links_only_indices = enterprise_gpos[links_only_mask].index
        self.report_gpos.loc[self.report_gpos.index.isin(links_only_indices), 'MatchType'] = 'Links Only - Review Required'
        
        after_operation_filter = len(self.report_gpos)
        logging.info(f"After operation filter (combined): {after_operation_filter} GPOs")
        
        # Log each GPO that matched the operation
        if logging.getLogger().isEnabledFor(logging.DEBUG):
            for _, gpo in self.report_gpos.iterrows():
                logging.debug(f"  Matched: {gpo.get('DisplayName')} (MatchType: {gpo.get('MatchType')})")
        
        # Filter out ENT GPOs from operation results
        ent_mask = self.report_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
        if ent_mask.any():
            ent_count = ent_mask.sum()
            # Log which GPOs are being filtered as ENT
            ent_gpos = self.report_gpos[ent_mask]['DisplayName'].tolist()
            logging.info(f"Filtering {ent_count} ENT GPOs: {ent_gpos}")
            print(f"  Filtering out {ent_count} ENT GPOs from operation results:")
            for gpo_name in ent_gpos:
                print(f"    - {gpo_name}")
            self.report_gpos = self.report_gpos[~ent_mask].copy()
            
            # Also filter ENT from links_only_gpos
            if not self.links_only_gpos.empty:
                ent_links_mask = self.links_only_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
                self.links_only_gpos = self.links_only_gpos[~ent_links_mask].copy()
        
        after_count = len(self.report_gpos)
        filtered_count = before_count - after_count
        
        # Count by match type
        name_match_count = (self.report_gpos['MatchType'] == 'Name Match').sum() if 'MatchType' in self.report_gpos.columns else after_count
        review_count = (self.report_gpos['MatchType'] == 'Links Only - Review Required').sum() if 'MatchType' in self.report_gpos.columns else 0
        
        print(f"\n  === Operation Filter Summary ===")
        print(f"  Enterprise/Shared Forest GPOs total: {before_count}")
        print(f"  Name matches: {name_match_count}")
        print(f"  Links-only (review required): {review_count}")
        print(f"  Total for report (after ENT filter): {after_count}")
        print(f"  Excluded from report: {filtered_count}")
        
        if after_count == 0:
            # Show what operations WERE detected to help troubleshoot
            detected_ops = enterprise_gpos['DetectedOperation'].value_counts().head(10)
            print(f"\n  WARNING: No GPOs found for operation '{operation_name}'")
            print(f"  Detected operations in baseline.corp:")
            for op, count in detected_ops.items():
                print(f"    - {op}: {count} GPOs")
            print(f"\n  TIP: Use one of the detected operation names above (e.g., --operation 'Foxtrot')")
        else:
            # Show bucket breakdown for this operation
            bucket_counts = self.report_gpos['Bucket'].value_counts()
            print(f"\n  Bucket breakdown for {operation_name}:")
            for bucket, count in bucket_counts.items():
                print(f"    {bucket}: {count}")
        
        # Filter all_settings to only this operation's settings
        if self.all_settings:
            settings_before = len(self.all_settings)
            operation_gpo_names = set(self.report_gpos['DisplayName'].values)
            self.all_settings = [
                s for s in self.all_settings 
                if s.get('GPOName') in operation_gpo_names
            ]
            settings_after = len(self.all_settings)
            print(f"  Filtered settings: {settings_before} -> {settings_after}")
    
    def _generate_domain_report(self, output_path: Path):
        """Generate single-domain or single-operation action plan (5 tabs - BUCKET FOCUSED)"""
        logging.info(f"Generating domain report: operation={self.operation}, domain={self.target_domain}")
        
        # For operation mode, skip domain verification (operations live in operations_gpos)
        if not self.operation:
            # Regular domain mode: verify domain exists
            if self.target_domain not in self.domains:
                print(f"ERROR: Domain '{self.target_domain}' not found in reports")
                print(f"Available domains: {', '.join(self.domains.keys())}")
                logging.error(f"Domain '{self.target_domain}' not found")
                sys.exit(1)
        
        # If operation specified, filter to only that operation's GPOs
        if self.operation:
            # Handle both LOCATION_MAPPING codes and direct operation names
            if self.operation in LOCATION_MAPPING:
                display_name = f"{LOCATION_MAPPING[self.operation]['full_name']} ({self.operation})"
            else:
                # Direct operation name (e.g., "Foxtrot", "Golf")
                display_name = self.operation
            print(f"\nFiltering to operation: {display_name}")
            self._filter_to_operation()
            # report_gpos is now set by _filter_to_operation()
            logging.info(f"After operation filter: report_gpos has {len(self.report_gpos)} rows")
        else:
            # Regular domain mode - filter domain_gpos to target domain into report_gpos
            # domain_gpos already has ENT filtered out from _filter_and_track_ent()
            # 
            # NOTE: For standalone domains (echo.corp.com, corp.alpha.com, etc.),
            # ALL GPOs in that domain belong to that operation - no Links-only matching needed.
            # We still add MatchType column for consistency with operation mode reports.
            self.report_gpos = self.domain_gpos[self.domain_gpos['Domain'] == self.target_domain].copy()
            self.report_gpos['MatchType'] = 'Domain GPO'  # All GPOs in standalone domain are direct matches
            display_name = self.target_domain
            logging.info(f"Domain mode: filtered domain_gpos to {self.target_domain}, got {len(self.report_gpos)} rows")
            
            # DIAGNOSTIC: Show bucket distribution before ENT filter
            if not self.report_gpos.empty and 'Bucket' in self.report_gpos.columns:
                bucket_dist = self.report_gpos['Bucket'].value_counts().to_dict()
                logging.info(f"Bucket distribution for {self.target_domain} (before ENT filter): {bucket_dist}")
                print(f"  Bucket distribution: {bucket_dist}")
            
            # Also filter out ENT from report_gpos (safety check)
            ent_mask = self.report_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
            if ent_mask.any():
                ent_count = ent_mask.sum()
                self.report_gpos = self.report_gpos[~ent_mask].copy()
                print(f"  Filtered {ent_count} ENT GPOs from {display_name}")
                
                # DIAGNOSTIC: Show bucket distribution after ENT filter
                if not self.report_gpos.empty and 'Bucket' in self.report_gpos.columns:
                    bucket_dist_after = self.report_gpos['Bucket'].value_counts().to_dict()
                    logging.info(f"Bucket distribution for {self.target_domain} (after ENT filter): {bucket_dist_after}")
                    print(f"  Bucket distribution after ENT filter: {bucket_dist_after}")
        
        # Detect single-domain mode and warn user
        domain_count = len(self.domains)
        is_single_domain = domain_count == 1
        
        # For operations, don't show single-domain warning (they're inherently single-operation)
        if is_single_domain and not self.operation:
            print(f"\n{'='*70}")
            print(f"  SINGLE DOMAIN MODE DETECTED")
            print(f"{'='*70}")
            print(f"Running with only 1 domain HTML file.")
            print(f"{'='*70}\n")
        
        writer = pd.ExcelWriter(output_path, engine='openpyxl')
        
        print(f"  - Creating Tab 1: Bucket Overview for {display_name}...")
        self._create_domain_tab1_bucket_overview(writer, display_name)
        
        print(f"  - Creating Tab 2: Server GPOs...")
        self._create_domain_tab2_server_gpos(writer)
        
        print(f"  - Creating Tab 3: Workstation GPOs...")
        self._create_domain_tab3_workstation_gpos(writer)
        
        print(f"  - Creating Tab 4: User GPOs...")
        self._create_domain_tab4_user_gpos(writer)
        
        print(f"  - Creating Tab 5: Review Required...")
        self._create_domain_tab5_review_required(writer)
        
        writer.close()
        print(f"\nâœ“ Domain report generated for {display_name}: {output_path}")

    def _create_executive_summary(self, writer):
        """Create Executive Summary with operations broken out"""
        print("  - Creating Executive Summary (with operations breakdown)...")

        summary_data = []

        # First, add operations from Shared Forest (if they exist)
        if not self.operations_gpos.empty:
            print("    Breaking out Shared Forest operations...")
            for code, info in LOCATION_MAPPING.items():
                ops_with_location = self.operations_gpos[
                    self.operations_gpos['LocationCodes'].apply(lambda x: code in x if x else False)
                ]
                if len(ops_with_location) > 0:
                    operation_name = f"{code} (Shared Forest)"
                    gpo_count = len(ops_with_location)
                    
                    # Count settings for this operation if available
                    settings_count = 0
                    if operation_name in self.operation_settings:
                        settings_count = len(self.operation_settings[operation_name])
                    
                    summary_data.append({
                        'Operation/Domain': operation_name,
                        'Full Name': info['full_name'],
                        'Active GPOs': gpo_count,
                        'Settings': settings_count,
                        'Owning Domain': 'baseline.corp/baseline.corp',
                        'Type': 'Operation'
                    })
        
        # Add shared ENT GPOs
        if not self.enterprise_standard_gpos.empty:
            summary_data.append({
                'Operation/Domain': 'Enterprise Standards (Shared)',
                'Full Name': 'Enterprise IT Standards',
                'Active GPOs': len(self.enterprise_standard_gpos),
                'Settings': 'N/A',
                'Owning Domain': 'baseline.corp/baseline.corp',
                'Type': 'Enterprise Standard'
            })
        
        if not self.ent_nonstandard_gpos.empty:
            summary_data.append({
                'Operation/Domain': 'ENT Non-Standard (Shared)',
                'Full Name': 'Non-Standard ENT Naming',
                'Active GPOs': len(self.ent_nonstandard_gpos),
                'Settings': 'N/A',
                'Owning Domain': 'baseline.corp/baseline.corp',
                'Type': 'Naming Issue'
            })

        # Add all other domains
        for domain, data in self.domains.items():
            if domain in ['baseline.corp', 'baseline.corp']:
                continue  # Skip - already handled above
            
            if 'summary' not in data or data['summary'].empty:
                continue

            total = len(data['summary'])
            active = len(self.active_gpos[self.active_gpos['Domain'] == domain])
            unique_count = len(self.domain_gpos[self.domain_gpos['Domain'] == domain])
            
            # Count settings
            settings_count = 0
            if domain in self.operation_settings:
                settings_count = len(self.operation_settings[domain])

            summary_data.append({
                'Operation/Domain': domain,
                'Full Name': domain,
                'Active GPOs': active,
                'Settings': settings_count,
                'Owning Domain': domain,
                'Type': 'Standalone Domain'
            })

        df = pd.DataFrame(summary_data)
        
        # Sort: Operations first, then standalone domains
        if not df.empty:
            df['SortOrder'] = df['Type'].map({
                'Operation': 1,
                'Enterprise Standard': 2,
                'Naming Issue': 3,
                'Standalone Domain': 4
            })
            df = df.sort_values(['SortOrder', 'Operation/Domain'])
            df = df.drop(columns=['SortOrder'])
        
        df.to_excel(writer, sheet_name='Executive Summary', index=False)
        print(f"    Created summary with {len(df)} entries")

    def _create_shared_gpos_inventory(self, writer):
        """Create Shared Enterprise GPOs inventory"""
        print("  - Creating Shared Enterprise GPOs (552 counted once)...")

        if self.shared_gpos.empty:
            df = pd.DataFrame({'Message': ['No shared GPOs detected']})
            df.to_excel(writer, sheet_name='Shared Enterprise GPOs', index=False)
            return

        # Determine GPO Type for shared GPOs
        def get_gpo_type(name):
            if pd.isna(name):
                return 'Unknown'
            if str(name).startswith('ENT - '):
                return 'Enterprise Standard'
            elif str(name).upper().startswith('ENT'):
                return 'ENT Non-Standard'
            else:
                return 'Operations'

        shared_display = self.shared_gpos.copy()
        shared_display.insert(1, 'GPO Type', shared_display['DisplayName'].apply(get_gpo_type))

        # Select relevant columns
        cols = ['DisplayName', 'GPO Type', 'GUID', 'LinksCount']
        for col in ['WMIFilter', 'Description', 'CreationTime']:
            if col in shared_display.columns:
                cols.append(col)

        shared_display = shared_display[cols]
        shared_display = shared_display.sort_values(['GPO Type', 'DisplayName'])

        shared_display.to_excel(writer, sheet_name='Shared Enterprise GPOs', index=False)
        print(f"    Listed {len(shared_display)} shared GPOs")

    def _create_operations_by_location(self, writer):
        """Create Operations GPOs grouped by location"""
        print("  - Creating Operations by Location...")

        if self.operations_gpos.empty:
            df = pd.DataFrame({'Message': ['No operations GPOs found']})
            df.to_excel(writer, sheet_name='Operations by Location', index=False)
            return

        ops_display = self.operations_gpos[['DisplayName', 'LocationCodesStr', 'GUID', 'LinksCount']].copy()
        ops_display = ops_display.sort_values(['LocationCodesStr', 'DisplayName'])

        # Add owning domain column
        def get_owning_domain(loc_str):
            if pd.isna(loc_str) or loc_str == 'Unknown':
                return 'baseline.corp/baseline.corp'
            codes = loc_str.split(', ')
            if codes:
                first_code = codes[0]
                target = LOCATION_MAPPING.get(first_code, {}).get('target_domain', 'baseline.corp')
                # Show current state
                return f'baseline.corp/baseline.corp -> {target}'
            return 'baseline.corp/baseline.corp'

        ops_display.insert(2, 'Owning Domain', ops_display['LocationCodesStr'].apply(get_owning_domain))

        ops_display.columns = ['GPO Name', 'Location Codes', 'Owning Domain', 'GUID', 'Links Count']
        ops_display.to_excel(writer, sheet_name='Operations by Location', index=False)
        print(f"    Listed {len(ops_display)} operations GPOs")

    def _create_ent_inventory(self, writer):
        """Create Enterprise Standard GPOs inventory"""
        print("  - Creating Enterprise Standard GPOs Inventory...")

        if self.enterprise_standard_gpos.empty:
            df = pd.DataFrame({'Message': ['No standard ENT GPOs found']})
            df.to_excel(writer, sheet_name='Enterprise Standard GPOs', index=False)
            return

        cols = ['DisplayName', 'GUID', 'LinksCount']
        for col in ['WMIFilter', 'Description', 'CreationTime']:
            if col in self.enterprise_standard_gpos.columns:
                cols.append(col)

        ent_display = self.enterprise_standard_gpos[cols].copy()
        ent_display = ent_display.sort_values('DisplayName')

        ent_display.to_excel(writer, sheet_name='Enterprise Standard GPOs', index=False)
        print(f"    Listed {len(ent_display)} standard ENT GPOs")

    def _create_nonstandard_ent(self, writer):
        """Create Non-Standard ENT Names tab"""
        print("  - Creating Non-Standard ENT Names...")

        if self.ent_nonstandard_gpos.empty:
            df = pd.DataFrame({'Message': ['All ENT GPOs follow standard naming convention']})
            df.to_excel(writer, sheet_name='Non-Standard ENT Names', index=False)
            return

        def get_naming_issue(name):
            if pd.isna(name):
                return 'Missing name'
            name_upper = str(name).upper()
            if name_upper.startswith('ENT-'):
                return 'Uses dash (ENT-) instead of space-dash-space (ENT - )'
            elif name_upper.startswith('ENT_'):
                return 'Uses underscore (ENT_) instead of space-dash-space (ENT - )'
            elif name_upper.startswith('ENT ') and not str(name).startswith('ENT - '):
                return 'Uses space only (ENT ) instead of space-dash-space (ENT - )'
            else:
                return 'Non-standard ENT prefix'

        nonstandard = self.ent_nonstandard_gpos[['DisplayName', 'GUID']].copy()
        nonstandard.insert(1, 'Naming Issue', nonstandard['DisplayName'].apply(get_naming_issue))
        nonstandard = nonstandard.sort_values(['Naming Issue', 'DisplayName'])

        nonstandard.to_excel(writer, sheet_name='Non-Standard ENT Names', index=False)
        print(f"    âš ï¸  Flagged {len(nonstandard)} GPOs with non-standard ENT naming")

    def _create_setting_analysis(self, writer):
        """
        Create setting analysis tabs with enhanced drill-down
        NEW v1.9.1: Adds Domains/Operations column and Sample GPOs column
        """
        print("  - Creating Setting Analysis (Domain-Unique GPOs Only)...")

        if not self.all_settings:
            print("    No domain-unique settings to analyze")
            return

        settings_df = pd.DataFrame(self.all_settings)

        # Group by setting and get counts
        setting_counts = settings_df.groupby(['SettingName', 'SettingValue']).agg({
            'Domain': lambda x: len(set(x))
        }).reset_index()
        setting_counts.columns = ['Setting Name', 'Setting Value', 'Domain Count']
        
        # NEW v1.9.1: Add associations
        def get_entities_for_setting(row):
            """Get list of entities that have this setting"""
            key = (row['Setting Name'], str(row['Setting Value']))
            if key in self.setting_to_entities:
                entities = sorted(list(self.setting_to_entities[key]))
                return ', '.join(entities)
            return ''
        
        def get_sample_gpos(row):
            """Get sample GPO names (up to 3) that contain this setting"""
            key = (row['Setting Name'], str(row['Setting Value']))
            if key in self.setting_to_gpos:
                gpos = sorted(list(self.setting_to_gpos[key]))
                # Return first 3 GPOs
                return ', '.join(gpos[:3]) + ('...' if len(gpos) > 3 else '')
            return ''
        
        setting_counts['Domains/Operations'] = setting_counts.apply(get_entities_for_setting, axis=1)
        setting_counts['GPO Names'] = setting_counts.apply(get_sample_gpos, axis=1)
        setting_counts = setting_counts.sort_values('Domain Count', ascending=False)

        # Calculate thresholds based on NON-shared domains
        total_domains = len([d for d in self.domains.keys() if d not in ['baseline.corp', 'baseline.corp']])

        # Enterprise Candidates (7+ domains or 66%+)
        threshold_enterprise = max(7, int(total_domains * 0.66))
        enterprise = setting_counts[setting_counts['Domain Count'] >= threshold_enterprise].copy()
        if not enterprise.empty:
            enterprise.to_excel(writer, sheet_name='Enterprise Candidates', index=False)
            print(f"    {len(enterprise)} enterprise candidate settings")

        # Inconsistent Settings (3-6 domains or 33-66%)
        threshold_low = max(3, int(total_domains * 0.33))
        inconsistent = setting_counts[
            (setting_counts['Domain Count'] >= threshold_low) &
            (setting_counts['Domain Count'] < threshold_enterprise)
        ].copy()
        if not inconsistent.empty:
            inconsistent.to_excel(writer, sheet_name='Inconsistent Settings', index=False)
            print(f"    {len(inconsistent)} inconsistent settings")

        # Domain-Specific Outliers (1-2 domains or <33%)
        outliers = setting_counts[setting_counts['Domain Count'] < threshold_low].copy()
        if not outliers.empty:
            outliers.to_excel(writer, sheet_name='Domain-Specific Outliers', index=False)
            print(f"    {len(outliers)} domain-specific outliers")

    def _create_performance_analysis(self, writer):
        """
        Create Performance Issues tab
        NEW v1.9.1: Break out operations like Executive Summary
        """
        print("  - Creating Performance Analysis (with operations breakdown)...")

        perf_issues = []

        # First, add operations from shared GPOs if they exist
        if not self.operations_gpos.empty:
            print("    Analyzing performance for operations...")
            for code, info in LOCATION_MAPPING.items():
                ops_with_location = self.operations_gpos[
                    self.operations_gpos['LocationCodes'].apply(lambda x: code in x if x else False)
                ]
                if len(ops_with_location) > 0:
                    operation_name = f"{code} (Shared Forest)"
                    
                    # Count metrics for this operation
                    wmi_count = ops_with_location['WMIFilter'].notna().sum() if 'WMIFilter' in ops_with_location.columns else 0
                    empty_count = (ops_with_location['Empty'].astype(str).str.lower() == 'true').sum() if 'Empty' in ops_with_location.columns else 0
                    problem_count = ops_with_location['Problem'].notna().sum() if 'Problem' in ops_with_location.columns else 0
                    
                    perf_issues.append({
                        'Operation/Domain': operation_name,
                        'Full Name': info['full_name'],
                        'WMI Filters': wmi_count,
                        'Empty GPOs': empty_count,
                        'Config Issues': problem_count,
                        'Total GPOs': len(ops_with_location),
                        'Type': 'Operation'
                    })
        
        # Add shared ENT GPOs performance
        if not self.enterprise_standard_gpos.empty:
            wmi_count = self.enterprise_standard_gpos['WMIFilter'].notna().sum() if 'WMIFilter' in self.enterprise_standard_gpos.columns else 0
            empty_count = (self.enterprise_standard_gpos['Empty'].astype(str).str.lower() == 'true').sum() if 'Empty' in self.enterprise_standard_gpos.columns else 0
            problem_count = self.enterprise_standard_gpos['Problem'].notna().sum() if 'Problem' in self.enterprise_standard_gpos.columns else 0
            
            perf_issues.append({
                'Operation/Domain': 'Enterprise Standards (Shared)',
                'Full Name': 'Enterprise IT Standards',
                'WMI Filters': wmi_count,
                'Empty GPOs': empty_count,
                'Config Issues': problem_count,
                'Total GPOs': len(self.enterprise_standard_gpos),
                'Type': 'Enterprise Standard'
            })
        
        if not self.ent_nonstandard_gpos.empty:
            wmi_count = self.ent_nonstandard_gpos['WMIFilter'].notna().sum() if 'WMIFilter' in self.ent_nonstandard_gpos.columns else 0
            empty_count = (self.ent_nonstandard_gpos['Empty'].astype(str).str.lower() == 'true').sum() if 'Empty' in self.ent_nonstandard_gpos.columns else 0
            problem_count = self.ent_nonstandard_gpos['Problem'].notna().sum() if 'Problem' in self.ent_nonstandard_gpos.columns else 0
            
            perf_issues.append({
                'Operation/Domain': 'ENT Non-Standard (Shared)',
                'Full Name': 'Non-Standard ENT Naming',
                'WMI Filters': wmi_count,
                'Empty GPOs': empty_count,
                'Config Issues': problem_count,
                'Total GPOs': len(self.ent_nonstandard_gpos),
                'Type': 'Naming Issue'
            })

        # Add standalone domains
        for domain, data in self.domains.items():
            if domain in ['baseline.corp', 'baseline.corp']:
                continue  # Already handled above
                
            if 'summary' not in data or data['summary'].empty:
                continue

            df = data['summary']
            active_df = self.active_gpos[self.active_gpos['Domain'] == domain]

            # Count metrics
            wmi_count = df['WMIFilter'].notna().sum() if 'WMIFilter' in df.columns else 0
            empty_count = (df['Empty'].astype(str).str.lower() == 'true').sum() if 'Empty' in df.columns else 0
            problem_count = df['Problem'].notna().sum() if 'Problem' in df.columns else 0

            perf_issues.append({
                'Operation/Domain': domain,
                'Full Name': domain,
                'WMI Filters': wmi_count,
                'Empty GPOs': empty_count,
                'Config Issues': problem_count,
                'Total GPOs': len(active_df),
                'Type': 'Standalone Domain'
            })

        if perf_issues:
            perf_df = pd.DataFrame(perf_issues)
            # Sort like Executive Summary
            perf_df['SortOrder'] = perf_df['Type'].map({
                'Operation': 1,
                'Enterprise Standard': 2,
                'Naming Issue': 3,
                'Standalone Domain': 4
            })
            perf_df = perf_df.sort_values(['SortOrder', 'Operation/Domain'])
            perf_df = perf_df.drop(columns=['SortOrder', 'Type'])
            perf_df.to_excel(writer, sheet_name='Performance Issues', index=False)
            print(f"    Analyzed performance for {len(perf_df)} entities")

    def _create_migration_gotchas(self, writer):
        """Create Migration Gotchas tab"""
        print("  - Creating Migration Gotchas...")

        gotchas = []

        unc_pattern = r'\\\\[\w\.-]+\\[\w\.$-]+'
        ip_pattern = r'\b(?:\d{1,3}\.){3}\d{1,3}\b'

        for setting in self.all_settings:
            value = str(setting.get('SettingValue', ''))

            unc_matches = re.findall(unc_pattern, value)
            if unc_matches:
                gotchas.append({
                    'Type': 'UNC Path',
                    'Domain': setting['Domain'],
                    'GPO': setting['GPOName'],
                    'Setting': setting['SettingName'],
                    'Value': value
                })

            ip_matches = re.findall(ip_pattern, value)
            if ip_matches:
                gotchas.append({
                    'Type': 'IP Address',
                    'Domain': setting['Domain'],
                    'GPO': setting['GPOName'],
                    'Setting': setting['SettingName'],
                    'Value': value
                })

        if gotchas:
            gotchas_df = pd.DataFrame(gotchas)
            gotchas_df.to_excel(writer, sheet_name='Migration Gotchas', index=False)

    def _create_security_analysis(self, writer):
        """Create Security & Permissions tab"""
        print("  - Creating Security Analysis...")

        security_issues = []

        for domain, data in self.domains.items():
            if 'permissions' not in data or data['permissions'].empty:
                continue

            df = data['permissions']
            df['Domain'] = domain
            security_issues.append(df)

        if security_issues:
            combined = pd.concat(security_issues, ignore_index=True)
            combined.to_excel(writer, sheet_name='Security & Permissions', index=False)

    def _create_link_analysis(self, writer):
        """Create OU Link Analysis tab"""
        print("  - Creating Link Analysis...")

        link_data = []

        for domain, data in self.domains.items():
            if 'links' not in data or data['links'].empty:
                continue

            df = data['links'].copy()
            df['Domain'] = domain
            link_data.append(df)

        if link_data:
            combined = pd.concat(link_data, ignore_index=True)

            # Add GPO Type classification
            if 'DisplayName' in combined.columns:
                def get_gpo_type(name, domain):
                    if pd.isna(name):
                        return 'Unknown'
                    if domain in ['baseline.corp', 'baseline.corp']:
                        if str(name).startswith('ENT - '):
                            return 'Enterprise Standard (Shared)'
                        elif str(name).upper().startswith('ENT'):
                            return 'ENT Non-Standard (Shared)'
                        else:
                            return 'Operations (Shared)'
                    else:
                        return 'Domain-Unique'

                combined.insert(1, 'GPO Type', combined.apply(lambda row: get_gpo_type(row['DisplayName'], row['Domain']), axis=1))

            combined.to_excel(writer, sheet_name='OU Link Analysis', index=False)

    def _create_migration_roadmap(self, writer):
        """Create Migration Roadmap for operations GPOs"""
        print("  - Creating Migration Roadmap...")

        if self.operations_gpos.empty:
            df = pd.DataFrame({'Message': ['No operations GPOs to migrate']})
            df.to_excel(writer, sheet_name='Migration Roadmap', index=False)
            return

        # Group by location
        location_summary = []

        for code, info in LOCATION_MAPPING.items():
            # Count GPOs for this location
            gpos_with_location = self.operations_gpos[
                self.operations_gpos['LocationCodes'].apply(lambda x: code in x)
            ]

            if len(gpos_with_location) > 0:
                location_summary.append({
                    'Location Code': code,
                    'Location Name': info['full_name'],
                    'Target Domain': info['target_domain'],
                    'GPO Count': len(gpos_with_location),
                    'Status': 'Pending Migration'
                })

        if location_summary:
            roadmap_df = pd.DataFrame(location_summary)
            roadmap_df = roadmap_df.sort_values('GPO Count', ascending=False)
            roadmap_df.to_excel(writer, sheet_name='Migration Roadmap', index=False)
            print(f"    Created roadmap for {len(location_summary)} locations")

    def _create_consolidation_recommendations(self, writer):
        """
        Create Consolidation Recommendations tab - Enhanced in v1.9.1
        Now splits "Unique" into "Domain-Only" and "Shared-Unique"
        """
        print("  - Creating Consolidation Recommendations...")
        
        if not self.all_settings or not self.enterprise_standards:
            df = pd.DataFrame({'Message': ['Run analysis first to generate recommendations']})
            df.to_excel(writer, sheet_name='Consolidation Recommendations', index=False)
            return
        
        recommendations = []
        settings_df = pd.DataFrame(self.all_settings)
        
        # For each operation/domain, generate enhanced recommendations
        for entity in self.operation_settings.keys():
            entity_settings = self.operation_settings[entity]
            
            if entity_settings.empty:
                continue
            
            # Get unique settings for this entity
            entity_setting_pairs = set(
                entity_settings.apply(lambda x: f"{x['SettingName']}={x['SettingValue']}", axis=1)
            )
            
            # Count settings this entity has that are enterprise standards
            match_enterprise = 0
            for std in self.enterprise_standards:
                std_pair = f"{std['SettingName']}={std['SettingValue']}"
                if std_pair in entity_setting_pairs:
                    match_enterprise += 1
            
            # NEW v1.9.1: Distinguish domain-only from shared-unique
            domain_only_count = 0
            shared_unique_count = 0
            
            for setting_pair in entity_setting_pairs:
                # Skip enterprise standards (already counted)
                is_enterprise = any(
                    f"{std['SettingName']}={std['SettingValue']}" == setting_pair
                    for std in self.enterprise_standards
                )
                if is_enterprise:
                    continue
                
                # Check how many entities have this setting
                # Parse back the setting
                parts = setting_pair.split('=', 1)
                if len(parts) == 2:
                    setting_name, setting_value = parts
                    key = (setting_name, setting_value)
                    
                    if key in self.setting_to_entities:
                        entity_count = len(self.setting_to_entities[key])
                        if entity_count == 1:
                            domain_only_count += 1
                        else:
                            shared_unique_count += 1
                    else:
                        domain_only_count += 1  # Assume unique if not tracked
            
            recommendations.append({
                'Operation/Domain': entity,
                'Total Settings': len(entity_setting_pairs),
                'Match Enterprise': match_enterprise,
                'Domain-Only': domain_only_count,
                'Shared-Unique': shared_unique_count,
                'Recommendation': 'Review for standardization' if match_enterprise > 5 else 'Mostly unique'
            })
        
        if recommendations:
            rec_df = pd.DataFrame(recommendations)
            rec_df = rec_df.sort_values('Match Enterprise', ascending=False)
            rec_df.to_excel(writer, sheet_name='Consolidation Recommendations', index=False)
            print(f"    Created recommendations for {len(recommendations)} entities")

    def _create_gap_analysis(self, writer):
        """Create Gap Analysis tab - what's each entity missing?"""
        print("  - Creating Gap Analysis...")
        
        if not self.enterprise_standards:
            df = pd.DataFrame({'Message': ['No enterprise standards identified']})
            df.to_excel(writer, sheet_name='Gap Analysis', index=False)
            return
        
        gap_data = []
        settings_df = pd.DataFrame(self.all_settings)
        
        # For each operation/domain
        for entity in self.operation_settings.keys():
            entity_settings = self.operation_settings[entity]
            
            if entity_settings.empty:
                continue
            
            # Get what they have
            entity_setting_pairs = set(
                entity_settings.apply(lambda x: f"{x['SettingName']}={x['SettingValue']}", axis=1)
            )
            
            # Check what enterprise standards they're missing
            missing_count = 0
            for std in self.enterprise_standards:
                std_pair = f"{std['SettingName']}={std['SettingValue']}"
                if std_pair not in entity_setting_pairs:
                    missing_count += 1
                    if missing_count <= 10:  # Show first 10 gaps
                        gap_data.append({
                            'Operation/Domain': entity,
                            'Missing Setting': std['SettingName'],
                            'Expected Value': std['SettingValue'],
                            'Found In': f"{std['EntityCount']} other entities",
                            'Priority': 'High' if std['EntityCount'] >= 10 else 'Medium'
                        })
        
        if gap_data:
            gap_df = pd.DataFrame(gap_data)
            gap_df = gap_df.sort_values(['Operation/Domain', 'Priority'])
            gap_df.to_excel(writer, sheet_name='Gap Analysis', index=False)
            print(f"    Identified {len(gap_df)} gaps across entities")

    def _create_settings_breakdown(self, writer):
        """
        NEW v1.9.1: Settings Breakdown by Entity
        Shows actual settings per domain/operation with overlap information
        """
        print("  - Creating Settings Breakdown by Entity...")
        
        if not self.entity_setting_details:
            df = pd.DataFrame({'Message': ['No setting details available']})
            df.to_excel(writer, sheet_name='Settings Breakdown', index=False)
            return
        
        breakdown_df = pd.DataFrame(self.entity_setting_details)
        
        # Add "Shared With" column - shows which other entities have this setting
        def get_shared_with(row):
            key = (row['SettingName'], row['SettingValue'])
            if key in self.setting_to_entities:
                entities = self.setting_to_entities[key]
                # Exclude current entity
                others = [e for e in entities if e != row['Entity']]
                if others:
                    return ', '.join(sorted(others)[:5]) + ('...' if len(others) > 5 else '')
                return 'Domain-Only'
            return 'Domain-Only'
        
        breakdown_df['Shared With'] = breakdown_df.apply(get_shared_with, axis=1)
        
        # Reorder columns for better readability
        columns = ['Entity', 'SettingName', 'SettingValue', 'Category', 'GPOName', 'Shared With']
        breakdown_df = breakdown_df[columns]
        breakdown_df = breakdown_df.sort_values(['Entity', 'SettingName'])
        
        breakdown_df.to_excel(writer, sheet_name='Settings Breakdown', index=False)
        print(f"    Created breakdown with {len(breakdown_df)} setting records")

    def _create_settings_detail_view(self, writer):
        """
        NEW v1.9.1: Settings Detail View - Expanded format for pivot analysis
        One row per entity per setting - enables Excel filtering and pivot tables
        """
        print("  - Creating Settings Detail View (pivot-friendly)...")
        
        if not self.entity_setting_details:
            df = pd.DataFrame({'Message': ['No setting details available']})
            df.to_excel(writer, sheet_name='Settings Detail View', index=False)
            return
        
        detail_df = pd.DataFrame(self.entity_setting_details)
        
        # Add domain count for each setting
        def get_domain_count(row):
            key = (row['SettingName'], row['SettingValue'])
            if key in self.setting_to_entities:
                return len(self.setting_to_entities[key])
            return 1
        
        detail_df['Domain Count'] = detail_df.apply(get_domain_count, axis=1)
        
        # Add classification
        def classify_setting(domain_count):
            if domain_count >= 8:
                return 'Enterprise Standard'
            elif domain_count >= 3:
                return 'Shared Across Multiple'
            elif domain_count == 2:
                return 'Shared Between Two'
            else:
                return 'Domain-Only'
        
        detail_df['Classification'] = detail_df['Domain Count'].apply(classify_setting)
        
        # Reorder and sort
        columns = ['Entity', 'SettingName', 'SettingValue', 'Category', 'GPOName', 
                   'Domain Count', 'Classification', 'SettingKey']
        detail_df = detail_df[columns]
        detail_df = detail_df.sort_values(['Classification', 'Domain Count', 'SettingName'], 
                                           ascending=[True, False, True])
        
        detail_df.to_excel(writer, sheet_name='Settings Detail View', index=False)
        print(f"    Created detail view with {len(detail_df)} records (pivot-ready)")

    def _create_raw_data(self, writer):
        """Create Raw Data tabs"""
        print("  - Creating Raw Data...")

        # Shared GPOs
        if not self.shared_gpos.empty:
            self.shared_gpos.to_excel(writer, sheet_name='Raw - Shared GPOs', index=False)

        # Domain-Unique GPOs
        if not self.domain_gpos.empty:
            self.domain_gpos.to_excel(writer, sheet_name='Raw - Domain-Unique GPOs', index=False)

    # ==========================================
    # EXECUTIVE MODE TAB METHODS (7 TABS)
    # ==========================================
    
    def _create_exec_tab1_migration_dashboard(self, writer):
        """Executive Tab 1: Migration Dashboard with bucket metrics"""
        from datetime import datetime
        
        # Calculate high-level metrics
        total_domains = len(self.domains)
        total_gpos_all = self.total_gpos_before_filter
        active_gpos_count = len(self.active_gpos) if not self.active_gpos.empty else 0
        
        # Bucket breakdown across ALL active GPOs
        bucket_counts = {}
        if not self.active_gpos.empty and 'Bucket' in self.active_gpos.columns:
            bucket_counts = self.active_gpos['Bucket'].value_counts().to_dict()
        
        server_count = bucket_counts.get('Server', 0)
        workstation_count = bucket_counts.get('Workstation', 0)
        user_count = bucket_counts.get('User', 0)
        dc_count = bucket_counts.get('Domain Controller', 0)
        domain_root_count = bucket_counts.get('Domain Root', 0)
        mixed_count = bucket_counts.get('Mixed', 0)
        unknown_count = bucket_counts.get('Unknown', 0)
        
        # Calculate migration readiness
        migrate_ready = server_count + workstation_count  # Clean buckets
        needs_review = user_count + mixed_count + unknown_count
        dont_migrate = dc_count + domain_root_count
        
        readiness_pct = (migrate_ready / active_gpos_count * 100) if active_gpos_count > 0 else 0
        
        # Count Shared Forest operations vs standalone domains
        shared_forest_ops = [code for code in LOCATION_MAPPING.keys() if LOCATION_MAPPING[code].get('source_domain') == 'baseline.corp']
        standalone_domains = [d for d in self.domains.keys() if d not in ['baseline.corp', 'baseline.corp']]
        
        # Create dashboard narrative
        dashboard = f"""GPO MIGRATION DASHBOARD
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
ENVIRONMENT OVERVIEW
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
Total Forests/Domains Analyzed:     {total_domains}
  â€¢ Standalone Domains:             {len(standalone_domains)}
  â€¢ Shared Forest Operations:          {len(shared_forest_ops)} (in baseline.corp)

Total GPOs Discovered:              {total_gpos_all:,}
Active GPOs (Enabled + Linked):     {active_gpos_count:,} ({active_gpos_count/total_gpos_all*100:.1f}%)

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
BUCKET DISTRIBUTION (All Active GPOs)
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
Server GPOs:                        {server_count:,}  â†’ MIGRATE (High Priority)
Workstation GPOs:                   {workstation_count:,}  â†’ MIGRATE (High Priority)
User GPOs:                          {user_count:,}  â†’ REVIEW (Optimization Candidates)
Domain Controller GPOs:             {dc_count:,}  â†’ DON'T MIGRATE
Domain Root GPOs:                   {domain_root_count:,}  â†’ DON'T MIGRATE
Mixed (Multi-Bucket):               {mixed_count:,}  â†’ MANUAL REVIEW
Unknown:                            {unknown_count:,}  â†’ MANUAL REVIEW

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
MIGRATION READINESS
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
Ready to Migrate:                   {migrate_ready:,} GPOs ({readiness_pct:.1f}%)
  (Server + Workstation buckets - clean categorization)

Needs Review:                       {needs_review:,} GPOs
  (User + Mixed + Unknown - require decisions)

Don't Migrate:                      {dont_migrate:,} GPOs
  (Domain Controller + Domain Root - infrastructure)

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
KEY METRICS FOR PLANNING
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
â€¢ Migration Scope:                  {migrate_ready + needs_review:,} GPOs (excluding infrastructure)
â€¢ Clean Migration (no review):      {readiness_pct:.1f}% of active GPOs
â€¢ Review Backlog:                   {needs_review:,} GPOs requiring decisions
â€¢ Infrastructure Exclusions:        {dont_migrate:,} GPOs (stay in current domain)

â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
NEXT STEPS
â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•â•
1. Review Tab 2 (Operations Summary) for per-operation breakdown
2. Review Tab 3 (Risk Assessment) to identify complex operations
3. Address Mixed/Unknown GPOs before migration
4. Plan User GPO optimization (consolidate where possible)
"""
        
        # Write to Excel
        df = pd.DataFrame({'Migration Dashboard': [dashboard]})
        df.to_excel(writer, sheet_name='1. Migration Dashboard', index=False, header=False)
        
        # Adjust column width
        worksheet = writer.sheets['1. Migration Dashboard']
        worksheet.column_dimensions['A'].width = 100
    
    def _create_exec_tab2_operations_summary(self, writer):
        """Executive Tab 2: Operations Summary with bucket breakdown per operation"""
        
        summary_data = []
        
        # Process Shared Forest operations
        if not self.active_gpos.empty:
            enterprise_gpos = self.active_gpos[
                self.active_gpos['Domain'].isin(['baseline.corp', 'baseline.corp'])
            ].copy()
            
            # Skip legacy alias codes to avoid duplicates
            for code, info in LOCATION_MAPPING.items():
                if info.get('source_domain') != 'baseline.corp':
                    continue
                
                # Filter GPOs for this operation using DisplayName matching
                search_terms = [code.lower(), info['full_name'].lower()]
                if 'name_prefixes' in info:
                    search_terms.extend([p.lower() for p in info['name_prefixes']])
                
                # Match by DisplayName
                op_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
                for term in search_terms:
                    if len(term) >= 2:
                        if len(term) <= 3:
                            pattern = r'(^|[\s\-_])' + re.escape(term) + r'($|[\s\-_])'
                            term_match = enterprise_gpos['DisplayName'].str.contains(pattern, case=False, na=False, regex=True)
                        else:
                            term_match = enterprise_gpos['DisplayName'].str.contains(term, case=False, na=False, regex=False)
                        op_mask = op_mask | term_match
                
                op_gpos = enterprise_gpos[op_mask].copy()
                
                # Filter out ENT GPOs (same as Web does)
                if not op_gpos.empty:
                    ent_mask = op_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
                    op_gpos = op_gpos[~ent_mask]
                
                if len(op_gpos) == 0:
                    continue
                
                # Get bucket distribution
                bucket_dist = op_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in op_gpos.columns else {}
                
                server = bucket_dist.get('Server', 0)
                workstation = bucket_dist.get('Workstation', 0)
                user = bucket_dist.get('User', 0)
                dc = bucket_dist.get('Domain Controller', 0)
                domain_root = bucket_dist.get('Domain Root', 0)
                mixed = bucket_dist.get('Mixed', 0)
                unknown = bucket_dist.get('Unknown', 0)
                
                migrate_ready = server + workstation
                needs_review = user + mixed + unknown
                dont_migrate = dc + domain_root
                
                summary_data.append({
                    'Operation': f"{code} - {info['full_name']}",
                    'Type': 'Shared Forest',
                    'Total GPOs': len(op_gpos),
                    'Server': server,
                    'Workstation': workstation,
                    'User': user,
                    'Mixed': mixed,
                    'Unknown': unknown,
                    'DC': dc,
                    'Domain Root': domain_root,
                    'Migrate Ready': migrate_ready,
                    'Needs Review': needs_review,
                    'Dont Migrate': dont_migrate,
                    'Readiness %': f"{migrate_ready/len(op_gpos)*100:.0f}%" if len(op_gpos) > 0 else "N/A"
                })
        
        # Process standalone domains
        for domain in self.domains.keys():
            if domain in ['baseline.corp', 'baseline.corp']:
                continue
            
            domain_gpos = self.active_gpos[self.active_gpos['Domain'] == domain].copy()
            
            # Filter out ENT GPOs (same as Web does)
            if not domain_gpos.empty:
                ent_mask = domain_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
                domain_gpos = domain_gpos[~ent_mask]
            
            if len(domain_gpos) == 0:
                continue
            
            bucket_dist = domain_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in domain_gpos.columns else {}
            
            server = bucket_dist.get('Server', 0)
            workstation = bucket_dist.get('Workstation', 0)
            user = bucket_dist.get('User', 0)
            dc = bucket_dist.get('Domain Controller', 0)
            domain_root = bucket_dist.get('Domain Root', 0)
            mixed = bucket_dist.get('Mixed', 0)
            unknown = bucket_dist.get('Unknown', 0)
            
            migrate_ready = server + workstation
            needs_review = user + mixed + unknown
            dont_migrate = dc + domain_root
            
            summary_data.append({
                'Operation': domain,
                'Type': 'Standalone',
                'Total GPOs': len(domain_gpos),
                'Server': server,
                'Workstation': workstation,
                'User': user,
                'Mixed': mixed,
                'Unknown': unknown,
                'DC': dc,
                'Domain Root': domain_root,
                'Migrate Ready': migrate_ready,
                'Needs Review': needs_review,
                'Dont Migrate': dont_migrate,
                'Readiness %': f"{migrate_ready/len(domain_gpos)*100:.0f}%" if len(domain_gpos) > 0 else "N/A"
            })
        
        if summary_data:
            df = pd.DataFrame(summary_data)
            df = df.sort_values('Total GPOs', ascending=False)
            df.to_excel(writer, sheet_name='2. Operations Summary', index=False)
        else:
            pd.DataFrame({'Note': ['No operations data available']}).to_excel(
                writer, sheet_name='2. Operations Summary', index=False)
    
    def _create_exec_tab3_risk_assessment(self, writer):
        """Executive Tab 3: Risk Assessment - operations ranked by complexity"""
        
        risk_data = []
        
        # Process Shared Forest operations
        if not self.active_gpos.empty:
            enterprise_gpos = self.active_gpos[
                self.active_gpos['Domain'].isin(['baseline.corp', 'baseline.corp'])
            ].copy()
            
            for code, info in LOCATION_MAPPING.items():
                if info.get('source_domain') != 'baseline.corp':
                    continue
                
                # Filter GPOs for this operation
                search_terms = [code.lower(), info['full_name'].lower()]
                if 'name_prefixes' in info:
                    search_terms.extend([p.lower() for p in info['name_prefixes']])
                
                op_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
                for term in search_terms:
                    if len(term) >= 2:
                        if len(term) <= 3:
                            pattern = r'(^|[\s\-_])' + re.escape(term) + r'($|[\s\-_])'
                            term_match = enterprise_gpos['DisplayName'].str.contains(pattern, case=False, na=False, regex=True)
                        else:
                            term_match = enterprise_gpos['DisplayName'].str.contains(term, case=False, na=False, regex=False)
                        op_mask = op_mask | term_match
                
                op_gpos = enterprise_gpos[op_mask]
                
                if len(op_gpos) == 0:
                    continue
                
                bucket_dist = op_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in op_gpos.columns else {}
                
                mixed = bucket_dist.get('Mixed', 0)
                unknown = bucket_dist.get('Unknown', 0)
                dc = bucket_dist.get('Domain Controller', 0)
                domain_root = bucket_dist.get('Domain Root', 0)
                
                # Calculate risk score (higher = more complex)
                total = len(op_gpos)
                risk_score = (mixed * 3) + (unknown * 2) + dc + domain_root
                risk_pct = (mixed + unknown) / total * 100 if total > 0 else 0
                
                # Determine risk level
                if risk_pct > 30:
                    risk_level = "HIGH"
                elif risk_pct > 15:
                    risk_level = "MEDIUM"
                else:
                    risk_level = "LOW"
                
                risk_data.append({
                    'Operation': f"{code} - {info['full_name']}",
                    'Type': 'Shared Forest',
                    'Total GPOs': total,
                    'Mixed GPOs': mixed,
                    'Unknown GPOs': unknown,
                    'DC GPOs': dc,
                    'Domain Root': domain_root,
                    'Review Required %': f"{risk_pct:.1f}%",
                    'Risk Level': risk_level,
                    'Risk Score': risk_score,
                    'Notes': self._get_risk_notes(mixed, unknown, dc, domain_root)
                })
        
        # Process standalone domains
        for domain in self.domains.keys():
            if domain in ['baseline.corp', 'baseline.corp']:
                continue
            
            domain_gpos = self.active_gpos[self.active_gpos['Domain'] == domain]
            
            if len(domain_gpos) == 0:
                continue
            
            bucket_dist = domain_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in domain_gpos.columns else {}
            
            mixed = bucket_dist.get('Mixed', 0)
            unknown = bucket_dist.get('Unknown', 0)
            dc = bucket_dist.get('Domain Controller', 0)
            domain_root = bucket_dist.get('Domain Root', 0)
            
            total = len(domain_gpos)
            risk_score = (mixed * 3) + (unknown * 2) + dc + domain_root
            risk_pct = (mixed + unknown) / total * 100 if total > 0 else 0
            
            if risk_pct > 30:
                risk_level = "HIGH"
            elif risk_pct > 15:
                risk_level = "MEDIUM"
            else:
                risk_level = "LOW"
            
            risk_data.append({
                'Operation': domain,
                'Type': 'Standalone',
                'Total GPOs': total,
                'Mixed GPOs': mixed,
                'Unknown GPOs': unknown,
                'DC GPOs': dc,
                'Domain Root': domain_root,
                'Review Required %': f"{risk_pct:.1f}%",
                'Risk Level': risk_level,
                'Risk Score': risk_score,
                'Notes': self._get_risk_notes(mixed, unknown, dc, domain_root)
            })
        
        if risk_data:
            df = pd.DataFrame(risk_data)
            df = df.sort_values('Risk Score', ascending=False)
            df.to_excel(writer, sheet_name='3. Risk Assessment', index=False)
        else:
            pd.DataFrame({'Note': ['No risk data available']}).to_excel(
                writer, sheet_name='3. Risk Assessment', index=False)
    
    def _get_risk_notes(self, mixed: int, unknown: int, dc: int, domain_root: int) -> str:
        """Generate risk notes based on bucket counts"""
        notes = []
        if mixed > 5:
            notes.append(f"{mixed} Mixed GPOs need splitting")
        if unknown > 3:
            notes.append(f"{unknown} Unknown GPOs need classification")
        if dc > 0:
            notes.append(f"{dc} DC GPOs - verify no migration needed")
        if domain_root > 0:
            notes.append(f"{domain_root} Domain Root GPOs - infrastructure review")
        return "; ".join(notes) if notes else "Low complexity"
    
    def _create_exec_tab4_bucket_comparison(self, writer):
        """Executive Tab 4: Bucket Comparison across operations"""
        
        comparison_data = []
        
        # Process Shared Forest operations
        if not self.active_gpos.empty:
            enterprise_gpos = self.active_gpos[
                self.active_gpos['Domain'].isin(['baseline.corp', 'baseline.corp'])
            ].copy()
            
            for code, info in LOCATION_MAPPING.items():
                if info.get('source_domain') != 'baseline.corp':
                    continue

                search_terms = [code.lower(), info['full_name'].lower()]
                if 'name_prefixes' in info:
                    search_terms.extend([p.lower() for p in info['name_prefixes']])
                
                op_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
                for term in search_terms:
                    if len(term) >= 2:
                        if len(term) <= 3:
                            pattern = r'(^|[\s\-_])' + re.escape(term) + r'($|[\s\-_])'
                            term_match = enterprise_gpos['DisplayName'].str.contains(pattern, case=False, na=False, regex=True)
                        else:
                            term_match = enterprise_gpos['DisplayName'].str.contains(term, case=False, na=False, regex=False)
                        op_mask = op_mask | term_match
                
                op_gpos = enterprise_gpos[op_mask]
                
                if len(op_gpos) == 0:
                    continue
                
                total = len(op_gpos)
                bucket_dist = op_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in op_gpos.columns else {}
                
                comparison_data.append({
                    'Operation': code,
                    'Full Name': info['full_name'],
                    'Total': total,
                    'Server %': f"{bucket_dist.get('Server', 0)/total*100:.0f}%" if total > 0 else "0%",
                    'Workstation %': f"{bucket_dist.get('Workstation', 0)/total*100:.0f}%" if total > 0 else "0%",
                    'User %': f"{bucket_dist.get('User', 0)/total*100:.0f}%" if total > 0 else "0%",
                    'Mixed %': f"{bucket_dist.get('Mixed', 0)/total*100:.0f}%" if total > 0 else "0%",
                    'Other %': f"{(bucket_dist.get('Unknown', 0) + bucket_dist.get('Domain Controller', 0) + bucket_dist.get('Domain Root', 0))/total*100:.0f}%" if total > 0 else "0%"
                })
        
        # Process standalone domains
        for domain in self.domains.keys():
            if domain in ['baseline.corp', 'baseline.corp']:
                continue
            
            domain_gpos = self.active_gpos[self.active_gpos['Domain'] == domain]
            
            if len(domain_gpos) == 0:
                continue
            
            total = len(domain_gpos)
            bucket_dist = domain_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in domain_gpos.columns else {}
            
            # Find short code for domain
            short_code = domain.split('.')[0].upper()[:4]
            
            comparison_data.append({
                'Operation': short_code,
                'Full Name': domain,
                'Total': total,
                'Server %': f"{bucket_dist.get('Server', 0)/total*100:.0f}%" if total > 0 else "0%",
                'Workstation %': f"{bucket_dist.get('Workstation', 0)/total*100:.0f}%" if total > 0 else "0%",
                'User %': f"{bucket_dist.get('User', 0)/total*100:.0f}%" if total > 0 else "0%",
                'Mixed %': f"{bucket_dist.get('Mixed', 0)/total*100:.0f}%" if total > 0 else "0%",
                'Other %': f"{(bucket_dist.get('Unknown', 0) + bucket_dist.get('Domain Controller', 0) + bucket_dist.get('Domain Root', 0))/total*100:.0f}%" if total > 0 else "0%"
            })
        
        if comparison_data:
            df = pd.DataFrame(comparison_data)
            df = df.sort_values('Total', ascending=False)
            df.to_excel(writer, sheet_name='4. Bucket Comparison', index=False)
        else:
            pd.DataFrame({'Note': ['No comparison data available']}).to_excel(
                writer, sheet_name='4. Bucket Comparison', index=False)
    
    # Keep old tab functions for backward compatibility but mark as deprecated
    def _create_exec_tab1_status(self, writer):
        """Executive Tab 1: Project Status with narrative summary"""
        from datetime import datetime
        
        # Calculate metrics
        total_domains = len(self.domains)
        non_shared_domains = [d for d in self.domains.keys() if d not in ['baseline.corp', 'baseline.corp']]
        total_gpos_all = self.total_gpos_before_filter
        active_gpos_count = len(self.active_gpos) if not self.active_gpos.empty else 0
        inactive_gpos_count = total_gpos_all - active_gpos_count
        domain_unique_count = len(self.domain_gpos) if not self.domain_gpos.empty else 0
        
        # Settings overlap analysis
        settings_df = pd.DataFrame(self.all_settings) if self.all_settings else pd.DataFrame()
        
        if not settings_df.empty:
            setting_counts = settings_df.groupby(['SettingName', 'SettingValue']).agg({
                'Domain': lambda x: len(set(x))
            }).reset_index()
            
            threshold_high = max(8, int(len(non_shared_domains) * 0.66))
            threshold_med = max(3, int(len(non_shared_domains) * 0.33))
            
            high_overlap = len(setting_counts[setting_counts['Domain'] >= threshold_high])
            medium_overlap = len(setting_counts[(setting_counts['Domain'] >= threshold_med) & (setting_counts['Domain'] < threshold_high)])
            low_overlap = len(setting_counts[setting_counts['Domain'] < threshold_med])
            total_unique = len(setting_counts)
        else:
            high_overlap = medium_overlap = low_overlap = total_unique = 0
            threshold_high = 8
        
        # Create narrative
        narrative = f"""CONSOLIDATION PROJECT STATUS
Technical Summary - Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}

ENVIRONMENT INVENTORY:
â€¢ Total Domains: {total_domains}
â€¢ Total GPOs Identified: {total_gpos_all}
â€¢ Active GPOs (Enabled + Linked): {active_gpos_count} ({active_gpos_count/total_gpos_all*100:.1f}% if total_gpos_all > 0 else 0)
â€¢ Inactive GPOs: {inactive_gpos_count} ({inactive_gpos_count/total_gpos_all*100:.1f}% if total_gpos_all > 0 else 0)

NON-SHARED DOMAINS ANALYZED: {len(non_shared_domains)}
(Excludes baseline.corp and baseline.corp which serve as shared repositories)

Analyzed domains: {', '.join(non_shared_domains)}

SETTING OVERLAP ANALYSIS:
â€¢ High Overlap ({threshold_high}+ domains): {high_overlap} settings -> Enterprise candidates
â€¢ Medium Overlap (3-{threshold_high-1} domains): {medium_overlap} settings -> Standardization decisions
â€¢ Low Overlap (1-2 domains): {low_overlap} settings -> Domain-specific
â€¢ Total Unique Settings Analyzed: {total_unique}

CONSOLIDATION METRICS:
â€¢ Settings Ready for Standardization: {high_overlap} (high overlap)
â€¢ Settings Requiring Decisions: {medium_overlap} (medium overlap)
â€¢ Settings Remaining Domain-Specific: {low_overlap}

NEXT ACTIONS:
Priority 1: Review {high_overlap} high-overlap settings for enterprise standardization
Priority 2: Resolve {medium_overlap} medium-overlap conflicts  
Priority 3: Domain-specific review of {low_overlap} settings
"""
        
        # Write to Excel
        df = pd.DataFrame({'Project Status': [narrative]})
        df.to_excel(writer, sheet_name='1. Project Status', index=False, header=False)
    
    def _create_exec_tab2_standardization(self, writer):
        """Executive Tab 2: Standardization Candidates (high overlap)"""
        if not self.all_settings:
            pd.DataFrame({'Note': ['No settings data available']}).to_excel(writer, sheet_name='2. Standardization Candidates', index=False)
            return
        
        settings_df = pd.DataFrame(self.all_settings)
        non_shared_domains = [d for d in self.domains.keys() if d not in ['baseline.corp', 'baseline.corp']]
        threshold_high = max(8, int(len(non_shared_domains) * 0.66))
        
        setting_counts = settings_df.groupby(['SettingName', 'SettingValue']).agg({
            'Domain': lambda x: list(set(x))
        }).reset_index()
        setting_counts['Domain Count'] = setting_counts['Domain'].apply(len)
        setting_counts['Domains'] = setting_counts['Domain'].apply(lambda x: ', '.join(sorted(x)))
        
        high_overlap_settings = setting_counts[setting_counts['Domain Count'] >= threshold_high].copy()
        high_overlap_settings = high_overlap_settings.sort_values('Domain Count', ascending=False)
        high_overlap_settings = high_overlap_settings[['SettingName', 'SettingValue', 'Domain Count', 'Domains']]
        
        if high_overlap_settings.empty:
            pd.DataFrame({'Note': [f'No settings found in {threshold_high}+ domains']}).to_excel(writer, sheet_name='2. Standardization Candidates', index=False)
        else:
            high_overlap_settings.to_excel(writer, sheet_name='2. Standardization Candidates', index=False)
    
    def _create_exec_tab3_conflicts(self, writer):
        """Executive Tab 3: Conflict Resolution (medium overlap)"""
        if not self.all_settings:
            pd.DataFrame({'Note': ['No settings data available']}).to_excel(writer, sheet_name='3. Conflict Resolution', index=False)
            return
        
        settings_df = pd.DataFrame(self.all_settings)
        non_shared_domains = [d for d in self.domains.keys() if d not in ['baseline.corp', 'baseline.corp']]
        threshold_high = max(8, int(len(non_shared_domains) * 0.66))
        threshold_med = max(3, int(len(non_shared_domains) * 0.33))
        
        setting_counts = settings_df.groupby(['SettingName', 'SettingValue']).agg({
            'Domain': lambda x: list(set(x))
        }).reset_index()
        setting_counts['Domain Count'] = setting_counts['Domain'].apply(len)
        setting_counts['Domains'] = setting_counts['Domain'].apply(lambda x: ', '.join(sorted(x)))
        
        medium_overlap_settings = setting_counts[
            (setting_counts['Domain Count'] >= threshold_med) &
            (setting_counts['Domain Count'] < threshold_high)
        ].copy()
        medium_overlap_settings = medium_overlap_settings.sort_values('Domain Count', ascending=False)
        medium_overlap_settings = medium_overlap_settings[['SettingName', 'SettingValue', 'Domain Count', 'Domains']]
        
        if medium_overlap_settings.empty:
            pd.DataFrame({'Note': [f'No settings found in {threshold_med}-{threshold_high-1} domains']}).to_excel(writer, sheet_name='3. Conflict Resolution', index=False)
        else:
            medium_overlap_settings.to_excel(writer, sheet_name='3. Conflict Resolution', index=False)
    
    def _create_exec_tab4_distribution(self, writer):
        """Executive Tab 4: Work Distribution by Domain"""
        if not self.all_settings:
            pd.DataFrame({'Note': ['No settings data available']}).to_excel(writer, sheet_name='4. Work Distribution', index=False)
            return
        
        settings_df = pd.DataFrame(self.all_settings)
        non_shared_domains = [d for d in self.domains.keys() if d not in ['baseline.corp', 'baseline.corp']]
        
        distribution_data = []
        for domain in non_shared_domains:
            domain_settings = settings_df[settings_df['Domain'] == domain]
            total_settings = len(domain_settings)
            
            # Count by overlap
            setting_counts = {}
            for _, row in domain_settings.iterrows():
                key = (row['SettingName'], row['SettingValue'])
                if key not in setting_counts:
                    # Count how many domains have this setting
                    count = len(settings_df[(settings_df['SettingName'] == row['SettingName']) & 
                                           (settings_df['SettingValue'] == row['SettingValue'])]['Domain'].unique())
                    setting_counts[key] = count
            
            high_overlap = sum(1 for c in setting_counts.values() if c >= 8)
            medium_overlap = sum(1 for c in setting_counts.values() if 3 <= c < 8)
            low_overlap = sum(1 for c in setting_counts.values() if c < 3)
            
            distribution_data.append({
                'Domain': domain,
                'Total Settings': total_settings,
                'High Overlap (8+)': high_overlap,
                'Medium Overlap (3-7)': medium_overlap,
                'Low Overlap (1-2)': low_overlap,
                'Unique to Domain': low_overlap
            })
        
        df = pd.DataFrame(distribution_data).sort_values('Unique to Domain', ascending=False)
        df.to_excel(writer, sheet_name='4. Work Distribution', index=False)
    
    def _create_exec_tab5_performance(self, writer):
        """Executive Tab 5: Performance Issues by Operation"""
        # Reuse existing performance analysis but organize by operation
        self._create_performance_analysis(writer)
        # Rename the sheet
        wb = writer.book
        if 'Performance Issues' in wb.sheetnames:
            ws = wb['Performance Issues']
            ws.title = '5. Performance Issues'
    
    def _create_exec_tab6_infrastructure(self, writer):
        """Executive Tab 6: Infrastructure Dependencies"""
        # Reuse existing migration gotchas
        self._create_migration_gotchas(writer)
        # Rename the sheet
        wb = writer.book
        if 'Migration Gotchas' in wb.sheetnames:
            ws = wb['Migration Gotchas']
            ws.title = '6. Infrastructure Dependencies'
    
    def _create_exec_tab7_roadmap(self, writer):
        """Executive Tab 7: Consolidation Roadmap"""
        roadmap_text = """CONSOLIDATION ROADMAP
4-Phase Implementation Plan (20 weeks)

PHASE 1: FOUNDATION (Weeks 1-4)
Objective: Establish enterprise standards and clean up performance issues
Tasks:
  â€¢ Standardize high-overlap settings (8+ domains) -> 40 hours
  â€¢ Remove empty GPOs and address WMI filters -> 20 hours
  â€¢ Document infrastructure dependencies (UNC paths, IPs) -> 16 hours
Deliverables: Enterprise GPO templates, Performance cleanup report
Success Metric: High-overlap settings reduced by 80%

PHASE 2: ALIGNMENT (Weeks 5-8)
Objective: Resolve conflicts and establish sub-standards
Tasks:
  â€¢ Resolve medium-overlap conflicts (3-7 domains) -> 60 hours
  â€¢ Establish operation-specific sub-standards -> 40 hours
  â€¢ Create migration playbooks per domain -> 24 hours
Deliverables: Conflict resolution decisions, Sub-standard GPOs
Success Metric: Medium-overlap conflicts reduced by 60%

PHASE 3: DOMAIN-SPECIFIC REVIEW (Weeks 9-16)
Objective: Review and categorize domain-unique settings
Tasks:
  â€¢ Domain teams review unique settings -> 120 hours (distributed)
  â€¢ Update infrastructure references -> 40 hours
  â€¢ Create keep/migrate/retire documentation -> 32 hours
Deliverables: Domain-specific decisions documented
Success Metric: All unique settings categorized

PHASE 4: MIGRATION & VALIDATION (Weeks 17-20)
Objective: Execute migration and validate
Tasks:
  â€¢ Pilot migration (2 domains) -> 32 hours
  â€¢ Full migration rollout -> 48 hours
  â€¢ Validation and remediation -> 40 hours
Deliverables: Consolidated GPO structure, Validation report
Success Metric: All domains migrated, <5% issues

TOTAL ESTIMATED EFFORT: ~512 hours over 20 weeks
CRITICAL PATH: Phase 1 -> Phase 2 (sequential), Phase 3 (parallel by domain)
"""
        
        df = pd.DataFrame({'Consolidation Roadmap': [roadmap_text]})
        df.to_excel(writer, sheet_name='7. Consolidation Roadmap', index=False, header=False)
    
    # ==========================================
    # DOMAIN MODE TAB METHODS (5 TABS - BUCKET FOCUSED)
    # v2.3.2: Reorganized by bucket type for migration decisions
    # ==========================================
    
    def _create_domain_tab1_bucket_overview(self, writer, display_name: str):
        """
        Domain Tab 1: Bucket Overview
        
        Clean, scannable format with bucket summary and tab navigation.
        Links-only GPOs are counted separately and go to Review Required.
        """
        from datetime import datetime
        
        # Get GPOs for this domain/operation (use report_gpos which is filtered for domain mode)
        gpos = self.report_gpos if not self.report_gpos.empty else pd.DataFrame()
        
        if gpos.empty:
            narrative = f"""GPO BUCKET OVERVIEW: {display_name.upper()}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Version: v2.3.2

No GPOs found for {display_name}.
"""
            df = pd.DataFrame({'Bucket Overview': [narrative]})
            df.to_excel(writer, sheet_name='1. Bucket Overview', index=False, header=False)
            return
        
        total_gpos = len(gpos)
        
        # Separate Name Match vs Links-only GPOs
        if 'MatchType' in gpos.columns:
            name_match_gpos = gpos[gpos['MatchType'] != 'Links Only - Review Required']
            links_only_gpos = gpos[gpos['MatchType'] == 'Links Only - Review Required']
            links_only_count = len(links_only_gpos)
        else:
            name_match_gpos = gpos
            links_only_gpos = pd.DataFrame()
            links_only_count = 0
        
        # Count by bucket type (only for Name Match GPOs - these go to bucket tabs)
        bucket_counts = name_match_gpos['Bucket'].value_counts().to_dict() if 'Bucket' in name_match_gpos.columns else {}
        
        server_count = bucket_counts.get('Server', 0)
        workstation_count = bucket_counts.get('Workstation', 0)
        user_count = bucket_counts.get('User', 0)
        domain_root_count = bucket_counts.get('Domain Root', 0)
        dc_count = bucket_counts.get('Domain Controller', 0)
        mixed_count = bucket_counts.get('Mixed', 0)
        unknown_count = bucket_counts.get('Unknown', 0)
        
        # Count by AppliesTo (all GPOs)
        applies_counts = gpos['AppliesTo'].value_counts().to_dict() if 'AppliesTo' in gpos.columns else {}
        computer_only = applies_counts.get('Computer', 0)
        user_only = applies_counts.get('User', 0)
        both_config = applies_counts.get('Both', 0)
        
        # Count mismatches (all GPOs)
        mismatch_count = gpos['ConfigMismatch'].sum() if 'ConfigMismatch' in gpos.columns else 0
        
        migrate_count = server_count + workstation_count
        dont_migrate_count = domain_root_count + dc_count
        review_bucket_count = mixed_count + unknown_count + user_count
        review_total = review_bucket_count + links_only_count
        
        # Build Links-only section if applicable
        links_only_section = ""
        if links_only_count > 0:
            links_only_section = f"""
================================================================================
LINKS-ONLY GPOs (Review Required)                           â†’ Tab 5
================================================================================
  GPOs linked to {display_name} OUs but without
  operation code in name:  {links_only_count:4}
  
  These GPOs APPLY to {display_name} but aren't named for it.
  Examples: "No Logon Banner", "Server - Power Settings"
  ACTION: Review to confirm they should migrate with {display_name}.
"""
        
        narrative = f"""GPO BUCKET OVERVIEW: {display_name.upper()}
================================================================================
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}
Version: v2.3.2

Total GPOs for {display_name}: {total_gpos}
  - Name Match GPOs: {len(name_match_gpos)} (in bucket tabs 2-4)
  - Links-Only GPOs: {links_only_count} (in Review Required tab 5)

================================================================================
BUCKET SUMMARY (Name Match GPOs)                            TAB
================================================================================
  Server:              {server_count:4}                                   â†’ Tab 2
  Workstation:         {workstation_count:4}                                   â†’ Tab 3
  User:                {user_count:4}                                   â†’ Tab 4
  Domain Root:         {domain_root_count:4}                                   â†’ Tab 5
  Domain Controller:   {dc_count:4}                                   â†’ Tab 5
  Mixed:               {mixed_count:4}                                   â†’ Tab 5
  Unknown:             {unknown_count:4}                                   â†’ Tab 5
{links_only_section}
================================================================================
CONFIGURATION TARGET (What the GPO configures)
================================================================================
  Computer settings only:   {computer_only:4}
  User settings only:       {user_only:4}
  Both Computer & User:     {both_config:4}

  Config Mismatches:        {mismatch_count:4}  (GPO target doesn't match OU type)

================================================================================
MIGRATION SUMMARY
================================================================================
  MIGRATE:       {migrate_count:4}  (Server + Workstation)
  DON'T MIGRATE: {dont_migrate_count:4}  (Domain Root + DC - infrastructure)
  REVIEW:        {review_total:4}  (User + Mixed + Unknown + Links-Only)

================================================================================
QUICK REFERENCE
================================================================================

BUCKET TYPES:
  Server         = Linked to OU containing "Server" or "Servers"
  Workstation    = Linked to OU containing "Computer", "Workstation", etc.
  User           = Linked to OU containing "User" or "Users"
  Domain Root    = Linked at domain level (no OU)
  Domain Controller = Linked to Domain Controllers OU
  Mixed          = Linked to multiple bucket types
  Unknown        = Cannot determine from OU structure

APPLIES TO:
  Computer       = GPO has Computer Configuration enabled
  User           = GPO has User Configuration enabled
  Both           = Both configurations enabled

CONFIG MISMATCH:
  A mismatch occurs when:
  - Computer-only GPO linked to User OU, or
  - User-only GPO linked to Computer OU (Server/Workstation)
  These are optimization candidates during consolidation.

================================================================================
MIGRATION STRATEGY
================================================================================
  1. Migrate Server GPOs (Tab 2) - Critical infrastructure
  2. Migrate Workstation GPOs (Tab 3) - End-user computing
  3. Review User GPOs (Tab 4) - Consider converting to Workstation
  4. Review Mixed/Unknown (Tab 5) - Manual decision required
  5. Leave Domain Root/DC GPOs - Infrastructure stays with domain
"""
        
        df = pd.DataFrame({'Bucket Overview': [narrative]})
        df.to_excel(writer, sheet_name='1. Bucket Overview', index=False, header=False)
    
    def _create_domain_tab2_server_gpos(self, writer):
        """
        Domain Tab 2: Server GPOs
        
        GPOs linked to Server OUs - HIGH PRIORITY for migration.
        Only includes Name Match GPOs - Links-only GPOs go to Review Required tab.
        """
        gpos = self.report_gpos if not self.report_gpos.empty else pd.DataFrame()
        
        # Diagnostic logging
        logging.info(f"Tab 2 Server GPOs - report_gpos count: {len(gpos)}")
        if not gpos.empty and 'Bucket' in gpos.columns:
            bucket_counts = gpos['Bucket'].value_counts().to_dict()
            logging.info(f"Tab 2 Server GPOs - bucket distribution: {bucket_counts}")
        
        if gpos.empty or 'Bucket' not in gpos.columns:
            logging.warning("Tab 2: No GPOs or no Bucket column")
            pd.DataFrame({'Note': ['No Server GPOs found']}).to_excel(
                writer, sheet_name='2. Server GPOs', index=False)
            return
        
        server_gpos = gpos[gpos['Bucket'] == 'Server'].copy()
        
        # Exclude Links-only GPOs - they go to Review Required tab
        if 'MatchType' in server_gpos.columns:
            server_gpos = server_gpos[server_gpos['MatchType'] != 'Links Only - Review Required']
        
        logging.info(f"Tab 2: Found {len(server_gpos)} Server bucket GPOs (after excluding Links-only)")
        
        if server_gpos.empty:
            pd.DataFrame({'Note': ['No Server GPOs found for this domain/operation']}).to_excel(
                writer, sheet_name='2. Server GPOs', index=False)
            return
        
        # Build inventory
        inventory = []
        for _, gpo in server_gpos.iterrows():
            gpo_name = gpo.get('DisplayName', 'Unknown')
            description = gpo.get('Description', '')
            if pd.isna(description):
                description = ''
            
            # Format links for readable display
            links = format_links_for_display(gpo.get('Links', ''))
            
            # Count settings
            gpo_settings = [s for s in self.all_settings if s.get('GPOName') == gpo_name]
            
            # Get AppliesTo and ConfigMismatch
            applies_to = gpo.get('AppliesTo', 'Unknown')
            config_mismatch = gpo.get('ConfigMismatch', False)
            
            # Determine migration priority based on mismatch
            if config_mismatch:
                priority = 'HIGH - REVIEW (Config Mismatch)'
            else:
                priority = 'HIGH'
            
            inventory.append({
                'GPO Name': gpo_name,
                'Description': description,
                'Linked To': links,
                'Applies To': applies_to,
                'Settings Count': len(gpo_settings),
                'Size (MB)': gpo.get('SizeMB', 'N/A'),
                'Created': gpo.get('CreationTime', 'N/A'),
                'Last Modified': gpo.get('ModificationTime', 'N/A'),
                'GUID': gpo.get('GUID', ''),
                'Migration Priority': priority
            })
        
        df = pd.DataFrame(inventory)
        df.to_excel(writer, sheet_name='2. Server GPOs', index=False)
        
        # Format the Linked To column for text wrap
        worksheet = writer.sheets['2. Server GPOs']
        # Set column width and text wrap for Linked To column (C)
        worksheet.column_dimensions['C'].width = 60
        for row in range(2, len(df) + 2):  # Skip header row
            cell = worksheet.cell(row=row, column=3)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_domain_tab3_workstation_gpos(self, writer):
        """
        Domain Tab 3: Workstation GPOs
        
        GPOs linked to Computer/Workstation OUs - HIGH PRIORITY for migration.
        Only includes Name Match GPOs - Links-only GPOs go to Review Required tab.
        """
        gpos = self.report_gpos if not self.report_gpos.empty else pd.DataFrame()
        
        if gpos.empty or 'Bucket' not in gpos.columns:
            pd.DataFrame({'Note': ['No Workstation GPOs found']}).to_excel(
                writer, sheet_name='3. Workstation GPOs', index=False)
            return
        
        workstation_gpos = gpos[gpos['Bucket'] == 'Workstation'].copy()
        
        # Exclude Links-only GPOs - they go to Review Required tab
        if 'MatchType' in workstation_gpos.columns:
            workstation_gpos = workstation_gpos[workstation_gpos['MatchType'] != 'Links Only - Review Required']
        
        if workstation_gpos.empty:
            pd.DataFrame({'Note': ['No Workstation GPOs found for this domain/operation']}).to_excel(
                writer, sheet_name='3. Workstation GPOs', index=False)
            return
        
        # Build inventory
        inventory = []
        for _, gpo in workstation_gpos.iterrows():
            gpo_name = gpo.get('DisplayName', 'Unknown')
            description = gpo.get('Description', '')
            if pd.isna(description):
                description = ''
            
            # Format links for readable display
            links = format_links_for_display(gpo.get('Links', ''))
            
            # Count settings
            gpo_settings = [s for s in self.all_settings if s.get('GPOName') == gpo_name]
            
            # Get AppliesTo and ConfigMismatch
            applies_to = gpo.get('AppliesTo', 'Unknown')
            config_mismatch = gpo.get('ConfigMismatch', False)
            
            # Determine migration priority based on mismatch
            if config_mismatch:
                priority = 'HIGH - REVIEW (Config Mismatch)'
            else:
                priority = 'HIGH'
            
            inventory.append({
                'GPO Name': gpo_name,
                'Description': description,
                'Linked To': links,
                'Applies To': applies_to,
                'Settings Count': len(gpo_settings),
                'Size (MB)': gpo.get('SizeMB', 'N/A'),
                'Created': gpo.get('CreationTime', 'N/A'),
                'Last Modified': gpo.get('ModificationTime', 'N/A'),
                'GUID': gpo.get('GUID', ''),
                'Migration Priority': priority
            })
        
        df = pd.DataFrame(inventory)
        df.to_excel(writer, sheet_name='3. Workstation GPOs', index=False)
        
        # Format the Linked To column for text wrap
        worksheet = writer.sheets['3. Workstation GPOs']
        worksheet.column_dimensions['C'].width = 60
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=3)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_domain_tab4_user_gpos(self, writer):
        """
        Domain Tab 4: User GPOs
        
        GPOs linked to User OUs - REVIEW for optimization.
        Consider converting to Workstation GPOs during consolidation.
        Only includes Name Match GPOs - Links-only GPOs go to Review Required tab.
        """
        gpos = self.report_gpos if not self.report_gpos.empty else pd.DataFrame()
        
        if gpos.empty or 'Bucket' not in gpos.columns:
            pd.DataFrame({'Note': ['No User GPOs found']}).to_excel(
                writer, sheet_name='4. User GPOs', index=False)
            return
        
        user_gpos = gpos[gpos['Bucket'] == 'User'].copy()
        
        # Exclude Links-only GPOs - they go to Review Required tab
        if 'MatchType' in user_gpos.columns:
            user_gpos = user_gpos[user_gpos['MatchType'] != 'Links Only - Review Required']
        
        if user_gpos.empty:
            pd.DataFrame({'Note': ['No User GPOs found for this domain/operation']}).to_excel(
                writer, sheet_name='4. User GPOs', index=False)
            return
        
        # Build inventory
        inventory = []
        for _, gpo in user_gpos.iterrows():
            gpo_name = gpo.get('DisplayName', 'Unknown')
            description = gpo.get('Description', '')
            if pd.isna(description):
                description = ''
            
            # Format links for readable display
            links = format_links_for_display(gpo.get('Links', ''))
            
            # Count settings
            gpo_settings = [s for s in self.all_settings if s.get('GPOName') == gpo_name]
            
            # Get AppliesTo and ConfigMismatch
            applies_to = gpo.get('AppliesTo', 'Unknown')
            config_mismatch = gpo.get('ConfigMismatch', False)
            
            # Determine optimization note based on AppliesTo
            if applies_to == 'Computer':
                opt_note = 'MISMATCH: Computer settings in User OU - Consider moving to Workstation OU'
            elif applies_to == 'Both':
                opt_note = 'Has both Computer and User settings - Review for optimization'
            else:
                opt_note = 'Consider converting to Workstation GPO'
            
            inventory.append({
                'GPO Name': gpo_name,
                'Description': description,
                'Linked To': links,
                'Applies To': applies_to,
                'Settings Count': len(gpo_settings),
                'Size (MB)': gpo.get('SizeMB', 'N/A'),
                'Created': gpo.get('CreationTime', 'N/A'),
                'Last Modified': gpo.get('ModificationTime', 'N/A'),
                'GUID': gpo.get('GUID', ''),
                'Migration Priority': 'MEDIUM - REVIEW' if config_mismatch else 'MEDIUM',
                'Optimization Note': opt_note
            })
        
        df = pd.DataFrame(inventory)
        df.to_excel(writer, sheet_name='4. User GPOs', index=False)
        
        # Format the Linked To column for text wrap
        worksheet = writer.sheets['4. User GPOs']
        worksheet.column_dimensions['C'].width = 60
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=3)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    def _create_domain_tab5_review_required(self, writer):
        """
        Domain Tab 5: Review Required
        
        Combined tab for:
        - Domain Root GPOs (DON'T MIGRATE)
        - Domain Controller GPOs (DON'T MIGRATE)
        - Mixed GPOs (MANUAL REVIEW)
        - Unknown GPOs (MANUAL REVIEW)
        - Links-Only GPOs (linked to operation OU but name doesn't match)
        """
        gpos = self.report_gpos if not self.report_gpos.empty else pd.DataFrame()
        
        if gpos.empty:
            pd.DataFrame({'Note': ['No GPOs requiring review']}).to_excel(
                writer, sheet_name='5. Review Required', index=False)
            return
        
        # Get GPOs that need review for various reasons
        review_gpos_list = []
        
        # 1. Bucket-based review (Domain Root, DC, Mixed, Unknown)
        if 'Bucket' in gpos.columns:
            review_buckets = ['Domain Root', 'Domain Controller', 'Mixed', 'Unknown']
            bucket_review = gpos[gpos['Bucket'].isin(review_buckets)].copy()
            if not bucket_review.empty:
                review_gpos_list.append(bucket_review)
        
        # 2. Links-only GPOs (matched by Links but not by name)
        if 'MatchType' in gpos.columns:
            links_only = gpos[gpos['MatchType'] == 'Links Only - Review Required'].copy()
            # Don't duplicate - only add if not already in bucket review
            if not links_only.empty and review_gpos_list:
                existing_names = set()
                for df in review_gpos_list:
                    existing_names.update(df['DisplayName'].tolist())
                links_only = links_only[~links_only['DisplayName'].isin(existing_names)]
            if not links_only.empty:
                review_gpos_list.append(links_only)
        
        if not review_gpos_list:
            pd.DataFrame({'Note': ['No GPOs requiring review - all GPOs categorized']}).to_excel(
                writer, sheet_name='5. Review Required', index=False)
            return
        
        review_gpos = pd.concat(review_gpos_list, ignore_index=True).drop_duplicates(subset=['DisplayName'])
        
        # Build inventory with review reason
        inventory = []
        for _, gpo in review_gpos.iterrows():
            gpo_name = gpo.get('DisplayName', 'Unknown')
            bucket = gpo.get('Bucket', 'Unknown')
            match_type = gpo.get('MatchType', 'Name Match')
            description = gpo.get('Description', '')
            if pd.isna(description):
                description = ''
            
            # Format links for readable display
            links = format_links_for_display(gpo.get('Links', ''))
            
            # Get AppliesTo
            applies_to = gpo.get('AppliesTo', 'Unknown')
            
            # Determine action based on bucket AND match type
            if match_type == 'Links Only - Review Required':
                action = "REVIEW - Linked to operation OU but name doesn't contain operation code"
            elif bucket == 'Domain Root':
                action = "DON'T MIGRATE - Infrastructure GPO"
            elif bucket == 'Domain Controller':
                action = "DON'T MIGRATE - DC configuration"
            elif bucket == 'Mixed':
                bucket_details = gpo.get('BucketDetails', '')
                action = f"MANUAL REVIEW - Multiple bucket types: {bucket_details}"
            elif bucket == 'Unknown':
                action = "MANUAL REVIEW - Cannot determine bucket type"
            else:
                action = "MANUAL REVIEW - Unexpected bucket type"
            
            # Count settings
            gpo_settings = [s for s in self.all_settings if s.get('GPOName') == gpo_name]
            
            inventory.append({
                'GPO Name': gpo_name,
                'Bucket': bucket,
                'Match Type': match_type,
                'Action': action,
                'Applies To': applies_to,
                'Description': description,
                'Linked To': links,
                'Settings Count': len(gpo_settings),
                'Size (MB)': gpo.get('SizeMB', 'N/A'),
                'GUID': gpo.get('GUID', '')
            })
        
        df = pd.DataFrame(inventory)
        df.to_excel(writer, sheet_name='5. Review Required', index=False)
        
        # Format the Linked To column for text wrap (column G)
        worksheet = writer.sheets['5. Review Required']
        worksheet.column_dimensions['G'].width = 60
        for row in range(2, len(df) + 2):
            cell = worksheet.cell(row=row, column=7)
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # =========================================================================
    # WEB API METHODS - JSON-serializable data for FastAPI layer
    # These methods are the SINGLE SOURCE OF TRUTH for web data
    # The analyzer.py service should only call these, no business logic there
    # =========================================================================

    def get_web_executive_dashboard(self) -> Dict[str, Any]:
        """
        Get Executive Dashboard data (Tab 1) for web API.
        Returns JSON-serializable dict with metrics and bucket distribution.
        """
        active = self.active_gpos
        
        # Handle empty or missing data
        if active is None or (hasattr(active, 'empty') and active.empty):
            return {
                "total_gpos": 0,
                "active_gpos": 0,
                "active_percentage": 0.0,
                "migration_ready": 0,
                "migration_ready_percentage": 0.0,
                "needs_review": 0,
                "dont_migrate": 0,
                "bucket_distribution": [],
                "domains_analyzed": 0,
                "operations_count": 0
            }
        
        total = getattr(self, 'total_gpos_before_filter', 0) or 0
        active_count = len(active)
        
        # Safely get bucket counts
        bucket_counts = {}
        if 'Bucket' in active.columns:
            bucket_counts = active['Bucket'].value_counts().to_dict()
        
        # Migration categories with safe defaults
        migrate_ready = bucket_counts.get('Server', 0) + bucket_counts.get('Workstation', 0)
        needs_review = (bucket_counts.get('User', 0) + 
                       bucket_counts.get('Mixed', 0) + 
                       bucket_counts.get('Unknown', 0))
        dont_migrate = (bucket_counts.get('Domain Controller', 0) + 
                       bucket_counts.get('Domain Root', 0))
        
        # Bucket distribution for chart
        bucket_dist = []
        for bucket, count in bucket_counts.items():
            bucket_dist.append({
                "bucket": str(bucket),
                "count": int(count),
                "percentage": round(float(count) / float(active_count) * 100, 1) if active_count > 0 else 0.0
            })
        
        # Sort by count descending
        bucket_dist.sort(key=lambda x: x['count'], reverse=True)
        
        # Safe calculation of operations count
        domains_count = len(self.domains) if hasattr(self, 'domains') else 0
        try:
            ops_count = len([k for k in LOCATION_MAPPING.keys() if k in SHARED_FOREST_OPERATIONS]) + max(0, domains_count - 1)
        except:
            ops_count = domains_count
        
        return {
            "total_gpos": int(total),
            "active_gpos": int(active_count),
            "active_percentage": round(float(active_count) / float(total) * 100, 1) if total > 0 else 0.0,
            "migration_ready": int(migrate_ready),
            "migration_ready_percentage": round(float(migrate_ready) / float(active_count) * 100, 1) if active_count > 0 else 0.0,
            "needs_review": int(needs_review),
            "dont_migrate": int(dont_migrate),
            "bucket_distribution": bucket_dist,
            "domains_analyzed": int(domains_count),
            "operations_count": int(ops_count)
        }

    def get_web_operations_summary(self) -> List[Dict[str, Any]]:
        """
        Get Operations Summary data (Tab 2) for web API.
        Returns list of per-operation bucket breakdowns.
        """
        operations_data = []
        
        # Safety check for empty data
        if self.active_gpos is None or self.active_gpos.empty:
            return []
        
        # Check required columns exist
        required_cols = ['Domain', 'Bucket', 'DisplayName']
        for col in required_cols:
            if col not in self.active_gpos.columns:
                return []
        
        # Get enterprise GPOs for Shared Forest operations
        try:
            enterprise_gpos = self.active_gpos[
                self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
            ].copy()
        except Exception:
            enterprise_gpos = pd.DataFrame()
        
        # Process Shared Forest operations
        for code in sorted(SHARED_FOREST_OPERATIONS):
            if code not in LOCATION_MAPPING:
                continue
            
            info = LOCATION_MAPPING[code]
            op_gpos = self._web_filter_gpos_for_operation(enterprise_gpos.copy(), code, info)
            
            if op_gpos.empty:
                continue
            
            bucket_counts = op_gpos['Bucket'].value_counts().to_dict()
            total = len(op_gpos)
            
            server = bucket_counts.get('Server', 0)
            workstation = bucket_counts.get('Workstation', 0)
            user = bucket_counts.get('User', 0)
            mixed = bucket_counts.get('Mixed', 0)
            unknown = bucket_counts.get('Unknown', 0)
            dc = bucket_counts.get('Domain Controller', 0)
            domain_root = bucket_counts.get('Domain Root', 0)
            
            migrate_ready = server + workstation
            needs_review = user + mixed + unknown
            dont_migrate = dc + domain_root
            
            operations_data.append({
                "operation": code,
                "full_name": info['full_name'],
                "operation_type": "Shared Forest",
                "total_gpos": total,
                "server": server,
                "workstation": workstation,
                "user": user,
                "mixed": mixed,
                "unknown": unknown,
                "domain_controller": dc,
                "domain_root": domain_root,
                "migrate_ready": migrate_ready,
                "needs_review": needs_review,
                "dont_migrate": dont_migrate,
                "readiness_percentage": round(migrate_ready / total * 100, 1) if total > 0 else 0
            })
        
        # Process standalone domains
        for domain in sorted(self.domains.keys()):
            if domain.lower() in ['baseline.corp', 'baseline.corp']:
                continue
            
            domain_gpos = self.active_gpos[self.active_gpos['Domain'].str.lower() == domain.lower()]
            
            # Filter out ENT GPOs
            ent_mask = domain_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
            domain_gpos = domain_gpos[~ent_mask]
            
            if domain_gpos.empty:
                continue
            
            bucket_counts = domain_gpos['Bucket'].value_counts().to_dict()
            total = len(domain_gpos)
            
            server = bucket_counts.get('Server', 0)
            workstation = bucket_counts.get('Workstation', 0)
            user = bucket_counts.get('User', 0)
            mixed = bucket_counts.get('Mixed', 0)
            unknown = bucket_counts.get('Unknown', 0)
            dc = bucket_counts.get('Domain Controller', 0)
            domain_root = bucket_counts.get('Domain Root', 0)
            
            migrate_ready = server + workstation
            needs_review = user + mixed + unknown
            dont_migrate = dc + domain_root
            
            # Find operation code for this domain
            op_code = domain
            for code, info in LOCATION_MAPPING.items():
                if info.get('source_domain', '').lower() == domain.lower():
                    op_code = code
                    break
            
            operations_data.append({
                "operation": op_code,
                "full_name": domain,
                "operation_type": "Standalone",
                "total_gpos": total,
                "server": server,
                "workstation": workstation,
                "user": user,
                "mixed": mixed,
                "unknown": unknown,
                "domain_controller": dc,
                "domain_root": domain_root,
                "migrate_ready": migrate_ready,
                "needs_review": needs_review,
                "dont_migrate": dont_migrate,
                "readiness_percentage": round(migrate_ready / total * 100, 1) if total > 0 else 0
            })
        
        # Sort by total GPOs descending
        operations_data.sort(key=lambda x: x['total_gpos'], reverse=True)
        
        return operations_data

    def get_web_risk_assessment(self) -> List[Dict[str, Any]]:
        """
        Get Risk Assessment data (Tab 3) for web API.
        Returns operations ranked by complexity.
        """
        ops_data = self.get_web_operations_summary()
        risk_data = []
        
        for op in ops_data:
            # Calculate risk score
            risk_score = (op['mixed'] * 3) + (op['unknown'] * 2) + op['domain_controller'] + op['domain_root']
            
            # Determine risk level
            total = op['total_gpos']
            review_pct = (op['needs_review'] / total * 100) if total > 0 else 0
            
            if review_pct > 30:
                risk_level = "HIGH"
            elif review_pct > 15:
                risk_level = "MEDIUM"
            else:
                risk_level = "LOW"
            
            # Generate risk notes
            risk_notes = self._web_get_risk_notes(op)
            
            risk_data.append({
                "operation": op['operation'],
                "full_name": op['full_name'],
                "total_gpos": total,
                "risk_score": risk_score,
                "risk_level": risk_level,
                "mixed_count": op['mixed'],
                "unknown_count": op['unknown'],
                "dc_count": op['domain_controller'],
                "domain_root_count": op['domain_root'],
                "risk_notes": risk_notes
            })
        
        # Sort by risk score descending
        risk_data.sort(key=lambda x: x['risk_score'], reverse=True)
        
        return risk_data

    def _web_get_risk_notes(self, op: Dict) -> str:
        """Generate risk-specific guidance notes."""
        notes = []
        
        if op['mixed'] > 0:
            notes.append(f"{op['mixed']} Mixed GPOs need bucket splitting")
        if op['unknown'] > 0:
            notes.append(f"{op['unknown']} Unknown GPOs need OU review")
        if op['domain_root'] > 5:
            notes.append(f"{op['domain_root']} Domain Root GPOs - assess scope")
        
        if not notes:
            notes.append("Low complexity - ready for migration")
        
        return "; ".join(notes)

    def get_web_bucket_comparison(self) -> List[Dict[str, Any]]:
        """
        Get Bucket Comparison data (Tab 4) for web API.
        Returns cross-operation bucket percentages.
        """
        ops_data = self.get_web_operations_summary()
        comparison_data = []
        
        for op in ops_data:
            total = op['total_gpos']
            if total == 0:
                continue
            
            other = op['domain_controller'] + op['domain_root'] + op['unknown']
            
            comparison_data.append({
                "operation": op['operation'],
                "full_name": op['full_name'],
                "total_gpos": total,
                "server_pct": round(op['server'] / total * 100, 1),
                "workstation_pct": round(op['workstation'] / total * 100, 1),
                "user_pct": round(op['user'] / total * 100, 1),
                "mixed_pct": round(op['mixed'] / total * 100, 1),
                "other_pct": round(other / total * 100, 1)
            })
        
        return comparison_data

    def get_web_available_operations(self) -> List[Dict[str, Any]]:
        """
        Get list of available operations with GPO counts for web API.
        """
        operations = []
        
        if self.active_gpos is None or self.active_gpos.empty:
            return []
        
        # Shared Forest operations
        enterprise_gpos = self.active_gpos[
            self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
        ]
        
        for code in sorted(SHARED_FOREST_OPERATIONS):
            if code not in LOCATION_MAPPING:
                continue
            info = LOCATION_MAPPING[code]
            gpos = self._web_filter_gpos_for_operation(enterprise_gpos.copy(), code, info)
            if not gpos.empty:
                operations.append({
                    "code": code,
                    "full_name": info['full_name'],
                    "source_domain": "baseline.corp",
                    "operation_type": "Shared Forest",
                    "gpo_count": len(gpos)
                })
        
        # Standalone domains
        for domain in sorted(self.domains.keys()):
            if domain.lower() in ['baseline.corp', 'baseline.corp']:
                continue
            
            domain_mask = self.active_gpos['Domain'].str.lower() == domain.lower()
            domain_gpos = self.active_gpos[domain_mask]
            ent_mask = domain_gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
            domain_gpos = domain_gpos[~ent_mask]
            
            if domain_gpos.empty:
                continue
            
            # Find code
            code = domain
            for c, i in LOCATION_MAPPING.items():
                if i.get('source_domain', '').lower() == domain.lower():
                    code = c
                    break
            
            operations.append({
                "code": code,
                "full_name": domain,
                "source_domain": domain,
                "operation_type": "Standalone",
                "gpo_count": len(domain_gpos)
            })
        
        return operations

    def get_web_domain_overview(self, operation_code: str) -> Dict[str, Any]:
        """
        Get Domain Overview data for a specific operation for web API.
        """
        gpos, info = self._web_get_operation_gpos(operation_code)
        
        if gpos.empty:
            return {
                "operation": operation_code,
                "full_name": info.get('full_name', operation_code),
                "total_gpos": 0,
                "bucket_counts": {},
                "migration_summary": {"MIGRATE": 0, "DON'T MIGRATE": 0, "REVIEW": 0}
            }
        
        bucket_counts = gpos['Bucket'].value_counts().to_dict()
        
        migrate = bucket_counts.get('Server', 0) + bucket_counts.get('Workstation', 0)
        dont_migrate = bucket_counts.get('Domain Controller', 0) + bucket_counts.get('Domain Root', 0)
        review = (bucket_counts.get('User', 0) + bucket_counts.get('Mixed', 0) + 
                 bucket_counts.get('Unknown', 0))
        
        return {
            "operation": operation_code,
            "full_name": info.get('full_name', operation_code),
            "total_gpos": len(gpos),
            "bucket_counts": bucket_counts,
            "migration_summary": {
                "MIGRATE": migrate,
                "DON'T MIGRATE": dont_migrate,
                "REVIEW": review
            }
        }

    def get_web_domain_gpos(
        self, 
        operation_code: str, 
        bucket: str,
        page: int = 1,
        limit: int = 50
    ) -> Tuple[List[Dict[str, Any]], int]:
        """
        Get GPOs for a specific operation filtered by bucket for web API.
        Returns (gpo_list, total_count).
        """
        gpos, info = self._web_get_operation_gpos(operation_code)
        
        if gpos.empty:
            return [], 0
        
        # Filter by bucket
        is_review = bucket.lower() == 'review'
        if is_review:
            review_buckets = ['Mixed', 'Unknown', 'Domain Controller', 'Domain Root']
            bucket_mask = gpos['Bucket'].isin(review_buckets)
            if 'MatchType' in gpos.columns:
                links_only_mask = gpos['MatchType'] == 'Links Only - Review Required'
                filtered = gpos[bucket_mask | links_only_mask]
            else:
                filtered = gpos[bucket_mask]
        else:
            filtered = gpos[gpos['Bucket'] == bucket]
            if 'MatchType' in filtered.columns:
                filtered = filtered[filtered['MatchType'] != 'Links Only - Review Required']
        
        total = len(filtered)
        
        # Paginate
        start = (page - 1) * limit
        end = start + limit
        page_data = filtered.iloc[start:end]
        
        # Convert to list of dicts
        result = []
        for _, row in page_data.iterrows():
            gpo_name = row.get('DisplayName', 'Unknown')
            bucket_val = row.get('Bucket', 'Unknown')
            match_type = row.get('MatchType', 'Name Match')
            applies_to = row.get('AppliesTo', 'Unknown')
            
            action = None
            if is_review:
                action = self._web_get_review_action(bucket_val, match_type, row)
            
            readiness = self._web_get_readiness(row)
            
            opt_note = None
            if bucket == 'User':
                opt_note = self._web_get_optimization_note(row)
            
            detection_reason = self._web_generate_detection_reason(bucket_val, row)
            
            gpo_row = {
                "gpo_name": gpo_name,
                "description": str(row.get('Description', '')) if pd.notna(row.get('Description')) else '',
                "linked_to": self._web_format_links(row.get('Links', '')),
                "applies_to": applies_to,
                "settings_count": self._web_get_settings_count(row),
                "size_mb": str(row.get('SizeMB', '')) if pd.notna(row.get('SizeMB')) else None,
                "created": str(row.get('CreationTime', '')) if pd.notna(row.get('CreationTime')) else None,
                "last_modified": str(row.get('ModificationTime', '')) if pd.notna(row.get('ModificationTime')) else None,
                "guid": str(row.get('GUID', '')) if pd.notna(row.get('GUID')) else None,
                "bucket": bucket_val,
                "match_type": match_type,
                "readiness": readiness,
                "optimization_note": opt_note,
                "detection_reason": detection_reason,
                "action": action
            }
            result.append(gpo_row)
        
        return result, total

    def get_web_gpo_details(self, operation_code: str, gpo_name: str) -> Optional[Dict[str, Any]]:
        """
        Get detailed information for a specific GPO including settings for web API.
        """
        gpos, info = self._web_get_operation_gpos(operation_code)
        
        if gpos.empty:
            return None
        
        # Find the specific GPO
        gpo_mask = gpos['DisplayName'].str.lower() == gpo_name.lower()
        gpo_matches = gpos[gpo_mask]
        
        if gpo_matches.empty:
            return None
        
        gpo_row = gpo_matches.iloc[0]
        
        # Parse link locations
        links_raw = str(gpo_row.get('Links', ''))
        link_locations = [l.strip() for l in links_raw.split('\n') if l.strip()]
        
        # Get settings for this GPO
        settings = []
        if self.all_settings:
            gpo_settings = [
                s for s in self.all_settings
                if s.get('GPOName', '').lower() == gpo_name.lower()
            ]
            for s in gpo_settings:
                settings.append({
                    "category": s.get('Category', 'Unknown'),
                    "setting_name": s.get('SettingName', 'Unknown'),
                    "setting_value": str(s.get('SettingValue', ''))[:500]
                })
        
        # Detect issues
        issues = self._web_detect_gpo_issues(gpo_row, settings)
        
        bucket = gpo_row.get('Bucket', 'Unknown')
        match_type = gpo_row.get('MatchType', 'Name Match')
        
        return {
            "gpo_name": gpo_row.get('DisplayName', gpo_name),
            "guid": str(gpo_row.get('GUID', '')) if pd.notna(gpo_row.get('GUID')) else None,
            "description": str(gpo_row.get('Description', '')) if pd.notna(gpo_row.get('Description')) else None,
            "applies_to": gpo_row.get('AppliesTo', 'Unknown'),
            "bucket": bucket,
            "readiness": self._web_get_readiness(gpo_row),
            "match_type": match_type,
            "created": str(gpo_row.get('CreationTime', '')) if pd.notna(gpo_row.get('CreationTime')) else None,
            "last_modified": str(gpo_row.get('ModificationTime', '')) if pd.notna(gpo_row.get('ModificationTime')) else None,
            "link_locations": link_locations,
            "settings": settings,
            "settings_count": len(settings),
            "issues": issues,
            "detection_reason": self._web_generate_detection_reason(bucket, gpo_row),
            "action": self._web_get_review_action(bucket, match_type, gpo_row) if bucket in ['Mixed', 'Unknown', 'Domain Controller', 'Domain Root'] or match_type == 'Links Only - Review Required' else None
        }

    def get_migration_domains(self) -> List[Dict[str, Any]]:
        """
        Get list of domains available for migration analysis for web API.
        Includes both standalone domains AND Shared Forest operations.

        GPO counts reflect the FULL migration scope:
        - All Active GPOs from the domain (including ENT copies)
        - Plus any Shared Forest GPOs matching the operation prefix
        """
        operations = self.get_web_available_operations()

        domains = []
        for op in operations:
            # For Shared Forest operations, use code as domain identifier
            # For standalone, use source_domain
            if op.get('operation_type') == 'Shared Forest':
                domain_id = op.get('code', '')
                # Shared Forest: count is already correct from get_web_available_operations
                gpo_count = op.get('gpo_count', 0)
            else:
                domain_id = op.get('source_domain', op.get('code', ''))
                # Standalone: calculate FULL migration scope (including ENT + Shared Forest matches)
                gpo_count = self._get_migration_scope_count(domain_id)

            domains.append({
                "domain": domain_id,
                "code": op.get('code', ''),  # Operation code for display
                "display_name": op.get('full_name', op.get('code', '')),
                "domain_type": op.get('operation_type', 'Unknown'),
                "gpo_count": gpo_count
            })
        
        # Sort: Shared Forest first, then standalone
        domains.sort(key=lambda x: (0 if x['domain_type'] == 'Shared Forest' else 1, x['display_name']))
        
        # Add Enterprise-Wide GPOs (policies that apply globally, not operation-specific)
        enterprise_wide = self._get_unmatched_enterprise_gpos()
        if enterprise_wide:
            domains.append({
                "domain": "_ENTERPRISE_WIDE",
                "display_name": f"🌐 Enterprise-Wide ({len(enterprise_wide)})",
                "domain_type": "Enterprise-Wide",
                "gpo_count": len(enterprise_wide),
                "unmatched_gpos": enterprise_wide  # List of GPO names for reference
            })
        
        return domains

    def _get_migration_scope_count(self, domain: str) -> int:
        """
        Calculate the FULL migration scope count for a standalone domain.

        This includes:
        1. ALL Active GPOs from the standalone domain (including ENT copies)
        2. Plus any Shared Forest GPOs from baseline.corp that match the operation prefix

        This ensures the dropdown count matches what run_migration_analysis() returns.
        """
        if self.active_gpos is None or self.active_gpos.empty:
            return 0

        total_gpos = set()

        # 1. Get ALL Active GPOs from standalone domain (no ENT filter)
        domain_mask = self.active_gpos['Domain'].str.lower() == domain.lower()
        domain_gpos = self.active_gpos[domain_mask]['DisplayName'].dropna().unique()
        total_gpos.update(domain_gpos)

        # 2. Find if this domain has shared_prefixes (GPOs in baseline.corp too)
        shared_prefixes = []
        for code, info in LOCATION_MAPPING.items():
            if info.get('source_domain', '').lower() == domain.lower():
                if 'shared_prefixes' in info:
                    shared_prefixes.extend([p.lower() for p in info['shared_prefixes']])
        shared_prefixes = list(set(shared_prefixes)) if shared_prefixes else None

        # 3. If shared_prefixes exist, count matching GPOs from baseline.corp/baseline.corp
        if shared_prefixes:
            ent_gpos = self.active_gpos[
                self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
            ]
            for _, row in ent_gpos.iterrows():
                gpo_name = row.get('DisplayName', '')
                gpo_name_lower = gpo_name.lower()

                # Skip ENT GPOs
                if is_enterprise_standard_gpo(gpo_name):
                    continue

                # Check if matches any prefix
                for prefix in shared_prefixes:
                    if len(prefix) <= 3:
                        pattern = r'(?:^|[\s\-_])' + re.escape(prefix) + r'(?:$|[\s\-_])'
                        if re.search(pattern, gpo_name_lower):
                            total_gpos.add(gpo_name)
                            break
                    else:
                        if prefix in gpo_name_lower:
                            total_gpos.add(gpo_name)
                            break

        return len(total_gpos)

    def _get_unmatched_enterprise_gpos(self) -> List[str]:
        """
        Find GPOs in baseline.corp that don't match any known operation.
        
        Uses TWO detection methods:
        1. Name prefix matching (OPF-, OPG-, OPA-, etc.)
        2. Link-based detection (DetectedOperation from OU structure)
        
        Returns list of GPO names that:
        - Are NOT ENT-prefixed
        - Do NOT match any Shared Forest operation by name prefix
        - Do NOT match any standalone operation's shared_prefixes
        - Do NOT have a valid DetectedOperation from link analysis
        
        These are truly "Enterprise-Wide" policies that apply globally.
        """
        if self.active_gpos is None or self.active_gpos.empty:
            return []
        
        # Get baseline.corp GPOs only
        enterprise_gpos = self.active_gpos[
            self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
        ]
        
        if enterprise_gpos.empty:
            return []
        
        # Build all known prefixes from ALL operations (Shared Forest AND standalone with shared_prefixes)
        all_prefixes = set()
        valid_operations = set()
        
        # Add Shared Forest operation prefixes
        for code in SHARED_FOREST_OPERATIONS:
            if code in LOCATION_MAPPING:
                info = LOCATION_MAPPING[code]
                all_prefixes.add(code.lower())
                all_prefixes.add(info['full_name'].lower())
                valid_operations.add(info['full_name'])
                if 'name_prefixes' in info:
                    all_prefixes.update([p.lower() for p in info['name_prefixes']])
                    valid_operations.update(info['name_prefixes'])
        
        # Add standalone operation shared_prefixes (e.g., OPA, OPB)
        for code, info in LOCATION_MAPPING.items():
            if 'shared_prefixes' in info:
                all_prefixes.add(code.lower())
                all_prefixes.add(info['full_name'].lower())
                valid_operations.add(info['full_name'])
                all_prefixes.update([p.lower() for p in info['shared_prefixes']])
                valid_operations.update(info['shared_prefixes'])
        
        # Check if DetectedOperation column exists
        has_detected_op = 'DetectedOperation' in enterprise_gpos.columns
        
        unmatched = []
        for _, row in enterprise_gpos.iterrows():
            gpo_name = row.get('DisplayName', '')
            if not gpo_name:
                continue
            
            gpo_name_lower = gpo_name.lower()
            
            # Skip ENT GPOs - they're the baseline
            if is_enterprise_standard_gpo(gpo_name):
                continue
            
            # Method 1: Check if matches any operation prefix by NAME
            matches_by_name = False
            for prefix in all_prefixes:
                if len(prefix) < 2:
                    continue
                if len(prefix) <= 3:
                    # Short prefix - require word boundary
                    pattern = r'(?:^|[\s\-_])' + re.escape(prefix) + r'(?:$|[\s\-_])'
                    if re.search(pattern, gpo_name_lower):
                        matches_by_name = True
                        break
                else:
                    if prefix in gpo_name_lower:
                        matches_by_name = True
                        break
            
            # Method 2: Check if has valid DetectedOperation from LINK analysis
            matches_by_link = False
            if has_detected_op:
                detected_op = row.get('DetectedOperation', 'Unknown')
                if detected_op and detected_op != 'Unknown':
                    # Check if detected operation matches a known operation
                    for valid_op in valid_operations:
                        if valid_op.lower() in detected_op.lower() or detected_op.lower() in valid_op.lower():
                            matches_by_link = True
                            break
            
            # Only unmatched if BOTH methods fail
            if not matches_by_name and not matches_by_link:
                unmatched.append(gpo_name)
        
        return sorted(unmatched)

    def run_migration_analysis(self, domain: str) -> Dict[str, Any]:
        """
        Run migration analysis for a domain/operation and return summary for web API.
        
        Handles three scenarios:
        1. Standalone domains (charlie.local, etc.) - compare domain HTML vs baseline.corp HTML
        2. Shared Forest operations (OPF, OPG, etc.) - compare operation GPOs vs ENT GPOs within baseline.corp
        3. _ENTERPRISE_WIDE - analyze global policies that aren't operation-specific
        
        Uses the existing SettingExtractionEngine and MigrationComparisonEngine.
        """
        # Handle Enterprise-Wide GPOs (global policies)
        if domain == '_ENTERPRISE_WIDE':
            return self._run_enterprise_wide_analysis()
        
        # Check if this is an Shared Forest operation
        is_shared_forest = domain.upper() in SHARED_FOREST_OPERATIONS
        
        if is_shared_forest:
            return self._run_shared_forest_migration_analysis(domain.upper())
        else:
            return self._run_standalone_migration_analysis(domain)
    
    def _run_enterprise_wide_analysis(self) -> Dict[str, Any]:
        """
        Analyze Enterprise-Wide GPOs (policies that apply globally, not operation-specific).
        
        These are GPOs in baseline.corp that:
        - Are NOT ENT-prefixed
        - Do NOT match any operation by name OR link
        
        We compare them against Enterprise Standard baseline to identify:
        - Settings that duplicate Enterprise Standard (DON'T MIGRATE)
        - Settings that conflict with Enterprise Standard (REVIEW)
        - Settings unique to these policies (MIGRATE - or consolidate into Enterprise Standard)
        """
        enterprise_wide_gpos = self._get_unmatched_enterprise_gpos()
        
        if not enterprise_wide_gpos:
            return {
                "domain": "_ENTERPRISE_WIDE",
                "total_settings": 0,
                "migrate": 0,
                "dont_migrate": 0,
                "review": 0,
                "total_gpos": 0,
                "p1_count": 0,
                "p2_count": 0,
                "p3_count": 0,
                "p4_count": 0,
                "gpo_summary": [],
                "warning": "No Enterprise-Wide GPOs found. All GPOs matched to operations."
            }
        
        # Find baseline.corp HTML file
        ent_html_file = None
        enterprise_domain = None
        
        for html_file in self.html_folder.glob('*.html'):
            with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)
            
            domain_match = re.search(r'"DomainName":\s*"([^"]+)"', content)
            if domain_match:
                detected_domain = domain_match.group(1)
                if 'baseline' in detected_domain.lower() or 'baseline.corp' in detected_domain.lower():
                    ent_html_file = html_file
                    enterprise_domain = detected_domain
                    break
        
        if not ent_html_file:
            return {"error": "No baseline.corp HTML file found"}
        
        # Convert list to set for fast lookup
        enterprise_wide_set = set(enterprise_wide_gpos)
        
        # Build set of ACTIVE GPO names - only process settings from active GPOs
        active_gpo_names = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            active_gpo_names = set(self.active_gpos['DisplayName'].dropna().unique())
        
        # Initialize extraction engine
        extraction_engine = SettingExtractionEngine(verbose=False)
        
        # Extract ALL settings from baseline.corp
        with open(ent_html_file, 'r', encoding='utf-8', errors='ignore') as f:
            ent_content = f.read()
        all_settings = extraction_engine.extract_from_html_content(ent_content, enterprise_domain or 'baseline.corp')
        
        # Filter settings into Enterprise-Wide vs Enterprise Standard (only from ACTIVE GPOs)
        enterprise_wide_settings = []
        ent_settings = []
        
        for setting in all_settings:
            gpo_name = setting.gpo_name if hasattr(setting, 'gpo_name') else setting.get('gpo_name', '')
            
            # CRITICAL: Skip settings from disabled/unlinked GPOs
            if gpo_name not in active_gpo_names:
                continue
            
            if is_enterprise_standard_gpo(gpo_name):
                ent_settings.append(setting)
            elif gpo_name in enterprise_wide_set:
                enterprise_wide_settings.append(setting)
        
        if not enterprise_wide_settings:
            return {
                "domain": "_ENTERPRISE_WIDE",
                "total_settings": 0,
                "migrate": 0,
                "dont_migrate": 0,
                "review": 0,
                "total_gpos": len(enterprise_wide_gpos),
                "p1_count": 0,
                "p2_count": 0,
                "p3_count": 0,
                "p4_count": len(enterprise_wide_gpos),
                "gpo_summary": [{'priority': 'P4', 'action': 'NO SETTINGS FOUND', 'operation_gpo': g, 
                                'ent_standard_overlap': '', 'shared_forest_operation_overlap': '',
                                'total_settings': 0, 'migrate': 0, 'dont_migrate': 0, 'review': 0}
                               for g in enterprise_wide_gpos],
                "warning": "Enterprise-Wide GPOs found but no settings extracted. May be empty or links-only policies."
            }
        
        # Initialize comparison engine and run analysis
        comparison_engine = MigrationComparisonEngine()
        comparison_engine.load_operation_settings(enterprise_wide_settings)
        comparison_engine.load_enterprise_settings(ent_settings)
        
        # Build source lookup from active_gpos (baseline.corp vs baseline.corp)
        gpo_source_lookup = {}
        links_lookup = {}
        if self.active_gpos is not None and not self.active_gpos.empty:
            enterprise_gpos_df = self.active_gpos[
                self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
            ]
            for _, row in enterprise_gpos_df.iterrows():
                gpo_name = row.get('DisplayName', '')
                domain_name = row.get('DomainName', 'baseline.corp')
                links = row.get('Links', '')
                if gpo_name:
                    gpo_source_lookup[gpo_name] = domain_name
                    links_lookup[gpo_name] = links
        
        # Use shared response builder with source lookup
        # Pass the enterprise_wide_gpos set to include GPOs with 0 settings
        enterprise_wide_set = set(enterprise_wide_gpos)
        result = self._build_migration_response(comparison_engine, "_ENTERPRISE_WIDE", gpo_source_lookup, enterprise_wide_set)
        
        # Add formatted DN path (same format as Domain dashboard's Linked To column)
        for gpo_row in result.get('gpo_summary', []):
            gpo_name = gpo_row.get('operation_gpo', '')
            raw_links = links_lookup.get(gpo_name, '')
            # Use same formatting as Domain dashboard for consistency
            gpo_row['dn_path'] = self._web_format_links(raw_links)
        
        # Add context-specific warning
        result["warning"] = (
            f"These {len(enterprise_wide_gpos)} GPOs apply globally across the enterprise. "
            "Consider consolidating unique settings into Enterprise Standard baselines."
        )
        
        return result
    
    def _run_standalone_migration_analysis(self, domain: str) -> Dict[str, Any]:
        """
        Run migration analysis for a standalone domain.
        Compares domain HTML vs baseline.corp HTML.
        
        If the operation has shared_prefixes defined, also pulls matching GPOs
        from baseline.corp and merges them into the analysis (single pane of glass).
        """
        # Find HTML files
        op_html_file = None
        ent_html_file = None
        enterprise_domain = None
        
        for html_file in self.html_folder.glob('*.html'):
            with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)
            
            domain_match = re.search(r'"DomainName":\s*"([^"]+)"', content)
            if domain_match:
                detected_domain = domain_match.group(1)
                
                if detected_domain.lower() == domain.lower():
                    op_html_file = html_file
                
                if 'baseline' in detected_domain.lower() or 'baseline.corp' in detected_domain.lower():
                    ent_html_file = html_file
                    enterprise_domain = detected_domain
        
        if not op_html_file:
            return {"error": f"Could not find HTML file for domain: {domain}"}
        
        if not ent_html_file:
            return {"error": "No enterprise baseline found"}
        
        # Find if this domain has shared_prefixes (GPOs in baseline.corp too)
        # Collect from ALL operations that share this source_domain
        shared_prefixes = []
        for code, info in LOCATION_MAPPING.items():
            if info.get('source_domain', '').lower() == domain.lower():
                if 'shared_prefixes' in info:
                    shared_prefixes.extend([p.lower() for p in info['shared_prefixes']])
        shared_prefixes = list(set(shared_prefixes)) if shared_prefixes else None  # Dedupe or None if empty
        
        # Initialize engines
        extraction_engine = SettingExtractionEngine(verbose=False)
        comparison_engine = MigrationComparisonEngine()
        
        # Build set of ACTIVE GPO names - only process settings from active GPOs
        active_gpo_names = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            active_gpo_names = set(self.active_gpos['DisplayName'].dropna().unique())
        
        # Build source lookup: GPO name -> source domain
        gpo_source_lookup = {}
        
        # Extract settings from operation domain - all get tagged with standalone domain
        with open(op_html_file, 'r', encoding='utf-8', errors='ignore') as f:
            op_content = f.read()
        all_op_settings = extraction_engine.extract_from_html_content(op_content, domain)
        
        # Filter to only ACTIVE GPOs from standalone domain
        op_settings = []
        for setting in all_op_settings:
            gpo_name = setting.gpo_name if hasattr(setting, 'gpo_name') else setting.get('gpo_name', '')
            # CRITICAL: Skip settings from disabled/unlinked GPOs
            if gpo_name not in active_gpo_names:
                continue
            op_settings.append(setting)
            if gpo_name:
                gpo_source_lookup[gpo_name] = domain
        
        # If this operation has shared_prefixes, also pull matching GPOs from baseline.corp
        shared_gpo_count = 0
        if shared_prefixes and ent_html_file:
            with open(ent_html_file, 'r', encoding='utf-8', errors='ignore') as f:
                ent_content = f.read()
            all_ent_settings = extraction_engine.extract_from_html_content(ent_content, enterprise_domain or 'baseline.corp')
            
            # Build lookup from active_gpos for actual DomainName (baseline.corp vs baseline.corp)
            # Also use this as the filter for which GPOs are active
            shared_domain_lookup = {}
            active_shared_gpos = set()
            if self.active_gpos is not None and not self.active_gpos.empty:
                ent_gpos_df = self.active_gpos[
                    self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
                ]
                for _, row in ent_gpos_df.iterrows():
                    gpo_name = row.get('DisplayName', '')
                    domain_name = row.get('DomainName', 'baseline.corp')
                    if gpo_name:
                        shared_domain_lookup[gpo_name] = domain_name
                        active_shared_gpos.add(gpo_name)
            
            # Filter for GPOs matching this operation's shared_prefixes (ACTIVE ONLY)
            shared_gpos_seen = set()
            for setting in all_ent_settings:
                gpo_name = setting.gpo_name if hasattr(setting, 'gpo_name') else setting.get('gpo_name', '')
                gpo_name_lower = gpo_name.lower()
                
                # CRITICAL: Skip settings from disabled/unlinked GPOs
                if gpo_name not in active_shared_gpos:
                    continue
                
                # Skip ENT GPOs
                if is_enterprise_standard_gpo(gpo_name):
                    continue
                
                # Check if matches any of the shared_prefixes
                matches = False
                for prefix in shared_prefixes:
                    if len(prefix) <= 3:
                        pattern = r'(?:^|[\s\-_])' + re.escape(prefix) + r'(?:$|[\s\-_])'
                        if re.search(pattern, gpo_name_lower):
                            matches = True
                            break
                    else:
                        if prefix in gpo_name_lower:
                            matches = True
                            break
                
                if matches:
                    op_settings.append(setting)
                    shared_gpos_seen.add(gpo_name)
                    # Tag with actual source domain (baseline.corp or baseline.corp)
                    gpo_source_lookup[gpo_name] = shared_domain_lookup.get(gpo_name, 'baseline.corp')
            
            shared_gpo_count = len(shared_gpos_seen)
        
        comparison_engine.load_operation_settings(op_settings)

        # Extract Enterprise Standard settings from enterprise baseline for comparison (ACTIVE ONLY)
        if ent_html_file != op_html_file:
            with open(ent_html_file, 'r', encoding='utf-8', errors='ignore') as f:
                ent_content = f.read()
            all_ent_settings = extraction_engine.extract_from_html_content(ent_content, enterprise_domain or 'baseline.corp')
            # Filter to only ACTIVE Enterprise Standard settings
            ent_settings = [s for s in all_ent_settings if
                (s.gpo_name if hasattr(s, 'gpo_name') else s.get('gpo_name', '')) in active_gpo_names and
                is_enterprise_standard_gpo(s.gpo_name if hasattr(s, 'gpo_name') else s.get('gpo_name', ''))
            ]
            comparison_engine.load_enterprise_settings(ent_settings)

        # Build set of ALL Active GPOs for this domain (to include those with 0 settings)
        # This includes: standalone domain GPOs + matched Shared Forest GPOs
        all_migration_gpos = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            # Get all Active GPOs from standalone domain
            domain_gpos = self.active_gpos[
                self.active_gpos['Domain'].str.lower() == domain.lower()
            ]['DisplayName'].dropna().unique()
            all_migration_gpos.update(domain_gpos)

            # Add Shared Forest GPOs that were matched (already tracked in gpo_source_lookup)
            if shared_prefixes:
                all_migration_gpos.update(gpo_source_lookup.keys())

        # Compare, classify, and build response with source lookup
        result = self._build_migration_response(comparison_engine, domain, gpo_source_lookup, all_migration_gpos)
        
        # Add note about Shared Forest GPOs if any were merged
        if shared_gpo_count > 0:
            result["warning"] = f"Includes {shared_gpo_count} GPO(s) from Shared Forest matching this operation's prefix."
        
        return result
    
    def _run_shared_forest_migration_analysis(self, operation_code: str) -> Dict[str, Any]:
        """
        Run migration analysis for an Shared Forest operation.
        Compares operation GPOs vs ENT GPOs within the same baseline.corp file.
        
        Uses TWO detection methods to find operation GPOs:
        1. Name prefix matching (OPF-, OPG-, etc.)
        2. Link-based detection (DetectedOperation from OU structure)
        
        This ensures GPOs with non-standard naming but linked to operation OUs are included.
        """
        # Get operation info
        if operation_code not in LOCATION_MAPPING:
            return {"error": f"Unknown operation code: {operation_code}"}
        
        info = LOCATION_MAPPING[operation_code]
        
        # Find baseline.corp HTML file
        ent_html_file = None
        enterprise_domain = None
        
        for html_file in self.html_folder.glob('*.html'):
            with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)
            
            domain_match = re.search(r'"DomainName":\s*"([^"]+)"', content)
            if domain_match:
                detected_domain = domain_match.group(1)
                if 'baseline' in detected_domain.lower() or 'baseline.corp' in detected_domain.lower():
                    ent_html_file = html_file
                    enterprise_domain = detected_domain
                    break
        
        if not ent_html_file:
            return {"error": "No baseline.corp HTML file found"}
        
        # Build search terms for this operation (for name-based matching)
        search_terms = [operation_code.lower(), info['full_name'].lower()]
        if 'name_prefixes' in info:
            search_terms.extend([p.lower() for p in info['name_prefixes']])
        
        # Get GPOs detected by LINK analysis for this operation
        link_matched_gpos = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            if 'DetectedOperation' in self.active_gpos.columns:
                # Get baseline.corp GPOs
                enterprise_gpos = self.active_gpos[
                    self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
                ]
                # Find GPOs where DetectedOperation matches this operation
                for _, row in enterprise_gpos.iterrows():
                    detected_op = str(row.get('DetectedOperation', '')).lower()
                    gpo_name = row.get('DisplayName', '')
                    
                    # Check if detected operation matches any of our search terms
                    # Use same word boundary logic as name matching
                    for term in search_terms:
                        if len(term) < 2:
                            # Skip single-char terms entirely for link matching
                            continue
                        if len(term) <= 3:
                            # Short term - require word boundary or exact match
                            # e.g., 'opf' should match 'foxtrot' but short codes shouldn't match unrelated operations
                            pattern = r'(?:^|[\s\-_])' + re.escape(term) + r'(?:$|[\s\-_])'
                            if re.search(pattern, detected_op) or detected_op == term:
                                link_matched_gpos.add(gpo_name)
                                break
                        else:
                            # Longer term - substring matching is safe
                            # 'foxtrot' in 'foxtrot' = True, short prefix check handled above
                            if term in detected_op or detected_op in term:
                                link_matched_gpos.add(gpo_name)
                                break
        
        # Build set of ACTIVE GPO names - only process settings from active GPOs
        active_gpo_names = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            active_gpo_names = set(self.active_gpos['DisplayName'].dropna().unique())
        
        # Initialize extraction engine
        extraction_engine = SettingExtractionEngine(verbose=False)
        
        # Extract ALL settings from baseline.corp
        with open(ent_html_file, 'r', encoding='utf-8', errors='ignore') as f:
            ent_content = f.read()
        all_settings = extraction_engine.extract_from_html_content(ent_content, enterprise_domain or 'baseline.corp')
        
        # Filter settings into operation vs Enterprise Standard (only from ACTIVE GPOs)
        operation_settings = []
        ent_settings = []
        
        for setting in all_settings:
            # ExtractedSetting is a dataclass, use attribute access
            gpo_name = setting.gpo_name if hasattr(setting, 'gpo_name') else setting.get('gpo_name', '')
            gpo_name_lower = gpo_name.lower()
            
            # CRITICAL: Skip settings from disabled/unlinked GPOs
            if gpo_name not in active_gpo_names:
                continue
            
            # Check if Enterprise Standard GPO
            if is_enterprise_standard_gpo(gpo_name):
                ent_settings.append(setting)
            else:
                # Method 1: Check if matches operation prefix by NAME
                matches_by_name = False
                for term in search_terms:
                    if len(term) < 2:
                        continue
                    if len(term) <= 3:
                        # Short term - require word boundary
                        pattern = r'(?:^|[\s\-_])' + re.escape(term) + r'(?:$|[\s\-_])'
                        if re.search(pattern, gpo_name_lower):
                            matches_by_name = True
                            break
                    else:
                        if term in gpo_name_lower:
                            matches_by_name = True
                            break
                
                # Method 2: Check if in link-matched GPOs
                matches_by_link = gpo_name in link_matched_gpos
                
                # Include if EITHER method matches
                if matches_by_name or matches_by_link:
                    operation_settings.append(setting)
        
        if not operation_settings:
            return {
                "domain": operation_code,
                "total_settings": 0,
                "migrate": 0,
                "dont_migrate": 0,
                "review": 0,
                "total_gpos": 0,
                "p1_count": 0,
                "p2_count": 0,
                "p3_count": 0,
                "p4_count": 0,
                "gpo_summary": [],
                "warning": f"No GPOs found matching operation prefixes: {', '.join(search_terms)}"
            }
        
        # Initialize comparison engine and load filtered settings
        comparison_engine = MigrationComparisonEngine()
        comparison_engine.load_operation_settings(operation_settings)
        comparison_engine.load_enterprise_settings(ent_settings)
        
        # Build source lookup from active_gpos (baseline.corp vs baseline.corp)
        gpo_source_lookup = {}
        if self.active_gpos is not None and not self.active_gpos.empty:
            ent_gpos_df = self.active_gpos[
                self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
            ]
            for _, row in ent_gpos_df.iterrows():
                gpo_name = row.get('DisplayName', '')
                domain_name = row.get('DomainName', 'baseline.corp')
                if gpo_name:
                    gpo_source_lookup[gpo_name] = domain_name

        # Build set of ALL operation GPOs (name-matched + link-matched) to include those with 0 settings
        # These are Active GPOs matching this operation that should appear even if they have no extractable settings
        all_operation_gpos = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            ent_gpos_df = self.active_gpos[
                self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
            ]
            for _, row in ent_gpos_df.iterrows():
                gpo_name = row.get('DisplayName', '')
                gpo_name_lower = gpo_name.lower()

                # Skip ENT GPOs
                if is_enterprise_standard_gpo(gpo_name):
                    continue

                # Check name matching
                matches_by_name = False
                for term in search_terms:
                    if len(term) < 2:
                        continue
                    if len(term) <= 3:
                        pattern = r'(?:^|[\s\-_])' + re.escape(term) + r'(?:$|[\s\-_])'
                        if re.search(pattern, gpo_name_lower):
                            matches_by_name = True
                            break
                    else:
                        if term in gpo_name_lower:
                            matches_by_name = True
                            break

                # Check link matching
                matches_by_link = gpo_name in link_matched_gpos

                if matches_by_name or matches_by_link:
                    all_operation_gpos.add(gpo_name)

        # Compare, classify, and build response with source lookup
        return self._build_migration_response(comparison_engine, operation_code, gpo_source_lookup, all_operation_gpos)
    
    def _build_migration_response(self, comparison_engine: 'MigrationComparisonEngine', domain: str, gpo_source_lookup: Dict[str, str] = None, all_active_gpos: set = None) -> Dict[str, Any]:
        """
        Build the migration analysis response from a comparison engine.
        Shared logic for both standalone and Shared Forest analysis.

        Args:
            comparison_engine: The comparison engine with loaded settings
            domain: The domain/operation being analyzed
            gpo_source_lookup: Optional dict mapping GPO name -> source domain
            all_active_gpos: Optional set of ALL active GPO names to include (even those with 0 settings)
        """
        # Compare and classify
        classifications = comparison_engine.compare_and_classify()
        summary = comparison_engine.get_summary()
        
        # Get DataFrame for GPO-level analysis
        df = comparison_engine.to_dataframe()
        
        # Calculate GPO-level summary
        total_gpos = 0
        p1_count = p2_count = p3_count = p4_count = 0
        
        if not df.empty:
            gpo_groups = df.groupby('Operation_GPO')
            total_gpos = len(gpo_groups)
            
            for gpo_name, gpo_df in gpo_groups:
                migrate_count = (gpo_df['Classification'] == 'MIGRATE').sum()
                review_count = (gpo_df['Classification'] == 'REVIEW').sum()
                
                if migrate_count > 0 and review_count == 0:
                    p1_count += 1
                elif migrate_count > 0 and review_count > 0:
                    p2_count += 1
                elif review_count > 0 and migrate_count == 0:
                    p3_count += 1
                else:
                    p4_count += 1
        
        # Build GPO summary data for web display
        gpo_summary_rows = []
        if not df.empty:
            # Use same logic as Excel report for consistency
            gpo_agg = df.groupby('Operation_GPO').agg({
                'Classification': 'count',
                'Setting_Key': lambda x: (df.loc[x.index, 'Classification'] == 'MIGRATE').sum(),
            }).rename(columns={'Classification': 'Total_Settings', 'Setting_Key': 'MIGRATE'})
            
            # Add DONT_MIGRATE and REVIEW counts
            gpo_agg['DONT_MIGRATE'] = df.groupby('Operation_GPO').apply(
                lambda x: (x['Classification'] == 'DONT_MIGRATE').sum()
            )
            gpo_agg['REVIEW'] = df.groupby('Operation_GPO').apply(
                lambda x: (x['Classification'] == 'REVIEW').sum()
            )
            
            # Build Enterprise Standard overlap using same is_enterprise_standard_gpo helper
            def get_ent_overlap(gpo_name):
                gpo_settings = df[df['Operation_GPO'] == gpo_name]
                overlap_settings = gpo_settings[gpo_settings['Classification'].isin(['DONT_MIGRATE', 'REVIEW'])]
                if overlap_settings.empty:
                    return ''
                ent_gpos = overlap_settings['Enterprise_GPO'].dropna().unique()
                ent_gpos = [g for g in ent_gpos if g and str(g).strip() and is_enterprise_standard_gpo(str(g))]
                return ', '.join(sorted(set(ent_gpos))) if ent_gpos else ''
            
            # Build Shared Forest operation overlap
            def get_shared_forest_overlap(gpo_name):
                gpo_settings = df[df['Operation_GPO'] == gpo_name]
                if 'Shared_Forest_Overlap_GPO' not in gpo_settings.columns:
                    return ''
                overlap_gpos = gpo_settings['Shared_Forest_Overlap_GPO'].dropna().unique()
                op_gpos = [g for g in overlap_gpos if g and str(g).strip()]
                return ', '.join(sorted(set(op_gpos))) if op_gpos else ''
            
            # Determine action and priority
            def get_action(row):
                if row['MIGRATE'] > 0 and row['DONT_MIGRATE'] == 0 and row['REVIEW'] == 0:
                    return 'MIGRATE GPO'
                elif row['MIGRATE'] == 0 and row['REVIEW'] == 0:
                    return 'SKIP (enterprise covers)'
                elif row['MIGRATE'] > 0 and row['REVIEW'] > 0:
                    return 'MIGRATE + REVIEW conflicts'
                elif row['MIGRATE'] > 0:
                    return 'MIGRATE (partial overlap)'
                elif row['REVIEW'] > 0:
                    return 'REVIEW conflicts only'
                else:
                    return 'REVIEW'
            
            def get_priority(action):
                if action in ['MIGRATE GPO', 'MIGRATE (partial overlap)']:
                    return 'P1'
                elif action == 'MIGRATE + REVIEW conflicts':
                    return 'P2'
                elif action == 'REVIEW conflicts only':
                    return 'P3'
                else:
                    return 'P4'
            
            # Build rows
            for gpo_name in gpo_agg.index:
                row = gpo_agg.loc[gpo_name]
                action = get_action(row)
                priority = get_priority(action)

                # Get source domain from lookup, default to the analysis domain
                source = gpo_source_lookup.get(gpo_name, domain) if gpo_source_lookup else domain

                # C-Enhanced: ENT GPOs show SKIP with drift detection
                # ENT GPOs in standalone domains are enterprise-managed, not operation decisions
                if is_enterprise_standard_gpo(gpo_name):
                    # Check for drift: if GPO has MIGRATE or REVIEW settings, it differs from baseline
                    has_drift = row['MIGRATE'] > 0 or row['REVIEW'] > 0
                    if has_drift:
                        action = 'SKIP (enterprise managed - DRIFT DETECTED)'
                        priority = 'P3'  # Flag for review due to drift
                    else:
                        action = 'SKIP (enterprise managed)'
                        priority = 'P4'  # No action needed

                gpo_summary_rows.append({
                    'priority': priority,
                    'action': action,
                    'operation_gpo': gpo_name,
                    'source_domain': source,
                    'ent_standard_overlap': get_ent_overlap(gpo_name),
                    'shared_forest_operation_overlap': get_shared_forest_overlap(gpo_name),
                    'total_settings': int(row['Total_Settings']),
                    'migrate': int(row['MIGRATE']),
                    'dont_migrate': int(row['DONT_MIGRATE']),
                    'review': int(row['REVIEW'])
                })
            
            # Sort by priority then migrate count descending
            priority_order = {'P1': 1, 'P2': 2, 'P3': 3, 'P4': 4}
            gpo_summary_rows.sort(key=lambda x: (priority_order.get(x['priority'], 5), -x['migrate']))

        # Add GPOs with zero settings (links-only or empty GPOs)
        # These are Active GPOs that exist but have no extractable settings
        if all_active_gpos:
            gpos_with_settings = set(row['operation_gpo'] for row in gpo_summary_rows)
            zero_setting_gpos = all_active_gpos - gpos_with_settings

            for gpo_name in sorted(zero_setting_gpos):
                source = gpo_source_lookup.get(gpo_name, domain) if gpo_source_lookup else domain

                # C-Enhanced: ENT GPOs with zero settings also show SKIP
                if is_enterprise_standard_gpo(gpo_name):
                    action = 'SKIP (enterprise managed)'
                    priority = 'P4'
                else:
                    action = 'NO SETTINGS (links only)'
                    priority = 'P4'

                gpo_summary_rows.append({
                    'priority': priority,
                    'action': action,
                    'operation_gpo': gpo_name,
                    'source_domain': source,
                    'ent_standard_overlap': '',
                    'shared_forest_operation_overlap': '',
                    'total_settings': 0,
                    'migrate': 0,
                    'dont_migrate': 0,
                    'review': 0
                })
                p4_count += 1
                total_gpos += 1

            # Re-sort after adding zero-setting GPOs
            priority_order = {'P1': 1, 'P2': 2, 'P3': 3, 'P4': 4}
            gpo_summary_rows.sort(key=lambda x: (priority_order.get(x['priority'], 5), -x['migrate']))

        # Calculate total from individual counts (get_summary doesn't include 'total' key)
        total_settings = summary.get('MIGRATE', 0) + summary.get('DONT_MIGRATE', 0) + summary.get('REVIEW', 0)
        
        return {
            "domain": domain,
            "total_settings": total_settings,
            "migrate": summary.get('MIGRATE', 0),
            "dont_migrate": summary.get('DONT_MIGRATE', 0),
            "review": summary.get('REVIEW', 0),
            "total_gpos": total_gpos,
            "p1_count": p1_count,
            "p2_count": p2_count,
            "p3_count": p3_count,
            "p4_count": p4_count,
            "gpo_summary": gpo_summary_rows,
            "_dataframe": df if gpo_source_lookup else None  # Internal: full dataframe for Excel export
        }

    def generate_migration_excel(self, domain: str, output_path: Path) -> Dict[str, Any]:
        """
        Generate Excel report using the SAME analysis as run_migration_analysis.
        
        This ensures Excel and web display show IDENTICAL data.
        Uses the web analysis code path which includes:
        - Active GPO filtering
        - Hybrid shared_prefixes merging
        - Per-GPO source tracking
        """
        import pandas as pd
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # Run the same analysis as web - this ensures consistency
        result = self.run_migration_analysis(domain)
        
        if "error" in result:
            return result
        
        # Get the internal dataframe
        df = result.pop('_dataframe', None)
        if df is None or df.empty:
            return {"error": "No data available for export"}
        
        # Generate Excel with same format as CLI but using consistent web data
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Tab 1: Migration Summary
            summary_data = {
                'Category': [
                    'SETTINGS ANALYSIS', '', '', '', '',
                    'GPO ANALYSIS', '', '', '', ''
                ],
                'Classification': [
                    'MIGRATE', "DON'T MIGRATE", 'REVIEW', 'TOTAL SETTINGS', '',
                    'P1 - Ready to Migrate', 'P2 - Migrate + Resolve Conflicts', 
                    'P3 - Audit Conflicts Only', 'P4 - Skip (Enterprise Covers)', 'TOTAL GPOs'
                ],
                'Count': [
                    result['migrate'], result['dont_migrate'], result['review'], 
                    result['total_settings'], '',
                    result['p1_count'], result['p2_count'], result['p3_count'], 
                    result['p4_count'], result['total_gpos']
                ],
                'Percentage': [
                    f"{result['migrate']/result['total_settings']*100:.1f}%" if result['total_settings'] > 0 else "0%",
                    f"{result['dont_migrate']/result['total_settings']*100:.1f}%" if result['total_settings'] > 0 else "0%",
                    f"{result['review']/result['total_settings']*100:.1f}%" if result['total_settings'] > 0 else "0%",
                    '100%', '',
                    '', '', '', '', ''
                ]
            }
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Migration_Summary', index=False)
            
            # Tab 2: GPO Summary - from web analysis (includes source_domain)
            gpo_summary_df = pd.DataFrame(result['gpo_summary'])
            if not gpo_summary_df.empty:
                # Rename columns to match expected Excel format
                column_map = {
                    'priority': 'Priority',
                    'action': 'Action', 
                    'operation_gpo': 'Operation_GPO',
                    'source_domain': 'Source_Domain',
                    'dn_path': 'DN_Path',
                    'ent_standard_overlap': 'ENT_Standard_Overlap',
                    'shared_forest_operation_overlap': 'Shared_Forest_Operation_Overlap',
                    'total_settings': 'Total_Settings',
                    'migrate': 'MIGRATE',
                    'dont_migrate': 'DONT_MIGRATE',
                    'review': 'REVIEW'
                }
                gpo_summary_df = gpo_summary_df.rename(columns=column_map)
                
                # Select and order columns (include Source_Domain and DN_Path if present)
                base_cols = ['Priority', 'Action', 'Operation_GPO']
                if 'Source_Domain' in gpo_summary_df.columns:
                    base_cols.append('Source_Domain')
                if 'DN_Path' in gpo_summary_df.columns:
                    base_cols.append('DN_Path')
                base_cols.extend(['ENT_Standard_Overlap', 'Shared_Forest_Operation_Overlap', 
                                  'Total_Settings', 'MIGRATE', 'DONT_MIGRATE', 'REVIEW'])
                
                # Only include columns that exist
                cols = [c for c in base_cols if c in gpo_summary_df.columns]
                gpo_summary_df = gpo_summary_df[cols]
                gpo_summary_df = gpo_summary_df.sort_values(['Priority', 'MIGRATE'], ascending=[True, False])
                gpo_summary_df.to_excel(writer, sheet_name='GPO_Summary', index=False)
            
            # Tab 3: Settings Analysis - detailed view
            if not df.empty:
                cols_to_export = ['Classification', 'Reason', 'Operation_GPO', 'Enterprise_GPO',
                                  'Category', 'Setting_Key', 'Setting_Name',
                                  'Operation_Value', 'Enterprise_Value']
                # Only include columns that exist
                cols_available = [c for c in cols_to_export if c in df.columns]
                df_export = df[cols_available].copy()
                df_export = df_export.sort_values(['Classification', 'Operation_GPO'])
                df_export.to_excel(writer, sheet_name='Settings_Analysis', index=False)
            
            # Tab 4: Review Required - just REVIEW items
            review_df = df[df['Classification'] == 'REVIEW'].copy() if not df.empty else pd.DataFrame()
            if not review_df.empty:
                cols_to_export = ['Operation_GPO', 'Enterprise_GPO', 'Category', 
                                  'Setting_Key', 'Setting_Name',
                                  'Operation_Value', 'Enterprise_Value', 'Reason']
                cols_available = [c for c in cols_to_export if c in review_df.columns]
                review_df = review_df[cols_available]
                review_df.to_excel(writer, sheet_name='Review_Required', index=False)
        
        return {
            "success": True,
            "output_path": str(output_path),
            "total_gpos": result['total_gpos'],
            "total_settings": result['total_settings']
        }

    def generate_impact_excel(self, domain: str, output_path: Path) -> Dict[str, Any]:
        """
        Generate Impact Analysis Excel report for GPO replacement planning.
        
        Shows what happens when operation GPOs are replaced with Enterprise Standard equivalents:
        - RETAINED: Settings that stay the same
        - LOST: Settings that will be lost (CRITICAL - must migrate)
        - CHANGED: Settings with different values (CRITICAL - review needed)
        - ADDED: New settings from Enterprise Standard (awareness)
        
        Handles two scenarios:
        1. Standalone domains (charlie.local, etc.) - compare domain HTML vs baseline.corp HTML
        2. Shared Forest operations (OPF, OPG, etc.) - compare operation GPOs vs ENT GPOs within baseline.corp

        Args:
            domain: Domain or operation code to analyze (e.g., 'corp.alpha.com' or 'OPF')
            output_path: Path for output Excel file
            
        Returns:
            Dict with success status and summary info
        """
        import pandas as pd
        
        # Load HTML files
        if not self.html_folder:
            return {"error": "No HTML folder configured"}
        
        html_files = list(self.html_folder.glob("*.html"))
        if not html_files:
            return {"error": f"No HTML files found in {self.html_folder}"}
        
        # Check if this is an Shared Forest operation
        is_shared_forest = domain.upper() in SHARED_FOREST_OPERATIONS
        
        if is_shared_forest:
            # Shared Forest operation: extract operation vs Enterprise Standard settings from baseline.corp
            result = self._get_shared_forest_impact_settings(domain.upper(), html_files)
        else:
            # Standalone domain: extract from domain HTML vs baseline.corp HTML
            result = self._get_standalone_impact_settings(domain, html_files)
        
        if "error" in result:
            return result
        
        operation_settings = result["operation_settings"]
        ent_settings = result["ent_settings"]
        
        if not operation_settings:
            return {"error": f"No settings extracted from {domain}"}
        
        # Run impact analysis
        impact_engine = ImpactAnalysisEngine()
        impact_engine.load_operation_settings(operation_settings)
        impact_engine.load_ent_settings(ent_settings)
        impact_engine.detect_overlaps()
        analyses = impact_engine.analyze_all()
        
        if not analyses:
            return {"error": "No GPOs found for impact analysis"}
        
        # Generate Excel
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            # Tab 1: Impact Summary with risk assessment legend
            summary_df = impact_engine.get_summary_dataframe()
            
            # Create legend dataframe
            legend_data = {
                'Risk_Assessment_Logic': [
                    'HIGH: Lost > 5 OR Changed > 3 — Significant settings will be lost or changed',
                    'MEDIUM: Lost > 0 OR Changed > 0 — Some settings require action before replacement',
                    'LOW: Lost = 0 AND Changed = 0 — Safe replacement, only gains new settings',
                    '',
                    'Column Definitions:',
                    'Retained = Settings with same key and value (no change)',
                    'Lost = Settings in operation GPO but NOT in Enterprise Standard (CRITICAL - must migrate elsewhere)',
                    'Changed = Settings in both but with DIFFERENT values (CRITICAL - review needed)',
                    'Added = Settings in Enterprise Standard but NOT in operation (new settings you will gain)',
                    ''
                ]
            }
            legend_df = pd.DataFrame(legend_data)
            legend_df.to_excel(writer, sheet_name='Impact_Summary', index=False, startrow=0)
            
            # Add actual summary data below legend
            summary_df.to_excel(writer, sheet_name='Impact_Summary', index=False, startrow=len(legend_df) + 2)
            
            # Tab 2: Settings Retained
            retained_df = impact_engine.get_retained_dataframe()
            if not retained_df.empty:
                retained_df.to_excel(writer, sheet_name='Settings_Retained', index=False)
            
            # Tab 3: Settings Lost (CRITICAL)
            lost_df = impact_engine.get_lost_dataframe()
            if not lost_df.empty:
                lost_df.to_excel(writer, sheet_name='Settings_Lost', index=False)
            
            # Tab 4: Settings Changed (CRITICAL)
            changed_df = impact_engine.get_changed_dataframe()
            if not changed_df.empty:
                changed_df.to_excel(writer, sheet_name='Settings_Changed', index=False)
            
            # Tab 5: Settings Added (AWARENESS)
            added_df = impact_engine.get_added_dataframe()
            if not added_df.empty:
                added_df.to_excel(writer, sheet_name='Settings_Added', index=False)
        
        # Calculate summary stats
        total_gpos = len(analyses)
        high_risk = sum(1 for a in analyses if a.risk_level == 'HIGH')
        medium_risk = sum(1 for a in analyses if a.risk_level == 'MEDIUM')
        low_risk = sum(1 for a in analyses if a.risk_level == 'LOW')
        total_lost = sum(a.lost_count for a in analyses)
        total_changed = sum(a.changed_count for a in analyses)
        
        return {
            "success": True,
            "output_path": str(output_path),
            "total_gpos": total_gpos,
            "high_risk": high_risk,
            "medium_risk": medium_risk,
            "low_risk": low_risk,
            "total_lost": total_lost,
            "total_changed": total_changed
        }

    def _get_standalone_impact_settings(self, domain: str, html_files: list) -> Dict[str, Any]:
        """
        Get operation and Enterprise Standard settings for standalone domain impact analysis.
        Compares domain HTML vs baseline.corp HTML.
        """
        # Find the domain HTML file and enterprise file
        domain_file = None
        enterprise_file = None
        
        for f in html_files:
            fname = f.stem.lower()
            if domain.lower().replace('.', '_') in fname or domain.lower() in fname:
                domain_file = f
            if 'enterprise' in fname or 'baseline' in fname:
                enterprise_file = f
        
        if not domain_file:
            return {"error": f"No HTML file found for domain: {domain}"}
        if not enterprise_file:
            return {"error": "No baseline.corp HTML file found for baseline comparison"}
        
        # Extract settings from both files
        extraction_engine = SettingExtractionEngine()
        
        with open(domain_file, 'r', encoding='utf-8', errors='ignore') as f:
            domain_html = f.read()
        with open(enterprise_file, 'r', encoding='utf-8', errors='ignore') as f:
            enterprise_html = f.read()
        
        operation_settings = extraction_engine.extract_from_html_content(domain_html, domain)
        ent_settings = extraction_engine.extract_from_html_content(enterprise_html, 'baseline.corp')
        
        # Filter Enterprise Standard settings to only include actual ENT GPOs
        ent_settings = [s for s in ent_settings if is_enterprise_standard_gpo(
            s.gpo_name if hasattr(s, 'gpo_name') else s.get('gpo_name', '')
        )]
        
        return {
            "operation_settings": operation_settings,
            "ent_settings": ent_settings
        }

    def _get_shared_forest_impact_settings(self, operation_code: str, html_files: list) -> Dict[str, Any]:
        """
        Get operation and Enterprise Standard settings for Shared Forest operation impact analysis.
        Both operation GPOs and ENT GPOs live in the same baseline.corp HTML file.
        
        Uses TWO detection methods to find operation GPOs:
        1. Name prefix matching (OPF-, OPG-, etc.)
        2. Link-based detection (DetectedOperation from OU structure)
        """
        # Get operation info
        if operation_code not in LOCATION_MAPPING:
            return {"error": f"Unknown operation code: {operation_code}"}
        
        info = LOCATION_MAPPING[operation_code]
        
        # Find baseline.corp HTML file
        enterprise_file = None
        enterprise_domain = None
        
        for html_file in html_files:
            with open(html_file, 'r', encoding='utf-8', errors='ignore') as f:
                content = f.read(50000)
            
            domain_match = re.search(r'"DomainName":\s*"([^"]+)"', content)
            if domain_match:
                detected_domain = domain_match.group(1)
                if 'baseline' in detected_domain.lower() or 'baseline.corp' in detected_domain.lower():
                    enterprise_file = html_file
                    enterprise_domain = detected_domain
                    break
        
        if not enterprise_file:
            return {"error": "No baseline.corp HTML file found"}
        
        # Build search terms for this operation (for name-based matching)
        search_terms = [operation_code.lower(), info['full_name'].lower()]
        if 'name_prefixes' in info:
            search_terms.extend([p.lower() for p in info['name_prefixes']])
        
        # Get GPOs detected by LINK analysis for this operation
        link_matched_gpos = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            if 'DetectedOperation' in self.active_gpos.columns:
                # Get baseline.corp GPOs
                enterprise_gpos = self.active_gpos[
                    self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
                ]
                # Find GPOs where DetectedOperation matches this operation
                for _, row in enterprise_gpos.iterrows():
                    detected_op = str(row.get('DetectedOperation', '')).lower()
                    gpo_name = row.get('DisplayName', '')
                    
                    # Check if detected operation matches any of our search terms
                    for term in search_terms:
                        if len(term) < 2:
                            continue
                        if len(term) <= 3:
                            # Short term - require word boundary
                            pattern = r'(?:^|[\s\-_])' + re.escape(term) + r'(?:$|[\s\-_])'
                            if re.search(pattern, detected_op) or detected_op == term:
                                link_matched_gpos.add(gpo_name)
                                break
                        else:
                            if term in detected_op or detected_op in term:
                                link_matched_gpos.add(gpo_name)
                                break
        
        # Build set of ACTIVE GPO names - only process settings from active GPOs
        active_gpo_names = set()
        if self.active_gpos is not None and not self.active_gpos.empty:
            active_gpo_names = set(self.active_gpos['DisplayName'].dropna().unique())
        
        # Initialize extraction engine
        extraction_engine = SettingExtractionEngine(verbose=False)
        
        # Extract ALL settings from baseline.corp
        with open(enterprise_file, 'r', encoding='utf-8', errors='ignore') as f:
            ent_content = f.read()
        all_settings = extraction_engine.extract_from_html_content(ent_content, enterprise_domain or 'baseline.corp')
        
        # Filter settings into operation vs Enterprise Standard (only from ACTIVE GPOs)
        operation_settings = []
        ent_settings = []
        
        for setting in all_settings:
            gpo_name = setting.gpo_name if hasattr(setting, 'gpo_name') else setting.get('gpo_name', '')
            gpo_name_lower = gpo_name.lower()
            
            # CRITICAL: Skip settings from disabled/unlinked GPOs
            if gpo_name not in active_gpo_names:
                continue
            
            # Check if Enterprise Standard GPO
            if is_enterprise_standard_gpo(gpo_name):
                ent_settings.append(setting)
            else:
                # Method 1: Check if matches operation prefix by NAME
                matches_by_name = False
                for term in search_terms:
                    if len(term) < 2:
                        continue
                    if len(term) <= 3:
                        # Short term - require word boundary
                        pattern = r'(?:^|[\s\-_])' + re.escape(term) + r'(?:$|[\s\-_])'
                        if re.search(pattern, gpo_name_lower):
                            matches_by_name = True
                            break
                    else:
                        if term in gpo_name_lower:
                            matches_by_name = True
                            break
                
                # Method 2: Check if in link-matched GPOs
                matches_by_link = gpo_name in link_matched_gpos
                
                # Include if EITHER method matches
                if matches_by_name or matches_by_link:
                    operation_settings.append(setting)
        
        if not operation_settings:
            return {
                "error": f"No GPOs found matching operation '{operation_code}'. "
                         f"Search terms: {', '.join(search_terms)}"
            }
        
        return {
            "operation_settings": operation_settings,
            "ent_settings": ent_settings
        }

    def _web_get_operation_gpos(self, operation_code: str) -> Tuple[pd.DataFrame, Dict]:
        """Get GPOs for a specific operation."""
        if operation_code.upper() in LOCATION_MAPPING:
            info = LOCATION_MAPPING[operation_code.upper()]
            code = operation_code.upper()
        else:
            info = {'full_name': operation_code, 'source_domain': operation_code}
            code = operation_code
        
        source_domain = info.get('source_domain', operation_code)
        
        if self.active_gpos is None or self.active_gpos.empty:
            return pd.DataFrame(), info
        
        if source_domain == 'baseline.corp':
            enterprise_gpos = self.active_gpos[
                self.active_gpos['Domain'].str.lower().isin(['baseline.corp', 'baseline.corp'])
            ].copy()
            gpos = self._web_filter_gpos_for_operation(enterprise_gpos, code, info)
        else:
            domain_mask = self.active_gpos['Domain'].str.lower() == source_domain.lower()
            gpos = self.active_gpos[domain_mask].copy()
            if not gpos.empty:
                ent_mask = gpos['DisplayName'].str.upper().str.contains('ENT', na=False)
                gpos = gpos[~ent_mask]
        
        return gpos, info

    def _web_filter_gpos_for_operation(
        self, 
        enterprise_gpos: pd.DataFrame, 
        code: str, 
        info: Dict
    ) -> pd.DataFrame:
        """Filter enterprise GPOs for a specific operation."""
        if enterprise_gpos.empty:
            return enterprise_gpos
        
        search_terms = [code.lower(), info['full_name'].lower()]
        if 'name_prefixes' in info:
            search_terms.extend([p.lower() for p in info['name_prefixes']])
        
        # Match by DisplayName
        name_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
        
        for term in search_terms:
            if len(term) < 2:
                continue
            
            if len(term) <= 3:
                pattern = r'(?:^|[\s\-_])' + re.escape(term) + r'(?:$|[\s\-_])'
                term_match = enterprise_gpos['DisplayName'].str.contains(
                    pattern, case=False, na=False, regex=True
                )
            else:
                term_match = enterprise_gpos['DisplayName'].str.contains(
                    term, case=False, na=False, regex=False
                )
            name_mask = name_mask | term_match
        
        # Match by Links column (GPOs linked to operation OUs but not named for operation)
        links_mask = pd.Series([False] * len(enterprise_gpos), index=enterprise_gpos.index)
        
        if 'Links' in enterprise_gpos.columns:
            for term in search_terms:
                if len(term) >= 4:  # Only 4+ char terms to avoid false positives
                    term_in_links = enterprise_gpos['Links'].str.contains(
                        term, case=False, na=False, regex=False
                    )
                    links_mask = links_mask | term_in_links
        
        # Links-only = linked to operation OUs but not named for operation
        links_only_mask = links_mask & ~name_mask
        
        # Combine both
        combined_mask = name_mask | links_only_mask
        matched = enterprise_gpos[combined_mask].copy()
        
        # Filter out ENT GPOs
        if not matched.empty:
            ent_mask = matched['DisplayName'].str.upper().str.contains('ENT', na=False)
            matched = matched[~ent_mask]
        
        # Set MatchType
        if not matched.empty:
            matched['MatchType'] = 'Name Match'
            links_only_indices = enterprise_gpos[links_only_mask].index
            matched.loc[matched.index.isin(links_only_indices), 'MatchType'] = 'Links Only - Review Required'
        
        return matched

    def _web_format_links(self, links_str: str) -> str:
        """Format links for display."""
        if not links_str or pd.isna(links_str):
            return ""
        
        links_str = str(links_str).strip()
        if not links_str:
            return ""
        
        dn_pattern = r'(?:OU|CN)=[^,]+(?:,(?:OU|CN)=[^,]+)*,DC=[^,]+(?:,DC=[^,]+)+'
        dn_matches = re.findall(dn_pattern, links_str, re.IGNORECASE)
        
        if not dn_matches:
            dn_matches = [links_str] if 'DC=' in links_str else []
        
        formatted_links = []
        for dn_path in dn_matches:
            parts = re.findall(r'(?:OU|CN)=([^,]+)', dn_path, re.IGNORECASE)
            
            if parts:
                ou_parts = list(reversed(parts))
                formatted = ' > '.join(ou_parts)
                formatted_links.append(formatted)
            else:
                dc_parts = re.findall(r'DC=([^,]+)', dn_path, re.IGNORECASE)
                if dc_parts:
                    formatted_links.append(f"[Domain Root: {'.'.join(dc_parts)}]")
        
        return ';\n'.join(formatted_links) + (';' if formatted_links else '')

    def _web_get_review_action(self, bucket: str, match_type: str, row: pd.Series) -> str:
        """Determine the action for Review Required tab."""
        if match_type == 'Links Only - Review Required':
            return "REVIEW: GPO linked to operation OU but name doesn't contain operation code."
        elif bucket == 'Domain Root':
            return "DON'T MIGRATE: Linked at domain root level - applies domain-wide."
        elif bucket == 'Domain Controller':
            return "DON'T MIGRATE: Domain Controller policy - critical infrastructure."
        elif bucket == 'Mixed':
            bucket_details = row.get('BucketDetails', '')
            try:
                details_dict = eval(bucket_details) if bucket_details else {}
                parts = [f"{v} {k}" for k, v in sorted(details_dict.items(), key=lambda x: -x[1])]
                breakdown = ", ".join(parts)
                return f"REVIEW: Mixed targeting ({breakdown}). Verify settings are appropriate."
            except:
                return "REVIEW: Multi-target GPO. Verify settings are appropriate."
        elif bucket == 'Unknown':
            return "REVIEW: Unable to classify - manually verify the target OU type."
        else:
            return "REVIEW: Manual verification required."

    def _web_generate_detection_reason(self, bucket: str, row: pd.Series) -> str:
        """Generate explanation for bucket assignment."""
        links = str(row.get('Links', ''))
        
        ou_names = []
        if links:
            ou_matches = re.findall(r'OU=([^,]+)', links)
            seen = set()
            for ou in ou_matches:
                if ou not in seen:
                    ou_names.append(ou)
                    seen.add(ou)
                if len(ou_names) >= 3:
                    break
        
        ou_display = ', '.join(ou_names) if ou_names else 'unknown'
        
        if bucket == 'Domain Root':
            return "Linked at domain root level (DC=)"
        elif bucket == 'Domain Controller':
            return "Linked to Domain Controllers OU"
        elif bucket == 'Server':
            return f"Detected 'Server' keyword in OU path: {ou_display}"
        elif bucket == 'Workstation':
            keywords_found = []
            for kw in ['Computer', 'Workstation', 'Desktop', 'Laptop', 'Client', 'PC', 'Kiosk']:
                if kw.lower() in links.lower():
                    keywords_found.append(kw)
            if keywords_found:
                return f"Detected '{keywords_found[0]}' keyword in OU: {ou_display}"
            return f"Detected workstation keyword in OU: {ou_display}"
        elif bucket == 'User':
            return f"Detected 'User' keyword in OU: {ou_display}"
        elif bucket == 'Mixed':
            bucket_details = row.get('BucketDetails', '')
            if bucket_details:
                try:
                    details_dict = eval(bucket_details) if isinstance(bucket_details, str) else bucket_details
                    parts = [f"{v} {k}" for k, v in sorted(details_dict.items(), key=lambda x: -x[1])]
                    return f"Multiple bucket types: {', '.join(parts)}"
                except:
                    pass
            return "Linked to multiple OU types"
        elif bucket == 'Unknown':
            if ou_names:
                return f"No recognized keywords in OUs: {ou_display}"
            return "Unable to determine OU type from links"
        
        return f"Bucket: {bucket}"

    def _web_get_readiness(self, row: pd.Series) -> str:
        """Determine migration readiness."""
        bucket = row.get('Bucket', '')
        applies_to = row.get('AppliesTo', 'Unknown')
        
        if bucket in ['Server', 'Workstation']:
            if bucket == 'Server' and applies_to == 'User':
                return 'Review First'
            return 'Ready'
        elif bucket == 'User':
            if applies_to in ['Computer', 'Both']:
                return 'Consider Splitting'
            return 'Ready'
        elif bucket in ['Mixed', 'Unknown']:
            return 'Review First'
        elif bucket in ['Domain Controller', 'Domain Root']:
            return 'Not Applicable'
        else:
            return 'Review First'

    def _web_get_optimization_note(self, row: pd.Series) -> str:
        """Generate optimization note for User GPOs."""
        applies_to = row.get('AppliesTo', 'Unknown')
        
        if applies_to == 'Computer':
            return 'INEFFICIENT: Computer-only settings in User OU have no effect.'
        elif applies_to == 'Both':
            return 'OPTIMIZATION: Computer settings are wasted in User OU.'
        else:
            return 'User-only settings OK.'

    def _web_get_settings_count(self, row: pd.Series) -> int:
        """Get settings count from row."""
        for col in ['SettingsCount', 'Settings', 'LinksCount']:
            if col in row and pd.notna(row[col]):
                try:
                    return int(row[col])
                except (ValueError, TypeError):
                    pass
        return 0

    def _web_detect_gpo_issues(self, gpo_row: pd.Series, settings: List[Dict]) -> List[Dict[str, Any]]:
        """Detect issues/warnings for a GPO."""
        issues = []
        
        bucket = gpo_row.get('Bucket', '')
        applies_to = gpo_row.get('AppliesTo', '')
        match_type = gpo_row.get('MatchType', '')
        
        # Check for WMI filter
        wmi_filter = gpo_row.get('WMIFilter', '')
        if pd.notna(wmi_filter) and str(wmi_filter).strip():
            issues.append({
                "severity": "warning",
                "message": "Has WMI Filter",
                "detail": f"WMI Filter: {wmi_filter}. May need recreation in target domain."
            })
        
        # Check for empty GPO
        settings_count = self._web_get_settings_count(gpo_row)
        if settings_count == 0 and len(settings) == 0:
            issues.append({
                "severity": "warning",
                "message": "Empty GPO",
                "detail": "This GPO has no configured settings."
            })
        
        # Check for bucket/applies_to mismatch
        if bucket == 'Server' and applies_to == 'User':
            issues.append({
                "severity": "info",
                "message": "User settings in Server OU",
                "detail": "User settings will only apply to users who log into these servers."
            })
        
        if bucket == 'User' and applies_to == 'Computer':
            issues.append({
                "severity": "error",
                "message": "Ineffective Configuration",
                "detail": "Computer settings have NO effect on user objects."
            })
        
        if match_type == 'Links Only - Review Required':
            issues.append({
                "severity": "warning",
                "message": "Links-Only Detection",
                "detail": "GPO detected by link location, not by name. Verify ownership."
            })
        
        if bucket == 'Mixed':
            issues.append({
                "severity": "info",
                "message": "Multi-target GPO",
                "detail": "This GPO applies to multiple OU types."
            })
        
        return issues


def get_download_url(output_path: Path) -> str:
    """
    Generate download URL if running in container with downloads folder.
    
    Returns clickable URL if:
    - Output path is in /app/data/downloads/
    
    Returns empty string otherwise (local development, different path, etc.)
    """
    downloads_folder = Path("/app/data/downloads")
    
    # Check if downloads folder exists (indicates container environment)
    if not downloads_folder.exists():
        return ""
    
    # Check if output is in downloads folder
    try:
        # Resolve both paths to handle any symlinks/relative paths
        output_resolved = output_path.resolve()
        downloads_resolved = downloads_folder.resolve()
        
        # Check if output is under downloads folder
        output_resolved.relative_to(downloads_resolved)
    except (ValueError, RuntimeError):
        return ""  # Not in downloads folder
    
    # Generate URL (port 9845 is the frontend nginx that proxies to backend)
    filename = output_path.name
    return f"http://localhost:9845/api/downloads/{filename}"


def main():
    parser = argparse.ArgumentParser(
        description='GPO Cross-Domain Consolidation Analyzer v2.3.2',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example usage:
  Executive mode:
    python gpo_analyzer.py --mode executive --output "exec.xlsx"
  
  Domain mode (regular domain):
    python gpo_analyzer.py --mode domain --domain "echo.corp.com" --output "echo.xlsx"

  Domain mode (operation - use LOCATION_MAPPING codes):
    python gpo_analyzer.py --mode domain --operation "OPF" --output "foxtrot.xlsx"
    python gpo_analyzer.py --mode domain --operation "OPG" --output "golf.xlsx"
    python gpo_analyzer.py --mode domain --operation "OPH" --output "hotel.xlsx"

  Migration mode (compare domain to enterprise baseline):
    python gpo_analyzer.py --mode migration --domain "corp.bravo.com" --output "bravo_migration.xlsx"
    
    Output tabs:
    - Migration_Summary: Classification counts
    - GPO_Summary: Per-GPO priority (P1-P4) and action
    - Settings_Analysis: All settings with classification
    - Review_Required: Conflicts needing review

  Impact mode (GPO replacement impact analysis):
    python gpo_analyzer.py --mode impact --domain "corp.alpha.com" --output "alpha_impact.xlsx"
    
    Output tabs:
    - Impact_Summary: Per-GPO risk assessment (HIGH/MEDIUM/LOW)
    - Settings_Retained: Settings that stay the same
    - Settings_Lost: CRITICAL - settings that will be lost
    - Settings_Changed: CRITICAL - settings with different values
    - Settings_Added: New settings from Enterprise Standard (awareness)

Container usage (save to downloads folder for clickable download link):
    python gpo_analyzer.py --mode executive --output data/downloads/exec.xlsx
    
    Output includes clickable URL: http://localhost:9845/api/downloads/exec.xlsx
    Files auto-delete after 60 minutes.
        """
    )

    parser.add_argument(
        '--mode',
        type=str,
        choices=['executive', 'domain', 'full', 'migration', 'impact'],
        default='full',
        help='Report mode: executive (7 tabs), domain (5 tabs), full (19 tabs), migration (4 tabs), or impact (5 tabs - GPO replacement analysis)'
    )
    
    parser.add_argument(
        '--domain',
        type=str,
        help='Target domain for domain mode (e.g., echo.corp.com)'
    )
    
    parser.add_argument(
        '--operation',
        type=str,
        help='Target operation for domain mode (use codes: OPF, OPG, OPH, OPI, OPJ, etc.)'
    )
    
    parser.add_argument(
        '--enterprise',
        type=str,
        help='Enterprise baseline domain for migration mode (default: auto-detect from shared forest or baseline.corp)'
    )

    parser.add_argument(
        '--html-folder',
        type=Path,
        required=True,
        help='Folder containing GPOZaurr HTML reports (one per domain)'
    )

    parser.add_argument(
        '--output',
        type=Path,
        default=Path('GPO_Analysis_v2.3.2.xlsx'),
        help='Output Excel file path'
    )

    parser.add_argument(
        '--log-file',
        type=Path,
        help='Optional log file path'
    )

    parser.add_argument(
        '--debug',
        action='store_true',
        help='Enable debug logging'
    )

    args = parser.parse_args()

    # SMART DETECTION: If user passed an operation code to --domain, auto-correct
    # This prevents confusion between --domain and --operation flags
    if args.domain and not args.operation:
        # Check if what they passed to --domain is actually an operation code
        if args.domain.upper() in LOCATION_MAPPING:
            mapping = LOCATION_MAPPING[args.domain.upper()]
            if mapping.get('source_domain') == 'baseline.corp':
                # They meant to use --operation, auto-correct
                print(f"\n*** AUTO-CORRECTION: '{args.domain}' is an operation code, not a domain name.")
                print(f"    Switching from --domain to --operation mode.")
                print(f"    (Use --domain for standalone domains like 'echo.corp.com')")
                print(f"    (Use --operation for baseline.corp operations like 'OPF', 'OPG', 'OPH')\n")
                args.operation = args.domain.upper()
                args.domain = None

    # SMART DETECTION: If user passed a domain name to --operation, auto-correct
    if args.operation and not args.domain:
        # Check if it looks like a domain name (contains a dot)
        if '.' in args.operation:
            print(f"\n*** AUTO-CORRECTION: '{args.operation}' looks like a domain name, not an operation code.")
            print(f"    Switching from --operation to --domain mode.")
            print(f"    (Use --operation for baseline.corp operations like 'OPF', 'OPG', 'OPH')")
            print(f"    (Use --domain for standalone domains like 'echo.corp.com')\n")
            args.domain = args.operation
            args.operation = None

    # Setup logging
    log_level = logging.DEBUG if args.debug else logging.INFO
    log_file = args.log_file or args.output.parent / 'gpo_analyzer.log'

    logging.basicConfig(
        level=log_level,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file, mode='w'),
            logging.StreamHandler(sys.stdout)
        ]
    )

    logging.info("=" * 60)
    logging.info("GPO CROSS-DOMAIN CONSOLIDATION ANALYZER v2.3.2")
    logging.info("=" * 60)

    # Validate input folder
    if not args.html_folder.exists():
        logging.error(f"Folder not found: {args.html_folder}")
        sys.exit(1)

    if not args.html_folder.is_dir():
        logging.error(f"Not a directory: {args.html_folder}")
        sys.exit(1)

    print("=" * 60)
    print("GPO CROSS-DOMAIN CONSOLIDATION ANALYZER v2.3.2")
    print("=" * 60)
    print("NEW: Migration Mode - Setting-level comparison for migrate/don't-migrate decisions")
    print("NEW: Bucket-based analysis (Server/Workstation/User/Domain Root)")
    print("NEW: Dynamic OU detection - use operation names from your OU structure")
    print("=" * 60)

    # MIGRATION MODE - Uses same code path as web for consistency
    if args.mode == 'migration':
        print("\n>>> MIGRATION MODE: Setting-level comparison analysis")
        print("    (Using unified analysis - consistent with web UI)")
        
        # Validate required arguments
        if not args.domain and not args.operation:
            print("ERROR: Migration mode requires --domain or --operation to specify the operation domain")
            sys.exit(1)
        
        # Determine domain/operation to analyze
        # For Shared Forest operations, use the operation code directly
        # For standalone domains, use the domain name
        if args.operation:
            analysis_target = args.operation.upper()
            print(f"    Operation: {analysis_target}")
        else:
            analysis_target = args.domain
            print(f"    Domain: {analysis_target}")
        
        # Initialize analyzer and parse data (same as web)
        print("\n  Initializing analyzer...")
        analyzer = GPOAnalyzer(args.html_folder, args.mode, args.domain, args.operation)
        
        print("  Parsing HTML reports...")
        analyzer.parse_html_reports()
        print(f"    Found {len(analyzer.domains)} domain(s)")
        
        print("  Filtering to active GPOs...")
        analyzer.filter_active_gpos()
        print(f"    Active GPOs: {len(analyzer.active_gpos)}")
        
        # Generate Excel using the SAME method as web export
        # This ensures CLI and web produce identical results
        print(f"\n  Running migration analysis for: {analysis_target}")
        result = analyzer.generate_migration_excel(analysis_target, args.output)
        
        if "error" in result:
            print(f"\nERROR: {result['error']}")
            sys.exit(1)
        
        # Print summary
        print("\n" + "=" * 60)
        print("MIGRATION ANALYSIS COMPLETE!")
        print("=" * 60)
        print(f"\n  Total GPOs analyzed: {result.get('total_gpos', 0)}")
        print(f"  Total settings: {result.get('total_settings', 0)}")
        print(f"\nReport saved to: {args.output.absolute()}")
        download_url = get_download_url(args.output)
        if download_url:
            print(f"✓ Download: {download_url}")
        print(f"Log saved to: {log_file.absolute()}")
        
        logging.info("=" * 60)
        logging.info("Migration analysis complete successfully")
        logging.info("=" * 60)
        return
    
    # IMPACT MODE - GPO replacement impact analysis
    if args.mode == 'impact':
        print("\n>>> IMPACT MODE: GPO replacement impact analysis")
        print("    Shows what happens when you replace operation GPOs with Enterprise Standard equivalents")
        
        # Validate required arguments
        if not args.domain and not args.operation:
            print("ERROR: Impact mode requires --domain or --operation to specify the operation domain")
            sys.exit(1)
        
        # Determine domain/operation to analyze
        if args.operation:
            analysis_target = args.operation.upper()
            print(f"    Operation: {analysis_target}")
        else:
            analysis_target = args.domain
            print(f"    Domain: {analysis_target}")
        
        # Initialize analyzer and parse data
        print("\n  Initializing analyzer...")
        analyzer = GPOAnalyzer(args.html_folder, args.mode, args.domain, args.operation)
        
        print("  Parsing HTML reports...")
        analyzer.parse_html_reports()
        print(f"    Found {len(analyzer.domains)} domain(s)")
        
        print("  Filtering to active GPOs...")
        analyzer.filter_active_gpos()
        print(f"    Active GPOs: {len(analyzer.active_gpos)}")
        
        # Generate Impact Excel
        print(f"\n  Running impact analysis for: {analysis_target}")
        result = analyzer.generate_impact_excel(analysis_target, args.output)
        
        if "error" in result:
            print(f"\nERROR: {result['error']}")
            sys.exit(1)
        
        # Print summary
        print("\n" + "=" * 60)
        print("IMPACT ANALYSIS COMPLETE!")
        print("=" * 60)
        print(f"\n  Total GPOs analyzed: {result.get('total_gpos', 0)}")
        print(f"\n  Risk Assessment:")
        print(f"    HIGH:   {result.get('high_risk', 0)} GPOs (significant changes)")
        print(f"    MEDIUM: {result.get('medium_risk', 0)} GPOs (some action needed)")
        print(f"    LOW:    {result.get('low_risk', 0)} GPOs (safe replacement)")
        print(f"\n  Total settings that will be LOST: {result.get('total_lost', 0)}")
        print(f"  Total settings that will CHANGE: {result.get('total_changed', 0)}")
        print(f"\nReport saved to: {args.output.absolute()}")
        download_url = get_download_url(args.output)
        if download_url:
            print(f"✓ Download: {download_url}")
        print(f"Log saved to: {log_file.absolute()}")
        
        logging.info("=" * 60)
        logging.info("Impact analysis complete successfully")
        logging.info("=" * 60)
        return
    
    # STANDARD MODES: executive, domain, full
    # Initialize analyzer
    analyzer = GPOAnalyzer(args.html_folder, args.mode, args.domain, args.operation)
    logging.info(f"Mode: {args.mode}, Domain: {args.domain}, Operation: {args.operation}")

    # Parse HTML reports
    analyzer.parse_html_reports()
    logging.info(f"Parsed {len(analyzer.domains)} domain(s): {', '.join(analyzer.domains.keys())}")

    # Filter to active GPOs (includes bucket analysis, shared GPO detection, 
    # domain GPO creation, ENT filtering, and shared GPO categorization)
    analyzer.filter_active_gpos()
    
    # Log state after filtering
    logging.info(f"=== POST-FILTER STATE ===")
    logging.info(f"  active_gpos: {len(analyzer.active_gpos)} rows")
    logging.info(f"  shared_gpos: {len(analyzer.shared_gpos)} rows")
    logging.info(f"  domain_gpos: {len(analyzer.domain_gpos)} rows")
    logging.info(f"  operations_gpos: {len(analyzer.operations_gpos)} rows")
    logging.info(f"  eit_standard_gpos: {len(analyzer.enterprise_standard_gpos)} rows")
    logging.info(f"  eit_nonstandard_gpos: {len(analyzer.ent_nonstandard_gpos)} rows")
    if 'Bucket' in analyzer.active_gpos.columns:
        bucket_counts = analyzer.active_gpos['Bucket'].value_counts().to_dict()
        logging.info(f"  Bucket distribution: {bucket_counts}")

    # Analyze settings
    analyzer.analyze_settings_patterns()
    logging.info(f"Settings extracted: {len(analyzer.all_settings)}")
    
    analyzer.analyze_for_decisions()
    logging.info(f"Enterprise standards: {len(analyzer.enterprise_standards)}")

    # Generate report based on mode
    logging.info(f"Generating {args.mode} report...")
    analyzer.generate_excel_report(args.output)

    # Log final state for domain mode
    if args.mode == 'domain':
        logging.info(f"=== DOMAIN MODE FINAL STATE ===")
        logging.info(f"  report_gpos: {len(analyzer.report_gpos)} rows")
        if not analyzer.report_gpos.empty and 'Bucket' in analyzer.report_gpos.columns:
            report_buckets = analyzer.report_gpos['Bucket'].value_counts().to_dict()
            logging.info(f"  Report bucket distribution: {report_buckets}")

    print("\n" + "=" * 60)
    print("ANALYSIS COMPLETE!")
    print("=" * 60)
    print(f"\nReport saved to: {args.output.absolute()}")
    download_url = get_download_url(args.output)
    if download_url:
        print(f"✓ Download: {download_url}")
    print(f"Log saved to: {log_file.absolute()}")
    
    print("\nKey Features:")
    print(f"  - Bucket analysis: GPOs categorized by OU link type")
    print(f"  - Operations detected from OU structure (no hardcoding)")
    
    # Use report_gpos for domain/operation mode (filtered), active_gpos for executive mode (all)
    summary_gpos = analyzer.report_gpos if not analyzer.report_gpos.empty else analyzer.active_gpos
    
    if not summary_gpos.empty and 'Bucket' in summary_gpos.columns:
        bucket_summary = summary_gpos['Bucket'].value_counts()
        print(f"\nBucket Summary (Report GPOs: {len(summary_gpos)}):")
        for bucket, count in bucket_summary.items():
            print(f"    {bucket}: {count}")
    
    logging.info("=" * 60)
    logging.info("Analysis complete successfully")
    logging.info("=" * 60)


if __name__ == '__main__':
    main()
